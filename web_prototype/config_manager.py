# -*- coding: utf-8 -*-
"""
Configuration Manager for Web Configurator
Manages configuration loading, saving, validation, and presets.
"""

import os
import json
import re
import uuid
from datetime import datetime
from typing import Dict, List, Tuple, Optional
import openpyxl


# QA-3 (Opcion B): tipo de agente capaz de un area. FUENTE DE VERDAD = el mapa explicito
# `work_area_equipment` de config.json (editable en la UI). La convencion de nombres
# (regex) queda SOLO como fallback de migracion para configs viejas sin el mapa.
# La MISMA logica vive en event_generator (motor) y fleet-manager.js (indicador C).
VALID_EQUIPMENT = ('GroundOperator', 'Forklift')
_GROUND_AREA_RE = re.compile(r'ground|piso|floor|suelo|terrestre|level[_-]?0|l0', re.I)


def _equipment_by_naming(area: str) -> str:
    """Fallback historico (convencion de nombres) cuando el mapa no define el area."""
    return 'GroundOperator' if _GROUND_AREA_RE.search(str(area)) else 'Forklift'


def _expected_equipment_for_area(area: str, mapping: dict = None) -> str:
    """Tipo requerido por el area: del mapa explicito si esta; si no, por convencion."""
    if mapping:
        t = mapping.get(area)
        if t:
            return t
    return _equipment_by_naming(area)


class WebConfigurationManager:
    """Manages configurations for the web configurator"""
    
    def __init__(self, project_root: str):
        """
        Initialize configuration manager
        
        Args:
            project_root: Root directory of the project
        """
        self.project_root = project_root
        self.config_path = os.path.join(project_root, "config.json")
        self.presets_dir = os.path.join(project_root, "data", "config_presets")
        
        # Create presets directory if it doesn't exist
        os.makedirs(self.presets_dir, exist_ok=True)
        
        print(f"[CONFIG_MANAGER] Initialized with root: {project_root}")
        print(f"[CONFIG_MANAGER] Config path: {self.config_path}")
        print(f"[CONFIG_MANAGER] Presets directory: {self.presets_dir}")
    
    def load_config(self) -> Dict:
        """
        Load current config.json
        
        Returns:
            Configuration dictionary
        """
        try:
            if not os.path.exists(self.config_path):
                print(f"[CONFIG_MANAGER] Config file not found, using defaults")
                return self._get_default_config()
            
            with open(self.config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
            
            print(f"[CONFIG_MANAGER] Loaded config from {self.config_path}")
            return config
            
        except Exception as e:
            print(f"[CONFIG_MANAGER ERROR] Error loading config: {e}")
            return self._get_default_config()
    
    def save_config(self, config: Dict) -> Tuple[bool, List[str]]:
        """
        Save configuration to config.json with validation
        
        Args:
            config: Configuration dictionary
            
        Returns:
            Tuple of (success, error_messages)
        """
        try:
            # Validate first
            is_valid, errors = self.validate_config(config)
            if not is_valid:
                return False, errors

            # MERGE: load existing config and overlay UI keys on top.
            # Contract: NEVER drop keys the UI does not manage (e.g. blocks
            # 'congestion' and 'outbound' used by the simulation engine).
            existing = self._load_existing_for_merge()
            merged = dict(existing)
            merged.update(config)
            preserved = sorted(set(existing.keys()) - set(config.keys()))
            if preserved:
                print(f"[CONFIG_MANAGER] Merge preserved keys not managed by UI: {preserved}")

            # Backup existing config (before overwriting)
            if os.path.exists(self.config_path):
                backup_path = f"{self.config_path}.backup"
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    backup_content = f.read()
                with open(backup_path, 'w', encoding='utf-8') as f:
                    f.write(backup_content)
                print(f"[CONFIG_MANAGER] Created backup: {backup_path}")

            # ATOMIC write: dump to temp file, then os.replace (rename).
            # Nota: json.dump en modo texto puede dejar bytes nulos residuales
            # en FUSE/NTFS (el cluster pre-asignado no se trunca correctamente).
            # Solucion: producir el string con json.dumps, codificar a bytes
            # explicitamente, y escribir en modo binario para garantizar
            # truncado exacto al tamano del contenido.
            json_str = json.dumps(merged, indent=4, ensure_ascii=False)
            json_bytes = json_str.encode('utf-8')
            tmp_path = f"{self.config_path}.tmp"
            with open(tmp_path, 'wb') as f:
                f.write(json_bytes)
                f.flush()
                os.fsync(f.fileno())

            # Verificar que el tmp no tenga bytes nulos (defensa adicional)
            with open(tmp_path, 'rb') as f:
                raw = f.read()
            if b'\x00' in raw:
                print("[CONFIG_MANAGER WARN] Null bytes detected in tmp; stripping (FUSE artifact)")
                raw = raw.replace(b'\x00', b'')
                with open(tmp_path, 'wb') as f:
                    f.write(raw)

            # NOTA: os.replace() en FUSE/WinFsp no trunca el archivo destino
            # (sobrescribe in-place dejando bytes residuales al final).
            # Solucion: escribir directamente a config_path en modo 'wb'
            # que garantiza truncado exacto al tamano del contenido.
            with open(self.config_path, 'wb') as f:
                f.write(raw)
                f.flush()
                os.fsync(f.fileno())

            # Cleanup del tmp
            try:
                os.remove(tmp_path)
            except Exception:
                pass

            # Post-verificacion: confirmar que no hay null bytes en el archivo final
            with open(self.config_path, 'rb') as f:
                final_raw = f.read()
            if b'\x00' in final_raw:
                print("[CONFIG_MANAGER WARN] Null bytes en config.json final - reintentando strip")
                clean = final_raw.replace(b'\x00', b'')
                with open(self.config_path, 'wb') as f:
                    f.write(clean)
                    f.flush()
                    os.fsync(f.fileno())

            print(f"[CONFIG_MANAGER] Saved config to {self.config_path}")
            return True, []

        except Exception as e:
            error_msg = f"Error saving config: {str(e)}"
            print(f"[CONFIG_MANAGER ERROR] {error_msg}")
            return False, [error_msg]

    def _load_existing_for_merge(self) -> Dict:
        """
        Load the current on-disk config to merge with incoming UI config.
        Fallback chain: config.json -> config.json.backup -> {} (with warning).
        Never raises: saving must not fail because the existing file is broken.
        """
        for path in (self.config_path, f"{self.config_path}.backup"):
            try:
                if os.path.exists(path):
                    with open(path, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    if isinstance(data, dict):
                        if path != self.config_path:
                            print(f"[CONFIG_MANAGER WARN] config.json unreadable; merging from backup: {path}")
                        return data
            except Exception as e:
                print(f"[CONFIG_MANAGER WARN] Could not read {path} for merge: {e}")
        print(f"[CONFIG_MANAGER WARN] No existing config readable; saving UI config as-is")
        return {}
    
    def validate_config(self, config: Dict) -> Tuple[bool, List[str]]:
        """
        Validate configuration structure and values
        
        Args:
            config: Configuration dictionary
            
        Returns:
            Tuple of (is_valid, error_messages)
        """
        errors = []
        
        try:
            # Validate total_ordenes
            if 'total_ordenes' not in config or config['total_ordenes'] <= 0:
                errors.append("total_ordenes must be greater than 0")
            
            # Validate distribucion_tipos
            if 'distribucion_tipos' in config:
                dist = config['distribucion_tipos']
                total_pct = 0
                
                for tipo in ['pequeno', 'mediano', 'grande']:
                    if tipo not in dist:
                        errors.append(f"Missing distribution type: {tipo}")
                        continue
                    
                    if 'porcentaje' not in dist[tipo]:
                        errors.append(f"Missing percentage for {tipo}")
                    else:
                        total_pct += dist[tipo]['porcentaje']
                    
                    if 'volumen' not in dist[tipo]:
                        errors.append(f"Missing volume for {tipo}")
                
                if total_pct != 100:
                    errors.append(f"Distribution percentages must sum to 100% (current: {total_pct}%)")
            else:
                errors.append("Missing distribucion_tipos")
            
            # MEJ-3: capacidad_carro ELIMINADA (el motor nunca la leyo; la
            # capacidad real viene de agent_types[].capacity). Ya no se exige.

            # BK-01: Validate radio_cercania (optional, solo si viene en el config)
            if 'radio_cercania' in config:
                rc = config['radio_cercania']
                if not isinstance(rc, int) or rc < 1 or rc > 500:
                    errors.append("radio_cercania must be an integer between 1 and 500")
            # H-6 fix: validar parametros de expansion gradual
            if 'radio_expansion_paso' in config:
                rep = config['radio_expansion_paso']
                if not isinstance(rep, int) or rep < 1 or rep > 1000:
                    errors.append("radio_expansion_paso must be an integer between 1 and 1000")
            if 'radio_max_expansiones' in config:
                rme = config['radio_max_expansiones']
                if not isinstance(rme, int) or rme < 0 or rme > 20:
                    errors.append("radio_max_expansiones must be an integer between 0 and 20")
            
            # Validate staging distribution
            if 'outbound_staging_distribution' in config:
                staging_dist = config['outbound_staging_distribution']
                total_staging = sum(staging_dist.values())
                
                if total_staging != 100:
                    errors.append(f"Staging distribution must sum to 100% (current: {total_staging}%)")
            
            # Validate agent_types if present
            if 'agent_types' in config:
                if not isinstance(config['agent_types'], list):
                    errors.append("agent_types must be a list")
                else:
                    for idx, agent in enumerate(config['agent_types']):
                        if 'type' not in agent:
                            errors.append(f"Agent {idx}: missing type")
                        if 'capacity' not in agent or agent['capacity'] <= 0:
                            errors.append(f"Agent {idx}: invalid capacity")
                        if 'discharge_time' not in agent or agent['discharge_time'] <= 0:
                            errors.append(f"Agent {idx}: invalid discharge_time")
            
            # Paso 2: basic validation of advanced engine blocks (if UI sends them)
            if 'congestion' in config:
                cong = config['congestion']
                if not isinstance(cong, dict):
                    errors.append("congestion must be an object")
                else:
                    if not isinstance(cong.get('enabled', False), bool):
                        errors.append("congestion.enabled must be boolean")
                    if cong.get('mode') is not None and cong.get('mode') not in ('off', 'timewindow'):
                        errors.append("congestion.mode must be 'off' or 'timewindow'")
            if 'outbound' in config:
                ob = config['outbound']
                if not isinstance(ob, dict):
                    errors.append("outbound must be an object")
                else:
                    if not isinstance(ob.get('enabled', False), bool):
                        errors.append("outbound.enabled must be boolean")
                    # truck_interval (opcional): expuesto en la UI; validar rango.
                    if 'truck_interval' in ob:
                        ti = ob['truck_interval']
                        if isinstance(ti, bool) or not isinstance(ti, (int, float)) \
                                or ti < 1 or ti > 3600:
                            errors.append("outbound.truck_interval must be a number between 1 and 3600")

            # Cobertura de areas (BK-04 + hardening QA, Opcion B): la flota debe tener >=1
            # agente y cubrir cada area con un agente del TIPO capaz. El "tipo requerido por
            # area" sale del mapa explicito work_area_equipment (fallback: convencion).
            agent_types = config.get('agent_types', [])
            seq = config.get('sequence_file', '')
            required_areas = self.extract_work_areas(seq) if seq else []
            wae = config.get('work_area_equipment', {})
            if not isinstance(wae, dict):
                errors.append("work_area_equipment debe ser un objeto area->tipo de equipo.")
                wae = {}
            else:
                for k, v in wae.items():
                    if v not in VALID_EQUIPMENT:
                        errors.append("work_area_equipment['" + str(k) + "'] = '" + str(v)
                                      + "' no es un tipo valido (GroundOperator o Forklift).")
                # Completitud: si el mapa esta definido, toda area del layout debe estar en el.
                if wae and required_areas:
                    for area in required_areas:
                        if area not in wae:
                            errors.append(
                                "El area '" + str(area) + "' no tiene tipo de equipo definido "
                                "en 'work_area_equipment'. Definelo (GroundOperator o Forklift).")
            if isinstance(agent_types, list) and agent_types:
                if required_areas:
                    for area in required_areas:
                        cubridores = [a for a in agent_types
                                      if isinstance(a, dict)
                                      and area in (a.get('work_area_priorities', {}) or {})]
                        if not cubridores:
                            errors.append(
                                "El area '" + str(area) + "' no tiene ningun agente "
                                "asignado. Asignala a un grupo en la pestana 'Flota de Agentes'.")
                            continue
                        exp = _expected_equipment_for_area(area, wae)
                        if not any(a.get('type') == exp for a in cubridores):
                            errors.append(
                                "El area '" + str(area) + "' esta cubierta por un tipo de "
                                "agente incorrecto; requiere un " + exp + ".")
            else:
                # QA-2: flota vacia => el fallback usa num_operarios_*; exige >=1 agente.
                n_terr = config.get('num_operarios_terrestres', 0) or 0
                n_fork = config.get('num_montacargas', 0) or 0
                if (n_terr + n_fork) < 1:
                    errors.append(
                        "La flota esta vacia (0 agentes). Agrega al menos un grupo de "
                        "agentes en la pestana 'Flota de Agentes'.")

            # C5: validacion del bloque tiempos (si la UI lo envia)
            if 'tiempos' in config:
                t = config['tiempos']
                if not isinstance(t, dict):
                    errors.append("tiempos must be an object")
                else:
                    for key in ('time_per_cell', 'speed_factor_ground', 'speed_factor_forklift',
                                'tiempo_horquilla', 'cell_size_m'):
                        val = t.get(key)
                        if val is not None:
                            try:
                                if float(val) <= 0:
                                    errors.append(f"tiempos.{key} must be > 0")
                            except (TypeError, ValueError):
                                errors.append(f"tiempos.{key} must be a positive number")
                    pick = t.get('tiempo_picking_por_linea')
                    if pick is not None:
                        try:
                            if float(pick) < 0:
                                errors.append("tiempos.tiempo_picking_por_linea must be >= 0")
                        except (TypeError, ValueError):
                            errors.append("tiempos.tiempo_picking_por_linea must be a number or null")

                    # INIT-4 C1: bloque OPCIONAL pick_time_model (escala de tiempo de pick).
                    # Defaults neutros en el motor; aqui solo validamos si la UI lo envia.
                    ptm = t.get('pick_time_model')
                    if ptm is not None:
                        if not isinstance(ptm, dict):
                            errors.append("tiempos.pick_time_model must be an object or null")
                        else:
                            base = ptm.get('base')
                            if base is not None:
                                try:
                                    if float(base) < 0:
                                        errors.append("tiempos.pick_time_model.base must be >= 0 or null")
                                except (TypeError, ValueError):
                                    errors.append("tiempos.pick_time_model.base must be a number or null")
                            for key in ('por_unidad', 'por_volumen', 'minimo'):
                                val = ptm.get(key)
                                if val is not None:
                                    try:
                                        if float(val) < 0:
                                            errors.append(f"tiempos.pick_time_model.{key} must be >= 0")
                                    except (TypeError, ValueError):
                                        errors.append(f"tiempos.pick_time_model.{key} must be a number")

            # INIT-4 (C2): flag de despacho por prioridad de pedido (bool opcional)
            if 'priority_dispatch_enabled' in config:
                if not isinstance(config['priority_dispatch_enabled'], bool):
                    errors.append("priority_dispatch_enabled must be a boolean")

            # INIT-4 (C3): bloque de olas (waves) opcional
            if 'waves' in config:
                w = config['waves']
                if not isinstance(w, dict):
                    errors.append("waves must be an object")
                else:
                    if 'enabled' in w and not isinstance(w['enabled'], bool):
                        errors.append("waves.enabled must be a boolean")
                    rt = w.get('release_times')
                    if rt is not None:
                        if not isinstance(rt, dict):
                            errors.append("waves.release_times must be an object (wave_id -> seconds)")
                        else:
                            # Acotar releases a un horizonte razonable para evitar hangs.
                            MAX_RELEASE = 1_000_000.0
                            for wid, sec in rt.items():
                                try:
                                    fv = float(sec)
                                    if fv < 0:
                                        errors.append(f"waves.release_times['{wid}'] must be >= 0")
                                    elif fv > MAX_RELEASE:
                                        errors.append(f"waves.release_times['{wid}'] exceeds max horizon ({MAX_RELEASE})")
                                except (TypeError, ValueError):
                                    errors.append(f"waves.release_times['{wid}'] must be a number")

            # MEJ-3: chequeo contra el esquema unico (src/core/config_schema.py).
            # Tipos invalidos en claves conocidas -> ERROR (bloquea el guardado,
            # el motor fallaria o lo ignoraria en silencio). Claves desconocidas
            # o legacy -> WARNING en el log del servidor (no bloquea).
            try:
                import sys as _sys
                _src = os.path.join(self.project_root, "src")
                if _src not in _sys.path:
                    _sys.path.insert(0, _src)
                from core.config_schema import validate_config_schema
                schema_errors, schema_warnings = validate_config_schema(config)
                for w in schema_warnings:
                    print(f"[CONFIG_MANAGER][SCHEMA][WARN] {w}")
                errors.extend(schema_errors)
            except ImportError as _e:
                print(f"[CONFIG_MANAGER][SCHEMA][WARN] esquema no disponible: {_e}")

            is_valid = len(errors) == 0

            if is_valid:
                print(f"[CONFIG_MANAGER] Configuration validation: PASSED")
            else:
                print(f"[CONFIG_MANAGER] Configuration validation: FAILED - {len(errors)} errors")
                for error in errors:
                    print(f"  - {error}")
            
            return is_valid, errors
            
        except Exception as e:
            error_msg = f"Validation error: {str(e)}"
            print(f"[CONFIG_MANAGER ERROR] {error_msg}")
            return False, [error_msg]
    
    def extract_work_areas(self, sequence_file: str) -> List[str]:
        """
        Extract work areas from sequence file (Excel/CSV)
        
        Args:
            sequence_file: Path to sequence file (relative to project root)
            
        Returns:
            List of work area names
        """
        try:
            # Construct absolute path
            if not os.path.isabs(sequence_file):
                sequence_file = os.path.join(self.project_root, sequence_file)
            
            if not os.path.exists(sequence_file):
                print(f"[CONFIG_MANAGER] Sequence file not found: {sequence_file}")
                return []
            
            print(f"[CONFIG_MANAGER] Extracting work areas from: {sequence_file}")
            
            # Read Excel file
            wb = openpyxl.load_workbook(sequence_file, read_only=True, data_only=True)
            
            # Look for PickingLocations or Warehouse_Logic sheet
            sheet_name = None
            if 'PickingLocations' in wb.sheetnames:
                sheet_name = 'PickingLocations'
            elif 'Warehouse_Logic' in wb.sheetnames:
                sheet_name = 'Warehouse_Logic'
            
            if not sheet_name:
                print(f"[CONFIG_MANAGER] No suitable sheet found in {sequence_file}")
                wb.close()
                return []
            
            ws = wb[sheet_name]
            work_areas = set()
            
            # Find WorkArea column
            headers = [cell.value for cell in ws[1]]
            wa_col_idx = None
            for idx, header in enumerate(headers):
                if header and 'WorkArea' in str(header):
                    wa_col_idx = idx
                    break
            
            if wa_col_idx is not None:
                # Extract work areas
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and len(row) > wa_col_idx:
                        work_area = row[wa_col_idx]
                        if work_area:
                            work_areas.add(str(work_area))
            
            wb.close()
            
            result = sorted(list(work_areas))
            print(f"[CONFIG_MANAGER] Extracted {len(result)} work areas: {result}")
            return result
            
        except Exception as e:
            print(f"[CONFIG_MANAGER ERROR] Error extracting work areas: {e}")
            return []
    
    def list_configurations(self) -> List[Dict]:
        """
        List all saved configuration presets
        
        Returns:
            List of configuration metadata dictionaries
        """
        try:
            configs = []
            
            if not os.path.exists(self.presets_dir):
                return configs
            
            for filename in os.listdir(self.presets_dir):
                if not filename.endswith('.json'):
                    continue
                
                filepath = os.path.join(self.presets_dir, filename)
                
                try:
                    with open(filepath, 'r', encoding='utf-8') as f:
                        preset_data = json.load(f)
                    
                    # Extract metadata
                    metadata = preset_data.get('metadata', {})
                    config_id = metadata.get('id', filename.replace('.json', ''))
                    
                    configs.append({
                        'id': config_id,
                        'name': metadata.get('name', filename),
                        'description': metadata.get('description', ''),
                        'created_at': metadata.get('created_at', ''),
                        'is_default': metadata.get('is_default', False),
                        'filename': filename
                    })
                    
                except Exception as e:
                    print(f"[CONFIG_MANAGER] Error loading preset {filename}: {e}")
                    continue
            
            # Sort by creation date, newest first
            configs.sort(key=lambda x: x.get('created_at', ''), reverse=True)
            
            print(f"[CONFIG_MANAGER] Listed {len(configs)} configurations")
            return configs
            
        except Exception as e:
            print(f"[CONFIG_MANAGER ERROR] Error listing configurations: {e}")
            return []
    
    def save_configuration(self, name: str, description: str, config: Dict, 
                          is_default: bool = False) -> Tuple[bool, str, List[str]]:
        """
        Save configuration as a preset
        
        Args:
            name: Configuration name
            description: Configuration description
            config: Configuration dictionary
            is_default: Whether this should be the default configuration
            
        Returns:
            Tuple of (success, config_id, error_messages)
        """
        try:
            # Validate configuration first
            is_valid, errors = self.validate_config(config)
            if not is_valid:
                return False, "", errors
            
            # Generate unique ID
            config_id = str(uuid.uuid4())
            
            # Create preset data structure
            preset_data = {
                'metadata': {
                    'id': config_id,
                    'name': name,
                    'description': description,
                    'created_at': datetime.now().isoformat(),
                    'is_default': is_default
                },
                'configuration': config
            }
            
            # If setting as default, unset other defaults
            if is_default:
                self._unset_all_defaults()
            
            # Save preset file
            filename = f"{config_id}.json"
            filepath = os.path.join(self.presets_dir, filename)
            
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(preset_data, f, indent=4, ensure_ascii=False)
            
            print(f"[CONFIG_MANAGER] Saved configuration '{name}' as {filename}")
            return True, config_id, []
            
        except Exception as e:
            error_msg = f"Error saving configuration: {str(e)}"
            print(f"[CONFIG_MANAGER ERROR] {error_msg}")
            return False, "", [error_msg]
    
    def load_configuration(self, config_id: str) -> Optional[Dict]:
        """
        Load a configuration preset by ID
        
        Args:
            config_id: Configuration ID
            
        Returns:
            Configuration dictionary or None if not found
        """
        try:
            filename = f"{config_id}.json"
            filepath = os.path.join(self.presets_dir, filename)
            
            if not os.path.exists(filepath):
                print(f"[CONFIG_MANAGER] Configuration {config_id} not found")
                return None
            
            with open(filepath, 'r', encoding='utf-8') as f:
                preset_data = json.load(f)
            
            config = preset_data.get('configuration', {})
            print(f"[CONFIG_MANAGER] Loaded configuration {config_id}")
            return config
            
        except Exception as e:
            print(f"[CONFIG_MANAGER ERROR] Error loading configuration {config_id}: {e}")
            return None
    
    def delete_configuration(self, config_id: str) -> Tuple[bool, List[str]]:
        """
        Delete a configuration preset
        
        Args:
            config_id: Configuration ID
            
        Returns:
            Tuple of (success, error_messages)
        """
        try:
            filename = f"{config_id}.json"
            filepath = os.path.join(self.presets_dir, filename)
            
            if not os.path.exists(filepath):
                return False, [f"Configuration {config_id} not found"]
            
            os.remove(filepath)
            print(f"[CONFIG_MANAGER] Deleted configuration {config_id}")
            return True, []
            
        except Exception as e:
            error_msg = f"Error deleting configuration: {str(e)}"
            print(f"[CONFIG_MANAGER ERROR] {error_msg}")
            return False, [error_msg]
    
    def set_default_configuration(self, config_id: str) -> Tuple[bool, List[str]]:
        """
        Set a configuration as the default
        
        Args:
            config_id: Configuration ID
            
        Returns:
            Tuple of (success, error_messages)
        """
        try:
            filename = f"{config_id}.json"
            filepath = os.path.join(self.presets_dir, filename)
            
            if not os.path.exists(filepath):
                return False, [f"Configuration {config_id} not found"]
            
            # Unset all defaults first
            self._unset_all_defaults()
            
            # Load and update this configuration
            with open(filepath, 'r', encoding='utf-8') as f:
                preset_data = json.load(f)
            
            preset_data['metadata']['is_default'] = True
            
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(preset_data, f, indent=4, ensure_ascii=False)
            
            print(f"[CONFIG_MANAGER] Set {config_id} as default configuration")
            return True, []
            
        except Exception as e:
            error_msg = f"Error setting default configuration: {str(e)}"
            print(f"[CONFIG_MANAGER ERROR] {error_msg}")
            return False, [error_msg]
    
    def get_default_configuration(self) -> Optional[Dict]:
        """
        Get the default configuration
        
        Returns:
            Configuration dictionary or None if no default is set
        """
        try:
            configs = self.list_configurations()
            
            for config_meta in configs:
                if config_meta.get('is_default', False):
                    return self.load_configuration(config_meta['id'])
            
            print(f"[CONFIG_MANAGER] No default configuration found")
            return None
            
        except Exception as e:
            print(f"[CONFIG_MANAGER ERROR] Error getting default configuration: {e}")
            return None
    
    def _unset_all_defaults(self):
        """Unset default flag from all configurations"""
        try:
            if not os.path.exists(self.presets_dir):
                return
            
            for filename in os.listdir(self.presets_dir):
                if not filename.endswith('.json'):
                    continue
                
                filepath = os.path.join(self.presets_dir, filename)
                
                try:
                    with open(filepath, 'r', encoding='utf-8') as f:
                        preset_data = json.load(f)
                    
                    if preset_data.get('metadata', {}).get('is_default', False):
                        preset_data['metadata']['is_default'] = False
                        
                        with open(filepath, 'w', encoding='utf-8') as f:
                            json.dump(preset_data, f, indent=4, ensure_ascii=False)
                        
                        print(f"[CONFIG_MANAGER] Unset default flag from {filename}")
                
                except Exception as e:
                    print(f"[CONFIG_MANAGER] Error processing {filename}: {e}")
                    continue
                    
        except Exception as e:
            print(f"[CONFIG_MANAGER ERROR] Error unsetting defaults: {e}")
    
    def _get_default_config(self) -> Dict:
        """Get hardcoded default configuration"""
        return {
            "total_ordenes": 300,
            "distribucion_tipos": {
                "pequeno": {"porcentaje": 60, "volumen": 5},
                "mediano": {"porcentaje": 30, "volumen": 25},
                "grande": {"porcentaje": 10, "volumen": 80}
            },
            "dispatch_strategy": "Optimizacion Global",
            "radio_cercania": 100,
            "radio_expansion_paso": 50,
            "radio_max_expansiones": 5,
            "tour_type": "Tour Mixto (Multi-Destino)",
            "layout_file": "layouts/WH1.tmx",
            "sequence_file": "layouts/Warehouse_Logic.xlsx",
            "work_area_equipment": {
                "Area_Ground": "GroundOperator",
                "Area_High": "Forklift",
                "Area_Special": "Forklift"
            },
            "num_operarios_terrestres": 2,
            "num_montacargas": 1,
            "num_operarios_total": 3,
            "outbound_staging_distribution": {
                "1": 100, "2": 0, "3": 0, "4": 0, "5": 0, "6": 0, "7": 0
            },
            "agent_types": []
        }

# -*- coding: utf-8 -*-
"""
INIT-7 F0 / IN-xx: contrato de dominio y datos del inbound (recepcion).

Pinea:
- El bloque 'inbound' del config_schema (opt-in, tipos, claves desconocidas).
- La hoja 'InboundDocks' del Excel canonico (3 muelles en fila superior).
- El importer Excel -> SQLite (tabla inbound_docks, sin tocar staging_areas).
- El loader del data_manager tolera DBs viejas SIN la tabla (opt-in real).
- El ASN de ejemplo (layouts/Inbound Test.json) cumple su propio contrato.

El motor NO lee nada de esto hasta F1: el gate byte-identico debe pasar
sin actualizar baseline.
"""
import json
import os
import sqlite3

import openpyxl
import pytest

from core.config_schema import validate_config_schema

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
XLSX_PATH = os.path.join(PROJECT_ROOT, "layouts", "Warehouse_Logic.xlsx")
ASN_PATH = os.path.join(PROJECT_ROOT, "layouts", "Inbound Test.json")

INBOUND_BLOCK = {
    "enabled": False,
    "arrival_mode": "deterministic",
    "asn_file_path": "layouts/Inbound Test.json",
    "truck_interval": 600.0,
    "pallets_per_truck": 10,
    "unload_time_per_pallet": 15.0,
    "slotting_strategy": "cercana_al_muelle",
}

CANONICAL_DOCKS = {1: (3, 1), 2: (15, 1), 3: (27, 1)}


def _canonical_config():
    with open(os.path.join(PROJECT_ROOT, "config.json"), "r", encoding="utf-8") as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# Config schema
# ---------------------------------------------------------------------------

def test_in01_bloque_inbound_completo_es_valido():
    cfg = _canonical_config()
    cfg["inbound"] = dict(INBOUND_BLOCK)
    errors, warnings = validate_config_schema(cfg)
    assert errors == [], errors
    assert warnings == [], warnings


def test_in02_clave_desconocida_dentro_de_inbound_avisa():
    cfg = _canonical_config()
    cfg["inbound"] = {"enabled": False, "slotting_strateg": "abc_rotacion"}  # typo
    errors, warnings = validate_config_schema(cfg)
    assert errors == []
    assert any("inbound.slotting_strateg" in w for w in warnings), warnings


def test_in03_tipo_invalido_en_inbound_es_error():
    cfg = _canonical_config()
    cfg["inbound"] = {"enabled": False, "truck_interval": "rapido"}
    errors, warnings = validate_config_schema(cfg)
    assert any("inbound.truck_interval" in e for e in errors), errors


# ---------------------------------------------------------------------------
# Excel canonico: hoja InboundDocks
# ---------------------------------------------------------------------------

def test_in04_excel_canonico_tiene_hoja_inbound_docks():
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    try:
        assert "InboundDocks" in wb.sheetnames
        ws = wb["InboundDocks"]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    headers = [str(c).strip() for c in rows[0]]
    assert headers[:3] == ["dock_id", "x", "y"]

    docks = {int(r[0]): (int(r[1]), int(r[2])) for r in rows[1:] if r and r[0]}
    assert docks == CANONICAL_DOCKS


# ---------------------------------------------------------------------------
# Importer Excel -> SQLite
# ---------------------------------------------------------------------------

def test_in05_importer_puebla_inbound_docks_sin_tocar_staging(tmp_path):
    from subsystems.database.database_manager import DatabaseManager
    from subsystems.database.importer import import_warehouse_data

    db_path = str(tmp_path / "test_warehouse.db")
    try:
        result = import_warehouse_data(XLSX_PATH, db_path=db_path)
        assert result.success, result.errors
        assert result.inbound_docks_imported == 3

        conn = sqlite3.connect(db_path)
        try:
            docks = {row[0]: (row[1], row[2]) for row in
                     conn.execute("SELECT dock_id, x, y FROM inbound_docks")}
            assert docks == CANONICAL_DOCKS

            # staging outbound intacto (7 zonas, ningun id pisado)
            n_outbound = conn.execute(
                "SELECT COUNT(*) FROM staging_areas WHERE staging_type = 'OUTBOUND'"
            ).fetchone()[0]
            assert n_outbound == result.staging_imported
            assert n_outbound >= 7
        finally:
            conn.close()
    finally:
        DatabaseManager.reset_instance()


def test_in06_importer_tolera_db_vieja_sin_tabla(tmp_path):
    """DB creada ANTES de INIT-7 (schema_version > 0, sin inbound_docks):
    _clear_tables debe crear la tabla en vez de reventar."""
    from subsystems.database.importer import ExcelImporter

    db_path = str(tmp_path / "old.db")
    conn = sqlite3.connect(db_path)
    conn.execute("CREATE TABLE dummy (id INTEGER)")
    ExcelImporter._ensure_inbound_docks_table(conn)
    conn.execute("DELETE FROM inbound_docks")  # no lanza: la tabla existe
    conn.close()


# ---------------------------------------------------------------------------
# DataManager: loader DB tolerante y sin defaults
# ---------------------------------------------------------------------------

def _bare_data_manager():
    from subsystems.simulation.data_manager import DataManager
    dm = object.__new__(DataManager)  # sin __init__: no TMX ni pygame
    dm.inbound_dock_locations = {}
    return dm


def test_in07_loader_db_vieja_sin_tabla_queda_vacio():
    dm = _bare_data_manager()
    conn = sqlite3.connect(":memory:")
    conn.row_factory = sqlite3.Row
    dm._load_inbound_docks_from_db(conn)  # tabla ausente: NO lanza
    conn.close()
    assert dm.inbound_dock_locations == {}


def test_in08_loader_db_lee_muelles():
    dm = _bare_data_manager()
    conn = sqlite3.connect(":memory:")
    conn.row_factory = sqlite3.Row
    conn.execute("CREATE TABLE inbound_docks (dock_id INTEGER PRIMARY KEY, "
                 "x INTEGER NOT NULL, y INTEGER NOT NULL)")
    conn.executemany("INSERT INTO inbound_docks VALUES (?, ?, ?)",
                     [(1, 3, 1), (2, 15, 1), (3, 27, 1)])
    dm._load_inbound_docks_from_db(conn)
    conn.close()
    assert dm.inbound_dock_locations == CANONICAL_DOCKS


# ---------------------------------------------------------------------------
# ASN de ejemplo: contrato del archivo
# ---------------------------------------------------------------------------

def test_in09_asn_de_ejemplo_cumple_contrato():
    with open(ASN_PATH, "r", encoding="utf-8") as f:
        asn = json.load(f)

    assert "trucks" in asn and len(asn["trucks"]) > 0

    # SKUs reales del Excel canonico
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    try:
        ws = wb["PickingLocations"]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    headers = [str(c).strip().lower() if c else "" for c in rows[0]]
    sku_idx = next(i for i, h in enumerate(headers) if "sku" in h)
    skus_reales = {str(r[sku_idx]) for r in rows[1:] if r and r[sku_idx]}

    prev_arrival = -1
    for truck in asn["trucks"]:
        assert truck["truck_id"]
        assert isinstance(truck["arrival_time"], (int, float))
        assert truck["arrival_time"] >= 0
        # orden cronologico (contrato del ASN: llegadas ordenadas)
        assert truck["arrival_time"] >= prev_arrival
        prev_arrival = truck["arrival_time"]
        # dock_id es OPCIONAL; si viene, debe ser un muelle canonico
        if "dock_id" in truck:
            assert truck["dock_id"] in CANONICAL_DOCKS
        assert len(truck["lines"]) > 0
        for line in truck["lines"]:
            assert line["sku_id"] in skus_reales, line["sku_id"]
            assert line["quantity"] > 0

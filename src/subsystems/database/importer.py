# -*- coding: utf-8 -*-
"""
Excel Importer Module - Warehouse_Logic.xlsx to SQLite
Digital Twin Warehouse Simulator

Imports data from Excel files into the SQLite database with validation.
Handles the current Excel format with columns:
- x, y, pick_sequence, equipment_required, sku_initial, qty_initial, WorkGroup, WorkArea

Author: Digital Twin Warehouse Team
Version: V1.0
"""

import os
import sqlite3
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass, field
from datetime import datetime

try:
    import openpyxl
except ImportError:
    raise ImportError("openpyxl is required. Install with: pip install openpyxl")

from .database_manager import DatabaseManager


@dataclass
class ImportResult:
    """Result of an import operation."""
    success: bool
    skus_imported: int = 0
    locations_imported: int = 0
    inventory_imported: int = 0
    staging_imported: int = 0
    warnings: List[str] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)
    
    def add_warning(self, msg: str):
        self.warnings.append(msg)
        print(f"[IMPORTER WARNING] {msg}")
    
    def add_error(self, msg: str):
        self.errors.append(msg)
        print(f"[IMPORTER ERROR] {msg}")


class ExcelImporter:
    """
    Imports Warehouse_Logic.xlsx into SQLite database.
    
    Expected Excel structure:
    - Sheet 'PickingLocations': x, y, pick_sequence, equipment_required, 
                                 sku_initial, qty_initial, WorkGroup, WorkArea
    - Sheet 'OutboundStaging' (optional): staging_id, x, y
    
    The importer generates location_ids based on pick_sequence if not provided.
    """
    
    def __init__(self, db_manager: Optional[DatabaseManager] = None):
        """
        Initialize importer.
        
        Args:
            db_manager: DatabaseManager instance. If None, uses singleton.
        """
        self.db = db_manager or DatabaseManager.get_instance()
    
    def import_from_excel(self, excel_path: str, 
                          clear_existing: bool = True) -> ImportResult:
        """
        Import data from Warehouse_Logic.xlsx into SQLite.
        
        Args:
            excel_path: Path to Excel file
            clear_existing: If True, clears existing data before import
            
        Returns:
            ImportResult with counts and any errors/warnings
        """
        result = ImportResult(success=False)
        
        # Validate file exists
        if not os.path.exists(excel_path):
            result.add_error(f"Excel file not found: {excel_path}")
            return result
        
        print(f"[IMPORTER] Starting import from: {excel_path}")
        
        try:
            workbook = openpyxl.load_workbook(excel_path, data_only=True)
        except Exception as e:
            result.add_error(f"Failed to open Excel file: {e}")
            return result
        
        # Ensure schema exists
        if not self.db.database_exists() or self.db.get_schema_version() == 0:
            print("[IMPORTER] Initializing database schema...")
            self.db.initialize_schema()
        
        try:
            with self.db.get_connection() as conn:
                if clear_existing:
                    self._clear_tables(conn)
                
                # Process PickingLocations sheet (required)
                if 'PickingLocations' in workbook.sheetnames:
                    sheet = workbook['PickingLocations']
                    self._import_picking_locations(sheet, conn, result)
                else:
                    result.add_error("Sheet 'PickingLocations' not found in Excel")
                    return result
                
                # Process OutboundStaging sheet (optional)
                if 'OutboundStaging' in workbook.sheetnames:
                    sheet = workbook['OutboundStaging']
                    self._import_staging_areas(sheet, conn, result)
                else:
                    result.add_warning("Sheet 'OutboundStaging' not found - skipping")
                
                result.success = len(result.errors) == 0
                
        except Exception as e:
            result.add_error(f"Import failed: {e}")
            import traceback
            traceback.print_exc()
        finally:
            workbook.close()
        
        self._print_summary(result)
        return result
    
    def _clear_tables(self, conn: sqlite3.Connection):
        """Clear all data tables (preserves schema)."""
        print("[IMPORTER] Clearing existing data...")
        # Order matters due to foreign keys
        conn.execute("DELETE FROM order_lines")
        conn.execute("DELETE FROM orders")
        conn.execute("DELETE FROM inventory")
        conn.execute("DELETE FROM locations")
        conn.execute("DELETE FROM sku_catalog")
        conn.execute("DELETE FROM staging_areas")
    
    def _find_header_row(self, sheet, marker_column: str, 
                         max_rows: int = 10) -> Optional[int]:
        """Find the row containing headers."""
        for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_rows, 
                                                   values_only=True), start=1):
            if row:
                for cell in row:
                    if cell and str(cell).lower() == marker_column.lower():
                        return idx
        return None
    
    def _get_headers(self, sheet, header_row: int) -> List[str]:
        """Get header names from a row."""
        headers = []
        for cell in sheet[header_row]:
            value = cell.value
            headers.append(str(value).strip() if value else '')
        return headers
    
    def _row_to_dict(self, row: tuple, headers: List[str]) -> Dict[str, Any]:
        """Convert a row tuple to a dict using headers."""
        result = {}
        for idx, header in enumerate(headers):
            if header and idx < len(row):
                result[header] = row[idx]
        return result
    
    def _import_picking_locations(self, sheet, conn: sqlite3.Connection, 
                                   result: ImportResult):
        """Import PickingLocations sheet."""
        print("[IMPORTER] Processing 'PickingLocations' sheet...")
        
        # Find header row
        header_row = self._find_header_row(sheet, 'x')
        if not header_row:
            result.add_error("Header row with 'x' column not found")
            return
        
        headers = self._get_headers(sheet, header_row)
        print(f"[IMPORTER] Found headers: {headers}")
        
        # Collect all unique SKUs and locations
        skus_seen = set()
        locations_data = []
        inventory_data = []
        
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not row[0]:  # Skip empty rows
                continue
            
            data = self._row_to_dict(row, headers)
            
            try:
                # Parse coordinates
                x = self._safe_int(data.get('x'), 0)
                y = self._safe_int(data.get('y'), 0)
                pick_sequence = self._safe_int(data.get('pick_sequence'), 0)
                
                # Generate location_id if not present
                location_id = data.get('location_id')
                if not location_id:
                    # Generate from sequence: LOC-001, LOC-002, etc.
                    location_id = f"LOC-{pick_sequence:03d}"
                
                # Get SKU and quantity
                sku_code = str(data.get('sku_initial', '')).strip()
                if not sku_code or sku_code.lower() == 'none':
                    sku_code = f"SKU{pick_sequence:03d}"  # Generate if missing
                
                qty = self._safe_int(data.get('qty_initial'), 1)
                
                # Get work area/group
                work_area = str(data.get('WorkArea', 'Area_Ground')).strip()
                work_group = str(data.get('WorkGroup', 'WG_Default')).strip()
                equipment = str(data.get('equipment_required', 'GroundOperator')).strip()
                
                # Collect SKU
                if sku_code not in skus_seen:
                    skus_seen.add(sku_code)
                
                # Collect location
                locations_data.append({
                    'location_id': location_id,
                    'location_type': 'PICKING',
                    'work_area': work_area,
                    'work_group': work_group,
                    'pick_sequence': pick_sequence,
                    'equipment_required': equipment,
                    'legacy_x': x,
                    'legacy_y': y
                })
                
                # Collect inventory
                inventory_data.append({
                    'location_id': location_id,
                    'sku_code': sku_code,
                    'qty_available': qty,
                    'qty_reserved': 0
                })
                
            except Exception as e:
                result.add_warning(f"Skipping invalid row: {data} - {e}")
                continue
        
        # Insert SKUs
        print(f"[IMPORTER] Inserting {len(skus_seen)} SKUs...")
        for sku in sorted(skus_seen):
            conn.execute("""
                INSERT OR IGNORE INTO sku_catalog (sku_code, description, equipment_required)
                VALUES (?, ?, 'GroundOperator')
            """, (sku, f"SKU {sku}"))
        result.skus_imported = len(skus_seen)
        
        # Insert locations
        print(f"[IMPORTER] Inserting {len(locations_data)} locations...")
        for loc in locations_data:
            try:
                conn.execute("""
                    INSERT OR REPLACE INTO locations 
                    (location_id, location_type, work_area, work_group, 
                     pick_sequence, equipment_required, legacy_x, legacy_y)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    loc['location_id'], loc['location_type'], loc['work_area'],
                    loc['work_group'], loc['pick_sequence'], loc['equipment_required'],
                    loc['legacy_x'], loc['legacy_y']
                ))
            except sqlite3.IntegrityError as e:
                result.add_warning(f"Duplicate location {loc['location_id']}: {e}")
        result.locations_imported = len(locations_data)
        
        # Insert inventory
        print(f"[IMPORTER] Inserting {len(inventory_data)} inventory records...")
        for inv in inventory_data:
            try:
                conn.execute("""
                    INSERT OR REPLACE INTO inventory 
                    (location_id, sku_code, qty_available, qty_reserved)
                    VALUES (?, ?, ?, ?)
                """, (
                    inv['location_id'], inv['sku_code'], 
                    inv['qty_available'], inv['qty_reserved']
                ))
            except sqlite3.IntegrityError as e:
                result.add_warning(f"Inventory error for {inv['location_id']}: {e}")
        result.inventory_imported = len(inventory_data)
    
    def _import_staging_areas(self, sheet, conn: sqlite3.Connection, 
                              result: ImportResult):
        """Import OutboundStaging sheet."""
        print("[IMPORTER] Processing 'OutboundStaging' sheet...")
        
        # Find header row (look for 'staging' in any cell)
        header_row = None
        for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, 
                                                   values_only=True), start=1):
            if row:
                for cell in row:
                    if cell and 'staging' in str(cell).lower():
                        header_row = idx
                        break
                if header_row:
                    break
        
        if not header_row:
            header_row = 1  # Assume first row is header
        
        headers = self._get_headers(sheet, header_row)
        count = 0
        
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not row[0]:
                continue
            
            data = self._row_to_dict(row, headers)
            
            try:
                # Find staging_id column (various naming conventions)
                staging_id = None
                for key in data.keys():
                    if 'staging' in key.lower() and 'id' in key.lower():
                        staging_id = self._safe_int(data[key])
                        break
                
                if staging_id is None:
                    staging_id = self._safe_int(data.get('staging_id'), 0)
                
                x = self._safe_int(data.get('x'), 0)
                y = self._safe_int(data.get('y'), 0)
                
                if staging_id > 0:
                    conn.execute("""
                        INSERT OR REPLACE INTO staging_areas 
                        (staging_id, staging_type, legacy_x, legacy_y)
                        VALUES (?, 'OUTBOUND', ?, ?)
                    """, (staging_id, x, y))
                    count += 1
                    
            except Exception as e:
                result.add_warning(f"Skipping staging row: {data} - {e}")
        
        result.staging_imported = count
        print(f"[IMPORTER] Imported {count} staging areas")
    
    @staticmethod
    def _safe_int(value, default: int = 0) -> int:
        """Safely convert value to int."""
        if value is None:
            return default
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return default
    
    def _print_summary(self, result: ImportResult):
        """Print import summary."""
        print("\n" + "=" * 50)
        print("[IMPORTER] IMPORT SUMMARY")
        print("=" * 50)
        print(f"  Status: {'SUCCESS' if result.success else 'FAILED'}")
        print(f"  SKUs imported: {result.skus_imported}")
        print(f"  Locations imported: {result.locations_imported}")
        print(f"  Inventory records: {result.inventory_imported}")
        print(f"  Staging areas: {result.staging_imported}")
        print(f"  Warnings: {len(result.warnings)}")
        print(f"  Errors: {len(result.errors)}")
        print("=" * 50 + "\n")


def import_warehouse_data(excel_path: str, db_path: Optional[str] = None,
                          clear_existing: bool = True) -> ImportResult:
    """
    Convenience function to import warehouse data.
    
    Args:
        excel_path: Path to Warehouse_Logic.xlsx
        db_path: Optional path to database (uses default if None)
        clear_existing: Clear existing data before import
        
    Returns:
        ImportResult with import statistics
    """
    # Reset singleton if db_path specified
    if db_path:
        DatabaseManager.reset_instance()
        db = DatabaseManager.get_instance(db_path)
    else:
        db = DatabaseManager.get_instance()
    
    importer = ExcelImporter(db)
    return importer.import_from_excel(excel_path, clear_existing)

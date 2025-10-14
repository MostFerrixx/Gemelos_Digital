# -*- coding: utf-8 -*-
"""
Analytics Engine para generar reportes ejecutivos en Excel.
Procesa eventos de simulacion y calcula metricas clave de rendimiento.
"""

import pandas as pd
import json
import argparse
from typing import List, Dict
from datetime import datetime, timedelta
from openpyxl.formatting.rule import ColorScaleRule


class AnalyticsEngine:
    """Motor de analiticas para generar reportes ejecutivos a partir de eventos de simulacion"""
    
    def __init__(self, event_log: list, config: dict):
        """
        Inicializa el motor de analiticas.
        
        Args:
            event_log: Lista de eventos capturados durante la simulacion
            config: Configuracion utilizada en la simulacion
        """
        self.event_log = event_log
        self.config = config
        self.events_df = None
        self.summary_kpis = None
        self.agent_performance = None
        self.heatmap_data = None
        
        print(f"[ANALYTICS-ENGINE] Inicializado con {len(event_log)} eventos")
    
    @classmethod
    def from_json_file(cls, json_filepath: str):
        """
        Constructor alternativo que carga eventos desde un archivo raw_events.json.
        
        Args:
            json_filepath: Ruta al archivo raw_events.json
            
        Returns:
            Nueva instancia de AnalyticsEngine cargada desde archivo
        """
        print(f"[ANALYTICS-ENGINE] Cargando eventos desde: {json_filepath}")
        
        try:
            with open(json_filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # Extraer eventos y configuracion del archivo JSON
            if isinstance(data, dict):
                # Formato: {"events": [...], "config": {...}}
                event_log = data.get('events', [])
                config = data.get('config', {})
            elif isinstance(data, list):
                # Formato: Lista directa de eventos (sin configuracion embebida)
                event_log = data
                config = {}
                print("[ANALYTICS-ENGINE] ADVERTENCIA: No se encontro configuracion en el archivo JSON")
            else:
                raise ValueError("Formato de archivo JSON no reconocido")
            
            print(f"[ANALYTICS-ENGINE] Cargados {len(event_log)} eventos del archivo")
            return cls(event_log=event_log, config=config)
            
        except FileNotFoundError:
            print(f"[ANALYTICS-ENGINE] ERROR: Archivo no encontrado: {json_filepath}")
            raise
        except json.JSONDecodeError as e:
            print(f"[ANALYTICS-ENGINE] ERROR: Archivo JSON malformado: {e}")
            raise
        except Exception as e:
            print(f"[ANALYTICS-ENGINE] ERROR al cargar archivo: {e}")
            raise
    
    def process_events(self):
        """
        Metodo principal que procesa todos los eventos y calcula metricas.
        """
        print("[ANALYTICS-ENGINE] Procesando eventos...")
        
        # Convertir eventos a DataFrame para facilitar analisis
        if not self.event_log:
            print("[ANALYTICS-ENGINE] ADVERTENCIA: No hay eventos para procesar")
            return
        
        self.events_df = pd.DataFrame(self.event_log)
        print(f"[ANALYTICS-ENGINE] DataFrame creado con {len(self.events_df)} filas")
        
        # Calcular metricas
        self.summary_kpis = self._calculate_summary_kpis()
        self.agent_performance = self._calculate_agent_performance()
        self.heatmap_data = self._calculate_heatmap_data()
        
        print("[ANALYTICS-ENGINE] Procesamiento completado")
    
    def _calculate_summary_kpis(self) -> pd.DataFrame:
        """
        Calcula las metricas clave para la hoja "Resumen Ejecutivo".
        
        Returns:
            DataFrame con metricas de resumen ejecutivo
        """
        print("[ANALYTICS-ENGINE] Calculando KPIs de resumen ejecutivo...")
        
        if self.events_df is None or self.events_df.empty:
            return pd.DataFrame()
        
        # Tiempo total de simulacion
        tiempo_inicio = self.events_df['timestamp'].min()
        tiempo_fin = self.events_df['timestamp'].max()
        tiempo_total_sim_segundos = tiempo_fin - tiempo_inicio
        
        print(f"[ANALYTICS-ENGINE] DEBUG: Tiempo simulación - Inicio: {tiempo_inicio}s, Fin: {tiempo_fin}s, Total: {tiempo_total_sim_segundos}s")
        
        # Total de tareas completadas
        work_order_completed_events = self.events_df[self.events_df['tipo'] == 'work_order_completed']
        total_tareas_completadas = len(work_order_completed_events)
        
        # Productividad (Tareas/Segundo)
        if tiempo_total_sim_segundos > 0:
            productividad_por_segundo = total_tareas_completadas / tiempo_total_sim_segundos
        else:
            productividad_por_segundo = 0
        
        # Crear DataFrame de resumen
        summary_data = {
            'Metrica': [
                'Tiempo Total de Simulacion (segundos)',
                'Total de Tareas Completadas (unidades)',
                'Productividad (Tareas/Segundo)',
                'Tiempo de Inicio (segundos)',
                'Tiempo de Fin (segundos)'
            ],
            'Valor': [
                round(tiempo_total_sim_segundos, 2),
                total_tareas_completadas,
                round(productividad_por_segundo, 4),
                round(tiempo_inicio, 2),
                round(tiempo_fin, 2)
            ]
        }
        
        summary_df = pd.DataFrame(summary_data)
        print(f"[ANALYTICS-ENGINE] KPIs calculados: {len(summary_df)} metricas")
        return summary_df
    
    def _calculate_agent_performance(self) -> pd.DataFrame:
        """
        REFACTOR: Calculates agent performance metrics based on the time difference
        between consecutive 'estado_agente' events.
        
        Returns:
            DataFrame with performance metrics per agent.
        """
        print("[ANALYTICS-ENGINE] Calculating agent performance using state-based time differences...")

        if self.events_df is None or self.events_df.empty:
            return pd.DataFrame()

        # --- Agent ID Normalization --- 
        def normalize_agent_id(agent_id):
            if isinstance(agent_id, str):
                return agent_id.replace("GroundOperator_", "").replace("Forklift_", "")
            return agent_id

        # Filter events and apply normalization
        agent_events = self.events_df[self.events_df['tipo'] == 'estado_agente'].copy()
        agent_events['agent_id'] = agent_events['agent_id'].apply(normalize_agent_id)

        wo_completed_events = self.events_df[self.events_df['tipo'] == 'work_order_completed'].copy()
        wo_completed_events['agent_id'] = wo_completed_events['agent_id'].apply(normalize_agent_id)

        if agent_events.empty:
            print("[ANALYTICS-ENGINE] No 'estado_agente' events found to calculate performance.")
            return pd.DataFrame()

        if 'status' not in agent_events.columns:
            print("[ANALYTICS-ENGINE] 'status' column not found in 'estado_agente' events.")
            return pd.DataFrame()

        agent_ids = agent_events['agent_id'].unique()
        performance_data = []
        
        sim_end_time = self.events_df['timestamp'].max()

        for agent_id in agent_ids:
            agent_df = agent_events[agent_events['agent_id'] == agent_id].sort_values(by='timestamp')
            agent_df['duration'] = agent_df['timestamp'].diff().shift(-1)
            
            last_event_index = agent_df.index[-1]
            last_event_timestamp = agent_df.loc[last_event_index, 'timestamp']
            last_duration = sim_end_time - last_event_timestamp
            agent_df.loc[last_event_index, 'duration'] = last_duration

            status_times = agent_df.groupby('status')['duration'].sum()

            tiempo_viaje = status_times.get('moving', 0)
            tiempo_picking = status_times.get('picking', 0)
            tiempo_descarga = status_times.get('dropping', 0) + status_times.get('unloading', 0)
            tiempo_ocioso = status_times.get('idle', 0)

            tareas_completadas = len(wo_completed_events[wo_completed_events['agent_id'] == agent_id])

            agent_info = agent_df.iloc[0]
            tipo_agente = agent_info['agent_type']
            eventos_totales = len(agent_df)

            tiempo_activo = tiempo_viaje + tiempo_picking + tiempo_descarga + status_times.get('lifting', 0)
            tiempo_suma_estados = tiempo_activo + tiempo_ocioso

            tiempo_total_simulacion = self.events_df['timestamp'].max() - self.events_df['timestamp'].min()
            utilizacion_capacidad = 0
            if tiempo_total_simulacion > 0:
                utilizacion_capacidad = (tiempo_activo / tiempo_total_simulacion) * 100

            performance_data.append({
                'Agent_ID': agent_id,
                'Tipo_Agente': tipo_agente,
                'Tareas_Completadas (unidades)': tareas_completadas,
                'Tiempo_Activo (segundos)': round(tiempo_activo, 2),
                'Tiempo_Picking (segundos)': round(tiempo_picking, 2),
                'Tiempo_Viaje (segundos)': round(tiempo_viaje, 2),
                'Tiempo_Descarga (segundos)': round(tiempo_descarga, 2),
                'Tiempo_Elevacion (segundos)': round(status_times.get('lifting', 0), 2),
                'Tiempo_Ocioso (segundos)': round(tiempo_ocioso, 2),
                'Tiempo_Suma_Estados (segundos)': round(tiempo_suma_estados, 2),
                'Utilizacion_Capacidad_Promedio (%)': round(utilizacion_capacidad, 2),
                'Eventos_Totales (unidades)': eventos_totales
            })

        performance_df = pd.DataFrame(performance_data)
        print(f"[ANALYTICS-ENGINE] State-based performance calculated for {len(performance_df)} agents")
        return performance_df    
    def _calculate_heatmap_data(self) -> pd.DataFrame:
        """
        Calcula los datos necesarios para generar heatmaps de actividad en el almacen.
        Analiza tiempo de transito y tiempo de trabajo por coordenada.
        
        Returns:
            DataFrame con columnas: x, y, tiempo_trabajo, tiempo_transito, tiempo_total
        """
        print("[ANALYTICS-ENGINE] Calculando datos de heatmap...")
        
        if self.events_df is None or self.events_df.empty:
            return pd.DataFrame()
        
        # Obtener dimensiones del almacen desde configuracion
        # Valores por defecto si no estan disponibles
        warehouse_width = self.config.get('warehouse_width', 50)
        warehouse_height = self.config.get('warehouse_height', 35)
        
        print(f"[ANALYTICS-ENGINE] Dimensiones del almacen: {warehouse_width}x{warehouse_height}")
        
        # Crear DataFrame base con todas las coordenadas del almacen
        heatmap_data = []
        for x in range(warehouse_width):
            for y in range(warehouse_height):
                heatmap_data.append({
                    'x (coordenada)': x,
                    'y (coordenada)': y,
                    'tiempo_trabajo (segundos)': 0.0,
                    'tiempo_transito (segundos)': 0.0,
                    'tiempo_total (segundos)': 0.0
                })
        
        heatmap_df = pd.DataFrame(heatmap_data)
        print(f"[ANALYTICS-ENGINE] Creada grilla base con {len(heatmap_df)} coordenadas")
        
        # Calcular tiempo de transito (eventos agent_moved)
        move_events = self.events_df[self.events_df['tipo'] == 'agent_moved']
        transito_count = 0
        
        for _, event in move_events.iterrows():
            if 'data' in event and event['data'] and 'position' in event['data']:
                try:
                    position = event['data']['position']
                    if isinstance(position, list) and len(position) >= 2:
                        x, y = int(position[0]), int(position[1])
                        # Verificar que las coordenadas esten dentro del rango valido
                        if 0 <= x < warehouse_width and 0 <= y < warehouse_height:
                            # Buscar la fila correspondiente e incrementar tiempo_transito
                            mask = (heatmap_df['x'] == x) & (heatmap_df['y'] == y)
                            heatmap_df.loc[mask, 'tiempo_transito'] += 1.0  # 1 unidad de tiempo por movimiento
                            transito_count += 1
                except (ValueError, TypeError, IndexError):
                    # Saltar eventos con coordenadas malformadas
                    continue
        
        print(f"[ANALYTICS-ENGINE] Procesados {transito_count} eventos de transito")
        
        # Calcular tiempo de trabajo (eventos task_completed)
        task_events = self.events_df[self.events_df['tipo'] == 'task_completed']
        trabajo_count = 0
        
        for _, event in task_events.iterrows():
            try:
                # Validar que event sea un diccionario antes de usar 'in'
                if not isinstance(event, dict):
                    continue
                    
                if 'data' in event and pd.notna(event['data']) and isinstance(event['data'], dict):
                    # Buscar coordenadas en task_ubicacion O location (ambos formatos de evento)
                    ubicacion = None
                    if 'task_ubicacion' in event['data']:
                        ubicacion = event['data']['task_ubicacion']
                    elif 'location' in event['data']:
                        ubicacion = event['data']['location']
                    
                    tiempo_picking = event['data'].get('tiempo_picking', 0)
                    
                    if ubicacion and isinstance(ubicacion, list) and len(ubicacion) >= 2 and tiempo_picking > 0:
                        x, y = int(ubicacion[0]), int(ubicacion[1])
                        # Verificar que las coordenadas esten dentro del rango valido
                        if 0 <= x < warehouse_width and 0 <= y < warehouse_height:
                            # Buscar la fila correspondiente e incrementar tiempo_trabajo
                            mask = (heatmap_df['x (coordenada)'] == x) & (heatmap_df['y (coordenada)'] == y)
                            heatmap_df.loc[mask, 'tiempo_trabajo (segundos)'] += tiempo_picking
                            trabajo_count += 1
            except (ValueError, TypeError, IndexError):
                # Saltar eventos con coordenadas o tiempo malformados
                continue
        
        print(f"[ANALYTICS-ENGINE] Procesados {trabajo_count} eventos de trabajo")
        
        # Calcular tiempo total (suma de trabajo y transito)
        heatmap_df['tiempo_total (segundos)'] = heatmap_df['tiempo_trabajo (segundos)'] + heatmap_df['tiempo_transito (segundos)']
        
        # Estadisticas de resumen
        total_tiempo_trabajo = heatmap_df['tiempo_trabajo (segundos)'].sum()
        total_tiempo_transito = heatmap_df['tiempo_transito (segundos)'].sum()
        coordenadas_activas = len(heatmap_df[heatmap_df['tiempo_total (segundos)'] > 0])
        
        print(f"[ANALYTICS-ENGINE] Heatmap calculado: {coordenadas_activas} coordenadas activas")
        print(f"[ANALYTICS-ENGINE] Tiempo total trabajo: {total_tiempo_trabajo:.2f}, transito: {total_tiempo_transito:.2f}")
        
        return heatmap_df
    
    def export_to_excel(self, filepath: str):
        """
        Exporta todos los DataFrames a un archivo Excel con multiples hojas.
        
        Args:
            filepath: Ruta del archivo Excel a crear
        """
        print(f"[ANALYTICS-ENGINE] Exportando reporte a: {filepath}")
        
        try:
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                
                # Hoja 1: Resumen Ejecutivo
                if self.summary_kpis is not None and not self.summary_kpis.empty:
                    self.summary_kpis.to_excel(writer, sheet_name='Resumen Ejecutivo', index=False)
                    print("[ANALYTICS-ENGINE] Hoja 'Resumen Ejecutivo' exportada")
                else:
                    # Crear hoja vacia si no hay datos
                    pd.DataFrame({'Mensaje': ['No hay datos de KPIs disponibles']}).to_excel(
                        writer, sheet_name='Resumen Ejecutivo', index=False)
                
                # Hoja 2: Rendimiento de Agentes
                if self.agent_performance is not None and not self.agent_performance.empty:
                    self.agent_performance.to_excel(writer, sheet_name='Rendimiento de Agentes', index=False)
                    print("[ANALYTICS-ENGINE] Hoja 'Rendimiento de Agentes' exportada")
                else:
                    # Crear hoja vacia si no hay datos
                    pd.DataFrame({'Mensaje': ['No hay datos de agentes disponibles']}).to_excel(
                        writer, sheet_name='Rendimiento de Agentes', index=False)
                
                # Hoja 3: Configuracion
                # Validar que config sea un diccionario antes de iterar
                if isinstance(self.config, dict) and self.config:
                    config_df = pd.DataFrame([
                        {'Parametro': key, 'Valor': str(value)}
                        for key, value in self.config.items()
                    ])
                else:
                    config_df = pd.DataFrame([{'Parametro': 'No config', 'Valor': 'N/A'}])
                config_df.to_excel(writer, sheet_name='Configuracion', index=False)
                print("[ANALYTICS-ENGINE] Hoja 'Configuracion' exportada")
                
                # Hoja 4: Datos de Heatmap
                if self.heatmap_data is not None and not self.heatmap_data.empty:
                    self.heatmap_data.to_excel(writer, sheet_name='HeatmapData', index=False)
                    print("[ANALYTICS-ENGINE] Hoja 'HeatmapData' exportada")
                else:
                    # Crear hoja vacia si no hay datos
                    pd.DataFrame({'Mensaje': ['No hay datos de heatmap disponibles']}).to_excel(
                        writer, sheet_name='HeatmapData', index=False)
                
                # Hoja 5: Visual Heatmap
                self._add_visual_heatmap_sheet(writer)
            
            print(f"[ANALYTICS-ENGINE] Reporte exportado exitosamente: {filepath}")
            return filepath
            
        except Exception as e:
            print(f"[ANALYTICS-ENGINE] ERROR al exportar: {e}")
            return None
    
    def export_to_json(self, filepath: str):
        """
        Exporta todos los datos a un archivo JSON con la misma información que el Excel.
        """
        print(f"[ANALYTICS-ENGINE] Exportando datos a JSON: {filepath}")
        
        try:
            # Crear estructura de datos equivalente al Excel
            json_data = {
                "metadata": {
                    "timestamp": pd.Timestamp.now().isoformat(),
                    "total_events": len(self.event_log) if self.event_log else 0,
                    "simulation_duration": self._get_simulation_duration()
                },
                "resumen_ejecutivo": {},
                "rendimiento_agentes": [],
                "configuracion": {},
                "heatmap_data": [],
                "visual_heatmap": {}
            }
            
            # Resumen Ejecutivo
            if self.summary_kpis is not None and not self.summary_kpis.empty:
                for _, row in self.summary_kpis.iterrows():
                    metrica = row.get('Métrica', row.get('Metrica', ''))
                    valor = row.get('Valor', 0)
                    json_data["resumen_ejecutivo"][metrica] = valor
            
            # Rendimiento de Agentes
            if self.agent_performance is not None and not self.agent_performance.empty:
                for _, row in self.agent_performance.iterrows():
                    agent_data = {}
                    for col in self.agent_performance.columns:
                        agent_data[col] = row[col]
                    json_data["rendimiento_agentes"].append(agent_data)
            
            # Configuración
            if isinstance(self.config, dict) and self.config:
                json_data["configuracion"] = self.config
            else:
                json_data["configuracion"] = {"No config": "N/A"}
            
            # Heatmap Data
            if self.heatmap_data is not None and not self.heatmap_data.empty:
                for _, row in self.heatmap_data.iterrows():
                    heatmap_row = {}
                    for col in self.heatmap_data.columns:
                        heatmap_row[col] = row[col]
                    json_data["heatmap_data"].append(heatmap_row)
            
            # Visual Heatmap (matriz simplificada)
            if hasattr(self, '_heatmap_matrix') and self._heatmap_matrix is not None:
                json_data["visual_heatmap"] = {
                    "dimensions": self._heatmap_matrix.shape,
                    "matrix": self._heatmap_matrix.tolist()
                }
            
            # Escribir archivo JSON
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, indent=2, ensure_ascii=False, default=str)
            
            print(f"[ANALYTICS-ENGINE] Datos JSON exportados exitosamente: {filepath}")
            return filepath
            
        except Exception as e:
            print(f"[ANALYTICS-ENGINE] ERROR al exportar JSON: {e}")
            return None
    
    def _get_simulation_duration(self):
        """
        Calcula la duración total de la simulación en segundos.
        """
        if self.events_df is not None and not self.events_df.empty:
            return self.events_df['timestamp'].max() - self.events_df['timestamp'].min()
        return 0.0
    
    def _add_visual_heatmap_sheet(self, writer):
        """
        Crea una hoja VisualHeatmap con representacion grafica del almacen coloreada
        segun intensidad de actividad usando formato condicional.
        
        Args:
            writer: pandas ExcelWriter object
        """
        print("[ANALYTICS-ENGINE] Creando hoja VisualHeatmap...")
        
        if self.heatmap_data is None or self.heatmap_data.empty:
            # Crear hoja vacia si no hay datos
            pd.DataFrame({'Mensaje': ['No hay datos de heatmap para visualizacion']}).to_excel(
                writer, sheet_name='VisualHeatmap', index=False)
            print("[ANALYTICS-ENGINE] Hoja 'VisualHeatmap' vacia creada")
            return
        
        # Pivotar datos de formato largo (x, y, tiempo_total) a matriz 2D
        # x sera las columnas, y sera las filas
        heatmap_matrix = self.heatmap_data.pivot(
            index='y (coordenada)',           # Filas (eje vertical)
            columns='x (coordenada)',         # Columnas (eje horizontal)
            values='tiempo_total (segundos)' # Valores de las celdas
        )
        
        # Ordenar indice y columnas para presentacion correcta del almacen
        # y=0 en la parte superior, x=0 en la izquierda
        heatmap_matrix = heatmap_matrix.sort_index(ascending=False)  # y invertido (arriba-abajo)
        heatmap_matrix = heatmap_matrix.sort_index(axis=1)           # x normal (izquierda-derecha)
        
        # Rellenar valores NaN con 0 (coordenadas sin actividad)
        heatmap_matrix = heatmap_matrix.fillna(0)
        
        print(f"[ANALYTICS-ENGINE] Matriz de heatmap creada: {heatmap_matrix.shape[0]}x{heatmap_matrix.shape[1]}")
        
        # Escribir matriz en nueva hoja sin indices ni cabeceras (puramente visual)
        heatmap_matrix.to_excel(
            writer, 
            sheet_name='VisualHeatmap', 
            index=False,    # Sin etiquetas de fila
            header=False    # Sin etiquetas de columna
        )
        
        # Obtener el objeto worksheet de openpyxl para aplicar formato condicional
        worksheet = writer.sheets['VisualHeatmap']
        
        # Calcular rango de celdas a formatear
        num_rows = heatmap_matrix.shape[0]
        num_cols = heatmap_matrix.shape[1]
        
        # Convertir a notacion de Excel (A1:ZZ99)
        from openpyxl.utils import get_column_letter
        start_cell = 'A1'
        end_col_letter = get_column_letter(num_cols)
        end_cell = f'{end_col_letter}{num_rows}'
        cell_range = f'{start_cell}:{end_cell}'
        
        print(f"[ANALYTICS-ENGINE] Aplicando formato condicional al rango: {cell_range}")
        
        # Crear regla de formato condicional: Verde-Amarillo-Rojo
        # Verde = baja actividad, Amarillo = media, Rojo = alta actividad
        color_scale = ColorScaleRule(
            start_type='min',        # Valores minimos
            start_color='00FF00',    # Verde para baja actividad
            mid_type='percentile',   # Percentil 50 (mediana)
            mid_value=50,
            mid_color='FFFF00',      # Amarillo para actividad media
            end_type='max',          # Valores maximos  
            end_color='FF0000'       # Rojo para alta actividad
        )
        
        # Aplicar la regla al rango de datos
        worksheet.conditional_formatting.add(cell_range, color_scale)
        
        # Ajustar ancho de columnas para mejor visualizacion
        for col in range(1, num_cols + 1):
            col_letter = get_column_letter(col)
            worksheet.column_dimensions[col_letter].width = 3  # Celdas estrechas tipo pixel
        
        # Ajustar alto de filas
        for row in range(1, num_rows + 1):
            worksheet.row_dimensions[row].height = 15  # Filas compactas
        
        print(f"[ANALYTICS-ENGINE] Hoja 'VisualHeatmap' creada con formato condicional aplicado")


def main():
    """Funcion principal para ejecucion independiente del AnalyticsEngine"""
    parser = argparse.ArgumentParser(
        description='Motor de Analiticas para Simulador de Almacen - Procesamiento independiente de eventos',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos de uso:
  python analytics_engine.py raw_events_20250826_011557.json report_output.xlsx
  python analytics_engine.py events.json analysis_report.xlsx
        """
    )
    
    parser.add_argument(
        'input_json', 
        help='Ruta al archivo raw_events.json de entrada'
    )
    parser.add_argument(
        'output_excel', 
        help='Ruta del archivo Excel de salida para el reporte'
    )
    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help='Mostrar informacion detallada del procesamiento'
    )
    
    args = parser.parse_args()
    
    try:
        print("=" * 70)
        print("ANALYTICS ENGINE - PROCESAMIENTO INDEPENDIENTE")
        print("=" * 70)
        
        # Cargar AnalyticsEngine desde archivo JSON
        engine = AnalyticsEngine.from_json_file(args.input_json)
        
        # Procesar eventos y calcular metricas
        engine.process_events()
        
        # Exportar reporte a Excel
        result = engine.export_to_excel(args.output_excel)
        
        if result:
            print("=" * 70)
            print("PROCESAMIENTO COMPLETADO EXITOSAMENTE")
            print(f"Reporte generado: {result}")
            print("=" * 70)
        else:
            print("ERROR: Fallo la exportacion del reporte")
            return 1
            
    except FileNotFoundError as e:
        print(f"ERROR: Archivo no encontrado - {e}")
        return 1
    except json.JSONDecodeError as e:
        print(f"ERROR: Archivo JSON malformado - {e}")
        return 1
    except Exception as e:
        print(f"ERROR INESPERADO: {e}")
        if args.verbose:
            import traceback
            traceback.print_exc()
        return 1
    
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())


print("[OK] Modulo 'analytics_engine' cargado - Motor de analiticas para reportes ejecutivos.")
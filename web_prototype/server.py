import sys
import os
import json
import math
import time
from typing import List, Dict, Any
from fastapi import FastAPI, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import uvicorn
import pytmx
import shutil
from pathlib import Path
from fastapi import FastAPI, HTTPException, UploadFile, File, WebSocket, WebSocketDisconnect

# Add project root to path to import existing modules if needed
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

# Import configuration manager
from web_prototype.config_manager import WebConfigurationManager
from web_prototype.simulation_runner import SimulationRunner

# Initialize configuration manager
PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
config_manager = WebConfigurationManager(PROJECT_ROOT)


# --- PYDANTIC MODELS ---

class ConfigData(BaseModel):
    """Model for configuration data"""
    config: Dict[str, Any]

class SaveConfigurationRequest(BaseModel):
    """Model for saving a configuration preset"""
    name: str
    description: str = ""
    config: Dict[str, Any]
    is_default: bool = False


app = FastAPI()

# CORS for development convenience
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# Anti-cache (desarrollo): evita que el navegador sirva versiones viejas de los estaticos
# (configurador/visor) y haya que forzar recarga (Ctrl+Shift+R). Aplica solo al front
# estatico; las APIs JSON ya son dinamicas y no se cachean igual.
@app.middleware("http")
async def add_no_cache_headers(request, call_next):
    response = await call_next(request)
    path = request.url.path
    if path == "/" or path.startswith("/web_configurator") or \
            path.endswith((".js", ".css", ".html", ".htm")):
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response

# --- DATA LOADING ---
TMX_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "data", "layouts", "layout_v2.tmx"))
# Using the replay JSONL file which contains the full simulation state including initial released WOs
REPLAY_FILE = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "output", "simulation_20251120_224852", "replay_20251120_224852.jsonl"))

class ReplayData:
    def __init__(self):
        self.events = []
        self.max_time = 0
        self.snapshots = {}  # {timestamp: state_dict}
        self.snapshot_interval = 60.0  # Create a snapshot every 60 seconds
        self.service_level = None  # INIT-5: resumen de nivel de servicio (backorders)
        self.load_data()
        self.precompute_snapshots()

    def precompute_snapshots(self):
        """
        Pre-computes state snapshots at PROGRESSIVE intervals to speed up seeking.
        Intervals decrease as simulation progresses to handle more events efficiently:
        - 0-5min: 60s intervals
        - 5-30min: 30s intervals  
        - 30min+: 15s intervals
        """
        print(f"Pre-computing progressive snapshots...")
        if not self.events:
            return

        # Initial state
        current_state = {
            "agents": {},
            "work_orders": {}
        }
        
        # Helper to determine next snapshot time based on current time
        def get_next_snapshot_time(current_t):
            if current_t < 300:  # 0-5 minutes
                return 60.0
            elif current_t < 1800:  # 5-30 minutes
                return 30.0
            else:  # 30+ minutes
                return 15.0
        
        # Process all events in order and save snapshots
        next_snapshot_time = 60.0  # First snapshot at 60s
        prev_time = None
        
        # We need to process events sequentially to build state
        for event in self.events:
            t = event.get('timestamp', 0)
            
            # Apply event to current_state
            self._apply_event_to_state(event, current_state)
            
            # Check if we just finished processing all t=0 events
            if prev_time == 0 and t > 0 and 0.0 not in self.snapshots:
                import copy
                self.snapshots[0.0] = copy.deepcopy(current_state)
            
            # If we passed a snapshot boundary, save current state
            while t >= next_snapshot_time and next_snapshot_time <= self.max_time:
                # Deep copy current state for the snapshot
                import copy
                self.snapshots[next_snapshot_time] = copy.deepcopy(current_state)
                print(f"  Snapshot at t={next_snapshot_time}s")
                
                # Calculate next interval based on current position
                interval = get_next_snapshot_time(next_snapshot_time)
                next_snapshot_time += interval
            
            prev_time = t
        
        # Save t=0 snapshot if we never exceeded t=0 (all events at t=0)
        if 0.0 not in self.snapshots:
            import copy
            self.snapshots[0.0] = copy.deepcopy(current_state)
            
        print(f"Created {len(self.snapshots)} progressive snapshots.")


    def _apply_event_to_state(self, event, state):
        """Helper to apply a single event to a state dict."""
        etype = event.get('event_type') or event.get('tipo') or event.get('type')
        data = event.get('data', {})
        agent_id = event.get('agent_id') or data.get('agent_id')
        
        if etype == 'agent_moved':
            if agent_id:
                if agent_id not in state['agents']:
                    state['agents'][agent_id] = {}
                pos = data.get('position') or [data.get('x', 0), data.get('y', 0)]
                state['agents'][agent_id]['position'] = pos
                state['agents'][agent_id]['type'] = data.get('agent_type', 'Unknown')
                state['agents'][agent_id]['status'] = data.get('status', 'idle')
                
        elif etype == 'estado_agente':
            if agent_id:
                if agent_id not in state['agents']:
                    state['agents'][agent_id] = {}
                pos = event.get('position') or data.get('position') or [0, 0]
                state['agents'][agent_id]['position'] = pos
                state['agents'][agent_id]['type'] = event.get('agent_type') or data.get('agent_type', 'Unknown')
                state['agents'][agent_id]['status'] = event.get('status') or data.get('status', 'idle')
                
        elif etype == 'work_order_update':
            wo_id = data.get('id') or event.get('id')
            if wo_id:
                # Use data if it has content, otherwise use the entire event (for synthetic events)
                state['work_orders'][wo_id] = data if data else event

        # F2.c: handlers de eventos de camion (outbound). Acumulan contadores en
        # state['outbound'] para que los snapshots y /api/snapshot los expongan.
        # Con outbound.enabled=false nunca se emiten estos tipos -> contadores = 0.
        elif etype == 'truck_arrived':
            ob = state.setdefault('outbound', {
                'trucks_dispatched': 0, 'pallets_shipped': 0, 'backlog': 0})
            ob['last_truck_id'] = event.get('truck_id')

        elif etype == 'truck_departed':
            ob = state.setdefault('outbound', {
                'trucks_dispatched': 0, 'pallets_shipped': 0, 'backlog': 0})
            if event.get('pallets_loaded', 0) > 0:
                ob['trucks_dispatched'] = ob.get('trucks_dispatched', 0) + 1
            ob['backlog'] = event.get('backlog', 0)

        elif etype == 'pallet_shipped':
            ob = state.setdefault('outbound', {
                'trucks_dispatched': 0, 'pallets_shipped': 0, 'backlog': 0})
            ob['pallets_shipped'] = ob.get('pallets_shipped', 0) + 1

    def load_data(self):
        print(f"Loading replay data from {REPLAY_FILE}...")
        try:
            self.events = []
            self.service_level = None
            with open(REPLAY_FILE, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        event = json.loads(line)
                        
                        # Handle SIMULATION_START to extract initial WOs
                        if event.get('type') == 'SIMULATION_START' or event.get('event_type') == 'SIMULATION_START':
                            # INIT-5: resumen de nivel de servicio (backorders) desde la metadata.
                            self.service_level = event.get('service_level')
                            # Extract initial WOs directly from event
                            initial_wos = event.get('initial_work_orders', [])
                            print(f"Found {len(initial_wos)} initial WOs in SIMULATION_START")
                            
                            # Create synthetic update events for these initial WOs so they appear in the timeline
                            timestamp = event.get('timestamp', 0.0)
                            for wo in initial_wos:
                                # Create a synthetic update event
                                synthetic_event = wo.copy()
                                synthetic_event['type'] = 'work_order_update'
                                synthetic_event['timestamp'] = timestamp
                                # Ensure status defaults to 'released' if missing
                                if 'status' not in synthetic_event:
                                    synthetic_event['status'] = 'released'
                                self.events.append(synthetic_event)
                            
                            # Also add the start event itself just in case
                            self.events.append(event)
                        else:
                            # Normal event
                            self.events.append(event)
                            
                    except json.JSONDecodeError:
                        continue
            
            # Sort events by timestamp
            self.events.sort(key=lambda x: x.get('timestamp', 0))
            
            if self.events:
                self.max_time = self.events[-1].get('timestamp', 0)
            print(f"Loaded {len(self.events)} events. Max time: {self.max_time}")
        except Exception as e:
            print(f"Error loading replay data: {e}")
            import traceback
            traceback.print_exc()

    def reload_data(self, new_file_path: str):
        """
        Reload replay data from a new JSONL file.
        Clears existing events and snapshots, then loads new data.
        """
        print(f"Reloading data from {new_file_path}...")
        
        # Clear existing data
        self.events = []
        self.snapshots = {}
        self.max_time = 0
        
        # Update file path
        global REPLAY_FILE
        REPLAY_FILE = new_file_path
        
        # Load new data
        self.load_data()
        self.precompute_snapshots()
        
        print(f"Reload complete. Loaded {len(self.events)} events, max time: {self.max_time}")

replay_data = ReplayData()

# --- API ENDPOINTS ---

# ========================================================================
# CONFIGURATOR ENDPOINTS
# ========================================================================

@app.get("/api/configurator/config")
def get_config():
    """Load current config.json"""
    try:
        config = config_manager.load_config()
        return {"success": True, "config": config}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/configurator/config")
def save_config(data: ConfigData):
    """Save/update config.json"""
    try:
        success, errors = config_manager.save_config(data.config)
        
        if success:
            return {"success": True, "message": "Configuration saved successfully"}
        else:
            return {"success": False, "errors": errors}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/configurator/work-areas")
def get_work_areas(sequence_file: str):
    """Extract work areas from sequence file"""
    try:
        work_areas = config_manager.extract_work_areas(sequence_file)
        return {"success": True, "work_areas": work_areas}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/configurator/configurations")
def list_configurations():
    """List all saved configuration presets"""
    try:
        configs = config_manager.list_configurations()
        return {"success": True, "configurations": configs}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/configurator/configurations")
def save_configuration(request: SaveConfigurationRequest):
    """Save a new configuration preset"""
    try:
        success, config_id, errors = config_manager.save_configuration(
            request.name,
            request.description,
            request.config,
            request.is_default
        )
        
        if success:
            return {
                "success": True,
                "config_id": config_id,
                "message": f"Configuration '{request.name}' saved successfully"
            }
        else:
            return {"success": False, "errors": errors}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/configurator/configurations/{config_id}")
def get_configuration(config_id: str):
    """Load a specific configuration preset"""
    try:
        config = config_manager.load_configuration(config_id)
        
        if config is None:
            raise HTTPException(status_code=404, detail="Configuration not found")
        
        return {"success": True, "config": config}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/configurator/configurations/{config_id}")
def delete_configuration(config_id: str):
    """Delete a configuration preset"""
    try:
        success, errors = config_manager.delete_configuration(config_id)
        
        if success:
            return {"success": True, "message": "Configuration deleted successfully"}
        else:
            return {"success": False, "errors": errors}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/configurator/configurations/{config_id}/set-default")
def set_default_configuration(config_id: str):
    """Set a configuration as the default"""
    try:
        success, errors = config_manager.set_default_configuration(config_id)
        
        if success:
            return {"success": True, "message": "Default configuration set successfully"}
        else:
            return {"success": False, "errors": errors}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/configurator/default")
def get_default_configuration():
    """Get the default configuration"""
    try:
        config = config_manager.get_default_configuration()
        
        if config is None:
            # Return hardcoded default if no default is set
            config = config_manager._get_default_config()
        
        return {"success": True, "config": config}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ========================================================================
# DETERMINISTIC ORDER FILE UPLOAD
# ========================================================================

class OrderFileUploadResponse(BaseModel):
    """Response model for order file upload"""
    success: bool
    message: str
    summary: Dict[str, Any] = {}
    exclusions: List[Dict[str, Any]] = []
    file_path: str = ""

@app.post("/api/upload-orders")
async def upload_orders_file(
    file: UploadFile = File(...),
    fulfillment_policy: str = "ship_partial"
):
    """
    Upload an order file (JSON or CSV) for deterministic simulation mode.
    
    Args:
        file: Order file (JSON or CSV format)
        fulfillment_policy: "ship_partial" or "fill_or_kill"
        
    Returns:
        Validation report with summary and exclusions
    """
    import json
    import csv
    from io import StringIO
    
    try:
        # Validate file extension
        filename = file.filename.lower()
        if not (filename.endswith('.json') or filename.endswith('.csv')):
            raise HTTPException(
                status_code=400,
                detail="Solo se permiten archivos .json o .csv"
            )
        
        # Read file content
        content = await file.read()
        content_str = content.decode('utf-8')
        
        # Parse based on file type
        orders = []
        total_items = 0
        
        if filename.endswith('.json'):
            # Parse JSON
            try:
                data = json.loads(content_str)
                orders_data = data.get('orders', data) if isinstance(data, dict) else data
                
                if not isinstance(orders_data, list):
                    raise HTTPException(status_code=400, detail="JSON debe contener un array 'orders'")
                
                for order in orders_data:
                    items = order.get('items', [])
                    orders.append({
                        'order_id': order.get('order_id', ''),
                        'items': items,
                        'staging_id': order.get('staging_id')
                    })
                    total_items += len(items)
                    
            except json.JSONDecodeError as e:
                raise HTTPException(status_code=400, detail=f"Error parsing JSON: {str(e)}")
                
        else:  # CSV
            # Parse CSV with smart grouping by order_id
            try:
                orders_dict = {}
                reader = csv.DictReader(StringIO(content_str))
                
                for row in reader:
                    order_id = row.get('order_id', '').strip()
                    if not order_id:
                        continue
                    
                    if order_id not in orders_dict:
                        staging_id = row.get('staging_id', '').strip()
                        orders_dict[order_id] = {
                            'order_id': order_id,
                            'items': [],
                            'staging_id': int(staging_id) if staging_id else None
                        }
                    
                    sku_id = row.get('sku_id', '').strip()
                    quantity_str = row.get('quantity', '1').strip()
                    
                    if sku_id:
                        try:
                            quantity = int(quantity_str) if quantity_str else 1
                        except ValueError:
                            quantity = 1
                        
                        orders_dict[order_id]['items'].append({
                            'sku_id': sku_id,
                            'quantity': quantity,
                            'work_area': row.get('work_area', '').strip() or None
                        })
                        total_items += 1
                
                orders = list(orders_dict.values())
                
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Error parsing CSV: {str(e)}")
        
        # Validate orders exist
        if not orders:
            raise HTTPException(status_code=400, detail="No se encontraron ordenes válidas en el archivo")
        
        # Save file for later use by simulation
        uploads_dir = os.path.join(PROJECT_ROOT, "uploads")
        os.makedirs(uploads_dir, exist_ok=True)
        
        saved_file_path = os.path.join(uploads_dir, f"orders_{file.filename}")
        with open(saved_file_path, 'wb') as f:
            f.write(content)
        
        # Load current catalog to validate SKUs - NOW USES SQLite!
        catalog_skus = set()
        try:
            import sqlite3
            db_path = os.path.join(PROJECT_ROOT, 'warehouse.db')
            
            if os.path.exists(db_path):
                # USE SQLITE - Fast and uses real SKUs
                conn = sqlite3.connect(db_path)
                cursor = conn.execute("SELECT sku_code FROM sku_catalog")
                catalog_skus = set(row[0] for row in cursor.fetchall())
                conn.close()
                print(f"[UPLOAD-ORDERS] Loaded {len(catalog_skus)} SKUs from SQLite")
            else:
                # FALLBACK: Read from Excel (legacy)
                print("[UPLOAD-ORDERS] No SQLite DB, using Excel fallback for SKU validation")
                try:
                    import openpyxl
                    config = config_manager.load_config()
                    sequence_file = config.get('sequence_file', 'layouts/Warehouse_Logic.xlsx')
                    excel_path = os.path.join(PROJECT_ROOT, sequence_file)
                    
                    if os.path.exists(excel_path):
                        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
                        if 'PickingLocations' in wb.sheetnames:
                            sheet = wb['PickingLocations']
                            # Find sku_initial column
                            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
                            sku_col = None
                            for i, h in enumerate(headers):
                                if h and 'sku' in str(h).lower():
                                    sku_col = i
                                    break
                            
                            if sku_col is not None:
                                for row in sheet.iter_rows(min_row=2, values_only=True):
                                    if row and sku_col < len(row) and row[sku_col]:
                                        catalog_skus.add(str(row[sku_col]))
                        wb.close()
                        print(f"[UPLOAD-ORDERS] Loaded {len(catalog_skus)} SKUs from Excel")
                except Exception as excel_err:
                    print(f"[UPLOAD-ORDERS] Excel fallback failed: {excel_err}")
        except Exception as db_err:
            print(f"[UPLOAD-ORDERS] SKU catalog load error: {db_err}")
        
        # Validate orders against catalog
        valid_orders = 0
        valid_items = 0
        exclusions = []
        skus_found = set()
        skus_missing = set()
        
        for order in orders:
            order_has_valid_items = False
            
            for item in order['items']:
                sku_id = item.get('sku_id', '')
                
                if catalog_skus and sku_id not in catalog_skus:
                    skus_missing.add(sku_id)
                    exclusions.append({
                        'order_id': order['order_id'],
                        'item_sku': sku_id,
                        'reason': f"SKU '{sku_id}' no encontrado en catálogo",
                        'action': 'excluded'
                    })
                else:
                    skus_found.add(sku_id)
                    valid_items += 1
                    order_has_valid_items = True
            
            if order_has_valid_items or not catalog_skus:
                valid_orders += 1
        
        # Apply policy logic for summary
        if fulfillment_policy == 'fill_or_kill':
            orders_with_exclusions = set(e['order_id'] for e in exclusions)
            final_orders = valid_orders - len(orders_with_exclusions)
        else:
            final_orders = valid_orders
        
        return {
            "success": True,
            "message": f"Archivo '{file.filename}' procesado exitosamente",
            "file_path": saved_file_path,
            "summary": {
                "total_orders_input": len(orders),
                "total_items_input": total_items,
                "total_orders_output": final_orders,
                "total_items_output": valid_items,
                "skus_found": len(skus_found),
                "skus_missing": list(skus_missing),
                "fulfillment_policy": fulfillment_policy
            },
            "exclusions": exclusions[:20]  # Limit to first 20 for UI
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error processing order file: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# ========================================================================
# SIMULATION/REPLAY ENDPOINTS
# ========================================================================

@app.get("/api/layout")
def get_layout():
    """Parses TMX file and returns geometry for rendering.

    FIX: resuelve el mapa desde la CONFIGURACION (config.json -> layout_file), que es
    el MISMO mapa que usa la simulacion. Antes usaba una ruta fija
    (data/layouts/layout_v2.tmx) que no existe -> caia al mapa viejo (30x30) y no
    coincidia con la simulacion (el v2 es 30x42), por eso el visor "no se veia bien".
    Orden: layout_file del config -> ruta historica -> WH1.tmx.
    """
    candidates = []
    try:
        cfg = config_manager.load_config()
        layout_rel = cfg.get('layout_file')
        if layout_rel:
            candidates.append(os.path.join(PROJECT_ROOT, layout_rel))
    except Exception as e:
        print(f"[API/LAYOUT] No se pudo leer layout_file del config: {e}")
    candidates.append(TMX_PATH)  # ruta historica (compat)
    candidates.append(os.path.join(PROJECT_ROOT, "layouts", "WH1.tmx"))  # fallback final
    tmx_file = next((p for p in candidates if p and os.path.exists(p)), None)
    if tmx_file is None:
        raise HTTPException(status_code=404, detail=f"TMX no encontrado. Probados: {candidates}")
    print(f"[API/LAYOUT] Sirviendo mapa: {tmx_file}")
    tm = pytmx.TiledMap(tmx_file)

    try:
        layers = []
        
        # Extract walls and zones
        for layer in tm.layers:
            if hasattr(layer, 'data'):
                layer_data = []
                for y, row in enumerate(layer.data):
                    for x, gid in enumerate(row):
                        if gid != 0:
                            # Simple mapping for prototype
                            # In real app, we'd map GID to properties
                            layer_data.append({"x": x, "y": y, "gid": gid})
                layers.append({"name": layer.name, "tiles": layer_data})
        
        return {
            "width": tm.width,
            "height": tm.height,
            "tile_width": tm.tilewidth,
            "tile_height": tm.tileheight,
            "layers": layers
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/snapshot")
def get_snapshot(t: float):
    """
    UNIFIED ENDPOINT: Returns both state and metrics in a single request.
    This reduces HTTP traffic by 50% compared to separate /api/state + /api/metrics calls.
    """
    # OPTIMIZATION: Use nearest snapshot
    snapshot_times = [st for st in replay_data.snapshots.keys() if st <= t]
    
    if snapshot_times:
        base_time = max(snapshot_times)
        # Deep copy to avoid modifying the snapshot
        import copy
        current_state = copy.deepcopy(replay_data.snapshots[base_time])
        # Only process events AFTER the snapshot
        relevant_events = [e for e in replay_data.events if base_time < e.get('timestamp', 0) <= t]
    else:
        current_state = {
            "agents": {},
            "work_orders": {}
        }
        relevant_events = [e for e in replay_data.events if e.get('timestamp', 0) <= t]
    
    for event in relevant_events:
        replay_data._apply_event_to_state(event, current_state)
    
    # Compute work_orders_asignadas for each agent
    # Based on desktop implementation in replay_engine.py
    for agent_id in current_state['agents']:
        assigned_wos = []
        for wo_id, wo in current_state['work_orders'].items():
            # Check if WO is assigned to this agent
            wo_agent = wo.get('assigned_agent_id') or wo.get('agent_id')
            wo_status = wo.get('status', 'released')
            
            # Normalize agent ID matching - WO agents are like "Forklift_Forklift-01" 
            # but state agents are "Forklift-01"
            # Extract the last part after underscore if present
            if wo_agent:
                if '_' in wo_agent:
                    wo_agent_short = wo_agent.split('_')[-1]
                else:
                    wo_agent_short = wo_agent
            else:
                wo_agent_short = None
            
            # Only include WOs that are PENDING (not picked or staged)
            # picked = completed picking, staged = delivered to staging
            pending_statuses = ['assigned', 'in_progress']
            if wo_agent_short == agent_id and wo_status in pending_statuses:
                assigned_wos.append({
                    'id': wo_id,
                    'location': wo.get('location') or wo.get('ubicacion', [0, 0]),
                    'status': wo_status,
                    'pick_sequence': wo.get('pick_sequence', 0)
                })
        
        # Sort by pick_sequence to maintain tour order
        assigned_wos.sort(key=lambda x: x['pick_sequence'])
        current_state['agents'][agent_id]['work_orders_asignadas'] = assigned_wos

    # Compute metrics (integrated from get_metrics)
    work_orders = current_state['work_orders']
    agents = current_state['agents']
    
    # Helper lambdas for status checks
    get_status = lambda wo: wo.get('status', 'released')
    is_released = lambda wo: get_status(wo) == 'released' or (get_status(wo) == 'assigned' and not wo.get('assigned_agent_id'))
    is_assigned = lambda wo: get_status(wo) == 'assigned' and wo.get('assigned_agent_id')
    
    # Count work orders by status
    wo_total = len(work_orders)
    wo_completed = sum(1 for wo in work_orders.values() if wo.get('status') == 'staged')
    
    # Count agents by status
    agent_total = len(agents)
    agent_active = sum(1 for agent in agents.values() if agent.get('status') in ['moving', 'picking', 'unloading'])
    agent_idle = sum(1 for agent in agents.values() if agent.get('status') == 'idle')
    
    # Calculate throughput (WOs per minute)
    if t > 0:
        throughput = (wo_completed / t) * 60.0
    else:
        throughput = 0.0
    
    # Count by agent type
    ground_operators = sum(1 for agent in agents.values() if 'Ground' in agent.get('type', ''))
    forklifts = sum(1 for agent in agents.values() if 'Forklift' in agent.get('type', ''))
    
    # === NEW: Dashboard de Agentes Metrics ===
    # Calculate utilizacion_promedio (% of agents not idle)
    operarios_working = sum(1 for agent in agents.values() if agent.get('status') in ['working', 'picking', 'lifting', 'unloading'])
    operarios_traveling = sum(1 for agent in agents.values() if agent.get('status') in ['moving', 'traveling'])
    operarios_idle = sum(1 for agent in agents.values() if agent.get('status') in ['idle', 'Esperando tour'])
    
    if agent_total > 0:
        utilizacion_promedio = ((operarios_working + operarios_traveling) / agent_total) * 100.0
    else:
        utilizacion_promedio = 0.0
    
    # Calculate WIP (Work In Progress = total - completed)
    wip = max(wo_total - wo_completed, 0)
    
    # Calculate tareas_completadas (tasks completed = wo_completed * 3)
    # Each WorkOrder has 3 tasks in the desktop version
    tareas_completadas = wo_completed * 3
    
    # Add cargo_volume to each agent for load bar visualization
    # Note: This would normally come from the event data, but we'll set a placeholder
    for agent_id, agent_data in agents.items():
        # If cargo_volume not already in agent data, initialize it
        if 'cargo_volume' not in agent_data:
            agent_data['cargo_volume'] = 0
        # capacidad would also come from agent config, set default
        if 'capacidad' not in agent_data:
            # Default capacity: GroundOperator = 100, Forklift = 200
            agent_type = agent_data.get('type', '')
            agent_data['capacidad'] = 200 if 'Forklift' in agent_type else 100

    # Return unified response with both state and metrics
    return {
        "timestamp": t,
        "max_time": replay_data.max_time,
        "service_level": replay_data.service_level,  # INIT-5: nivel de servicio (backorders)
        "state": {
            "agents": agents,  # Now includes cargo_volume and capacidad
            "work_orders": current_state['work_orders']
        },
        "metricas": {
            # Core metrics (matching desktop estado_visual["metricas"])
            "tiempo": t,
            "workorders_completadas": wo_completed,
            "total_wos": wo_total,
            "tareas_completadas": tareas_completadas,
            "operarios_idle": operarios_idle,
            "operarios_working": operarios_working,
            "operarios_traveling": operarios_traveling,
            "utilizacion_promedio": round(utilizacion_promedio, 1),
            "wip": wip,
            # Throughput (WOs per minute) - CRITICAL: Division by zero protection
            "throughput_min": round(throughput, 2) if t > 0 else 0.0,
        },
        # Legacy metrics structure (keep for backward compatibility)
        "metrics": {
            "simulation_time": t,
            "work_orders": {
                "total": wo_total,
                "staged": wo_completed,
                "picked": sum(1 for wo in work_orders.values() if get_status(wo) == 'picked'),
                "in_progress": sum(1 for wo in work_orders.values() if get_status(wo) == 'in_progress'),
                "assigned": sum(1 for wo in work_orders.values() if is_assigned(wo)),
                "released": sum(1 for wo in work_orders.values() if is_released(wo)),
                "completion_rate": (wo_completed / wo_total * 100) if wo_total > 0 else 0
            },
            "agents": {
                "total": agent_total,
                "active": agent_active,
                "idle": agent_idle,
                "ground_operators": ground_operators,
                "forklifts": forklifts
            },
            "performance": {
                "throughput_per_minute": round(throughput, 2),
                "avg_time_per_wo": round(t / wo_completed, 2) if wo_completed > 0 else 0
            },
            # F2.c: contadores de camion acumulados hasta t (0 si outbound off)
            "outbound": {
                "trucks_dispatched": current_state.get('outbound', {}).get('trucks_dispatched', 0),
                "pallets_shipped": current_state.get('outbound', {}).get('pallets_shipped', 0),
                "backlog": current_state.get('outbound', {}).get('backlog', 0),
            },
            "service_level": replay_data.service_level  # INIT-5: nivel de servicio (backorders)
        }
    }


@app.get("/api/state")
def get_state(t: float):
    """Returns the state of the world at timestamp t."""
    # Naive implementation: Replay all events from 0 to t
    # In production, we would use snapshots/keyframes for performance
    
    # OPTIMIZATION: Use nearest snapshot
    snapshot_times = [st for st in replay_data.snapshots.keys() if st <= t]
    
    if snapshot_times:
        base_time = max(snapshot_times)
        # Deep copy to avoid modifying the snapshot
        import copy
        current_state = copy.deepcopy(replay_data.snapshots[base_time])
        # Only process events AFTER the snapshot
        relevant_events = [e for e in replay_data.events if base_time < e.get('timestamp', 0) <= t]
    else:
        current_state = {
            "agents": {},
            "work_orders": {}
        }
        relevant_events = [e for e in replay_data.events if e.get('timestamp', 0) <= t]
    
    for event in relevant_events:
        replay_data._apply_event_to_state(event, current_state)
    
    # Compute work_orders_asignadas for each agent
    # Based on desktop implementation in replay_engine.py
    for agent_id in current_state['agents']:
        assigned_wos = []
        for wo_id, wo in current_state['work_orders'].items():
            # Check if WO is assigned to this agent
            wo_agent = wo.get('assigned_agent_id') or wo.get('agent_id')
            wo_status = wo.get('status', 'released')
            
            # Normalize agent ID matching - WO agents are like "Forklift_Forklift-01" 
            # but state agents are "Forklift-01"
            # Extract the last part after underscore if present
            if wo_agent:
                if '_' in wo_agent:
                    wo_agent_short = wo_agent.split('_')[-1]
                else:
                    wo_agent_short = wo_agent
            else:
                wo_agent_short = None
            
            # Only include WOs that are PENDING (not picked or staged)
            # picked = completed picking, staged = delivered to staging
            pending_statuses = ['assigned', 'in_progress']
            if wo_agent_short == agent_id and wo_status in pending_statuses:
                assigned_wos.append({
                    'id': wo_id,
                    'location': wo.get('location') or wo.get('ubicacion', [0, 0]),
                    'status': wo_status,
                    'pick_sequence': wo.get('pick_sequence', 0)
                })
        
        # Sort by pick_sequence to maintain tour order
        assigned_wos.sort(key=lambda x: x['pick_sequence'])
        current_state['agents'][agent_id]['work_orders_asignadas'] = assigned_wos

    return {
        "timestamp": t,
        "max_time": replay_data.max_time,
        "service_level": replay_data.service_level,  # INIT-5: nivel de servicio (backorders)
        "agents": current_state['agents'],
        "work_orders": current_state['work_orders']
    }


@app.get("/api/metrics")
def get_metrics(t: float):
    """Returns comprehensive metrics at timestamp t."""
    # Reuse state calculation
    state_data = get_state(t)
    
    work_orders = state_data['work_orders']
    agents = state_data['agents']
    
    # Count work orders by status
    # ACTUAL statuses in log: 'assigned', 'in_progress', 'picked', 'staged'
    # 'staged' = completed (tiempo_fin is set)
    # 'picked' = picking done, awaiting staging
    # 'in_progress' = actively being picked
    # 'assigned' = pending
    wo_total = len(work_orders)
    wo_completed = sum(1 for wo in work_orders.values() if wo.get('status') == 'staged')
    wo_in_progress = sum(1 for wo in work_orders.values() if wo.get('status') in ['in_progress', 'picked'])
    wo_pending = sum(1 for wo in work_orders.values() if wo.get('status') == 'assigned')
    
    # Count agents by status
    agent_total = len(agents)
    agent_active = sum(1 for agent in agents.values() if agent.get('status') in ['moving', 'picking', 'unloading'])
    agent_idle = sum(1 for agent in agents.values() if agent.get('status') == 'idle')
    
    # Calculate throughput (WOs per minute)
    if t > 0:
        throughput = (wo_completed / t) * 60.0
    else:
        throughput = 0.0
    
    # Count by agent type
    ground_operators = sum(1 for agent in agents.values() if 'Ground' in agent.get('type', ''))
    forklifts = sum(1 for agent in agents.values() if 'Forklift' in agent.get('type', ''))
    
    # Helper lambdas for status checks
    # MATCH PYTHON DASHBOARD LOGIC: Status defaults to 'released' if missing
    get_status = lambda wo: wo.get('status', 'released')
    
    is_released = lambda wo: get_status(wo) == 'released' or (get_status(wo) == 'assigned' and not wo.get('assigned_agent_id'))
    is_assigned = lambda wo: get_status(wo) == 'assigned' and wo.get('assigned_agent_id')

    return {
        "simulation_time": t,
        "service_level": replay_data.service_level,  # INIT-5: nivel de servicio (backorders)
        "work_orders": {
            "total": wo_total,
            "staged": wo_completed,
            "picked": sum(1 for wo in work_orders.values() if get_status(wo) == 'picked'),
            "in_progress": sum(1 for wo in work_orders.values() if get_status(wo) == 'in_progress'),
            "assigned": sum(1 for wo in work_orders.values() if is_assigned(wo)),
            "released": sum(1 for wo in work_orders.values() if is_released(wo)),
            "completion_rate": (wo_completed / wo_total * 100) if wo_total > 0 else 0
        },
        "agents": {
            "total": agent_total,
            "active": agent_active,
            "idle": agent_idle,
            "ground_operators": ground_operators,
            "forklifts": forklifts
        },
        "performance": {
            "throughput_per_minute": round(throughput, 2),
            "avg_time_per_wo": round(t / wo_completed, 2) if wo_completed > 0 else 0
        }
    }




# ========================================================================
# UPLOAD & CLEANUP
# ========================================================================

@app.on_event("startup")
async def startup_event():
    """Clean up uploads directory on startup"""
    uploads_dir = os.path.join(PROJECT_ROOT, "uploads")
    if os.path.exists(uploads_dir):
        print(f"Cleaning up uploads directory: {uploads_dir}")
        try:
            shutil.rmtree(uploads_dir)
            os.makedirs(uploads_dir, exist_ok=True)
            print("Uploads directory cleaned.")
        except Exception as e:
            print(f"Error cleaning uploads directory: {e}")

@app.post("/api/upload_replay")
async def upload_replay_file(file: UploadFile = File(...)):
    """
    Upload a new JSONL replay file and reload the simulation.
    
    Returns:
        - success: bool
        - message: str
        - max_time: float (if successful)
    """
    try:
        # Validate file extension
        if not file.filename.endswith('.jsonl'):
            raise HTTPException(
                status_code=400, 
                detail="Solo se permiten archivos .jsonl"
            )
        
        # Create uploads directory if it doesn't exist
        uploads_dir = os.path.join(PROJECT_ROOT, "uploads")
        os.makedirs(uploads_dir, exist_ok=True)
        
        # Save file to uploads directory
        file_path = os.path.join(uploads_dir, f"uploaded_{file.filename}")
        
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        print(f"File saved to: {file_path}")
        
        # Reload replay data with new file
        replay_data.reload_data(file_path)
        
        return {
            "success": True,
            "message": f"Archivo '{file.filename}' cargado exitosamente",
            "max_time": replay_data.max_time,
            "event_count": len(replay_data.events)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error uploading file: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))



# ========================================================================
# SIMULATION RUNNER ENDPOINTS
# ========================================================================

@app.websocket("/ws/simulation-runner")
async def websocket_simulation_runner(websocket: WebSocket):
    await websocket.accept()
    runner = SimulationRunner()
    
    try:
        # Check if simulation is already running
        if runner.is_running():
            await websocket.send_json({
                "type": "error",
                "message": "System busy: A simulation is already running"
            })
            return
        
        # Wait for START command
        data = await websocket.receive_json()
        if data.get("command") != "START":
            await websocket.send_json({
                "type": "error", 
                "message": "Invalid command. Expected START."
            })
            return
        
        # Run simulation and stream events
        async for event in runner.run_simulation_async():
            await websocket.send_json(event)
            
    except WebSocketDisconnect:
        print("Client disconnected, cancelling simulation...")
        runner.cancel_current_simulation()
    except Exception as e:
        print(f"WebSocket error: {e}")
        try:
            await websocket.send_json({
                "type": "error",
                "message": f"Internal server error: {str(e)}"
            })
        except:
            pass
    finally:
        try:
            await websocket.close()
        except:
            pass


@app.get("/api/simulation-status")
def get_simulation_status():
    """Check if a simulation is currently running"""
    runner = SimulationRunner()
    return {
        "running": runner.is_running()
    }


@app.get("/api/validate-replay")
def validate_replay_file(file: str):
    """Validate that a replay file exists and is readable"""
    # Security check: prevent path traversal
    if ".." in file or file.startswith("/") or file.startswith("\\"):
         # Allow relative paths but ensure they are within project root
         pass
         
    file_path = os.path.join(PROJECT_ROOT, file)
    
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Replay file not found")
    
    if not file.endswith('.jsonl'):
        raise HTTPException(status_code=400, detail="Invalid file format")
    
    return {"valid": True, "path": file}


@app.post("/api/load_replay")
def load_replay_file(file: str):
    """Load an existing replay file from the server"""
    # Security check: prevent path traversal
    if ".." in file or file.startswith("/") or file.startswith("\\"):
         # Allow relative paths but ensure they are within project root
         pass
         
    file_path = os.path.join(PROJECT_ROOT, file)
    
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Replay file not found")
    
    if not file.endswith('.jsonl'):
        raise HTTPException(status_code=400, detail="Invalid file format")
        
    try:
        print(f"Loading replay file: {file_path}")
        replay_data.reload_data(file_path)
        
        return {
            "success": True,
            "message": f"Archivo '{file}' cargado exitosamente",
            "max_time": replay_data.max_time,
            "event_count": len(replay_data.events)
        }
    except Exception as e:
        print(f"Error loading file: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ========================================================================
# SYSTEM CONTROL ENDPOINTS
# ========================================================================

@app.post("/api/system/restart")
def restart_server():
    """
    Trigger server restart by touching the server file.
    Requires reload=True in uvicorn.run() to work.
    """
    try:
        # Touch this file to trigger Uvicorn reload
        server_file = __file__
        os.utime(server_file, None)
        
        return {
            "success": True,
            "message": "Server restart triggered. Reloading...",
            "estimated_time": 3  # seconds
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to trigger restart: {str(e)}")


@app.get("/api/system/health")
def health_check():
    """
    Health check endpoint to verify server is responsive.
    Used by frontend to detect when server has restarted.
    """
    return {
        "status": "ok",
        "timestamp": time.time(),
        "uptime": time.time() - replay_data.events[0].get('timestamp', 0) if replay_data.events else 0
    }


# Serve static files (Frontend)
app.mount("/", StaticFiles(directory=os.path.join(os.path.dirname(__file__), "static"), html=True), name="static")

if __name__ == "__main__":
    uvicorn.run(
        "web_prototype.server:app", 
        host="0.0.0.0", 
        port=8000,
        reload=True,  # Enable auto-reload for development
        reload_dirs=[PROJECT_ROOT]  # Watch project directory for changes
    )

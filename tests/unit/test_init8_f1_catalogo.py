# -*- coding: utf-8 -*-
"""
INIT-8 F1 / T8-xx: catalogo fisico por SKU (hoja SkuCatalog).

Pinea:
- La hoja canonica: 50 SKUs, clases validas, rangos fisicos coherentes por
  clase, todos los sku_code existen en PickingLocations.
- Importer: actualiza weight_kg y category en DB pero NO volume_m3
  (estrategia de baseline F1: el volumen se activa en F2 JUNTO con el
  modelo de tiempos -- docs/PLAN_INIT8_TIEMPOS.md).
- Fallback Excel del data_manager enriquece el catalogo en memoria.
- SKU del motor: atributos peso/clase con defaults neutros.
"""
import os
import sqlite3

import openpyxl
import pytest

from subsystems.simulation.warehouse import SKU

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
XLSX_PATH = os.path.join(PROJECT_ROOT, "layouts", "Warehouse_Logic.xlsx")

CLASES_VALIDAS = {"pequeno", "mediano", "voluminoso", "pesado", "extra_grande"}
# rangos por clase (vol_min, vol_max, peso_min, peso_max) del catalogo sintetico
RANGOS = {
    "pequeno":      (0.002, 0.008, 0.1, 0.8),
    "mediano":      (0.01, 0.04, 1.0, 8.0),
    "voluminoso":   (0.12, 0.35, 4.0, 12.0),
    "pesado":       (0.03, 0.09, 18.0, 45.0),
    "extra_grande": (0.60, 1.30, 50.0, 90.0),
}


def _leer_hoja():
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    try:
        assert "SkuCatalog" in wb.sheetnames
        rows = list(wb["SkuCatalog"].iter_rows(values_only=True))
        picking = list(wb["PickingLocations"].iter_rows(values_only=True))
    finally:
        wb.close()
    headers = [str(c).strip() for c in rows[0]]
    assert headers == ["sku_code", "volumen_m3", "peso_kg", "clase_manejo"]
    data = [dict(zip(headers, r)) for r in rows[1:] if r and r[0]]
    return data, picking


def test_t801_hoja_canonica_completa_y_coherente():
    data, picking = _leer_hoja()
    assert len(data) == 50

    p_headers = [str(c).strip() if c else "" for c in picking[0]]
    i_sku = p_headers.index("sku_initial")
    skus_reales = {str(r[i_sku]) for r in picking[1:] if r and r[i_sku]}

    for row in data:
        clase = row["clase_manejo"]
        assert clase in CLASES_VALIDAS, clase
        assert row["sku_code"] in skus_reales, row["sku_code"]
        vmin, vmax, pmin, pmax = RANGOS[clase]
        assert vmin <= row["volumen_m3"] <= vmax, (row["sku_code"], clase)
        assert pmin <= row["peso_kg"] <= pmax, (row["sku_code"], clase)

    # un refrigerador y una polera YA no son iguales en los datos
    voles = {r["volumen_m3"] for r in data}
    pesos = {r["peso_kg"] for r in data}
    assert len(voles) > 40 and len(pesos) > 40  # variedad real, no plana


def test_t802_importer_volumen_peso_clase_activos(tmp_path):
    """Contrato F2 (2026-07-11): el importer trae los TRES atributos fisicos.
    El volumen real fue LA actualizacion intencional de baseline de INIT-8
    (fluye a SKU.volumen => capacidad/tours), junto al modelo de tiempos."""
    from subsystems.database.database_manager import DatabaseManager
    from subsystems.database.importer import import_warehouse_data

    db_path = str(tmp_path / "t8.db")
    try:
        result = import_warehouse_data(XLSX_PATH, db_path=db_path)
        assert result.success, result.errors

        conn = sqlite3.connect(db_path)
        try:
            n_con_peso = conn.execute(
                "SELECT COUNT(*) FROM sku_catalog WHERE weight_kg IS NOT NULL"
            ).fetchone()[0]
            n_clases = conn.execute(
                "SELECT COUNT(DISTINCT category) FROM sku_catalog").fetchone()[0]
            vol_distintos = conn.execute(
                "SELECT COUNT(DISTINCT volume_m3) FROM sku_catalog").fetchone()[0]
            fridge = conn.execute(
                "SELECT weight_kg, volume_m3 FROM sku_catalog "
                "WHERE category = 'extra_grande' LIMIT 1").fetchone()
        finally:
            conn.close()

        assert n_con_peso == 50           # peso importado
        assert n_clases == 5              # las 5 clases de manejo
        assert vol_distintos > 40         # volumen REAL activo (F2)
        assert fridge[0] >= 50.0          # linea blanca pesa de verdad
        assert fridge[1] >= 0.60          # y ocupa de verdad
    finally:
        DatabaseManager.reset_instance()


def test_t803_fallback_excel_enriquece_catalogo():
    from subsystems.simulation.data_manager import DataManager

    dm = object.__new__(DataManager)  # sin __init__: no TMX ni pygame
    dm.sku_catalog = {
        "SKU001": {"sku_code": "SKU001", "volume_m3": 0.01,
                   "weight_kg": None, "category": "GENERAL",
                   "equipment_required": "GroundOperator"},
    }
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    try:
        dm._process_sku_catalog_attrs(wb["SkuCatalog"])
    finally:
        wb.close()

    e = dm.sku_catalog["SKU001"]
    assert e["weight_kg"] is not None and e["weight_kg"] > 0
    assert e["category"] in CLASES_VALIDAS
    assert e["volume_m3"] != 0.01  # F2: el fallback tambien activa el volumen real


def test_t804_sku_atributos_fisicos_con_defaults_neutros():
    legacy = SKU("SKU-X", volumen=1)
    assert legacy.peso == 0.0 and legacy.clase == "GENERAL"  # neutro

    fridge = SKU("SKU-R", volumen=120, peso=85.5, clase="extra_grande")
    assert fridge.peso == 85.5 and fridge.clase == "extra_grande"
    assert "extra_grande" in repr(fridge)

# -*- coding: utf-8 -*-
"""
Analytics Engine para generar reportes ejecutivos en Excel.
Procesa eventos de simulación y calcula métricas clave de rendimiento.
"""

import pandas as pd
from typing import List, Dict, Any
import json
import argparse
from typing import List, Dict
from datetime import datetime, timedelta
from openpyxl.formatting.rule import ColorScaleRule


class AnalyticsEngine:
    """Motor de analíticas para generar reportes ejecutivos a partir de eventos de simulación"""
    
    def __init__(self, event_log: List[Dict[str, Any]], config: Dict[str, Any], warehouse_width: int = None, warehouse_height: int = None):
        """
        Initialize the analytics engine.

        Args:
            event_log: A list of event dictionaries from the simulation.
            config: The simulation configuration.
            warehouse_width: The width of the warehouse grid.
            warehouse_height: The height of the warehouse grid.
        """
        self.event_log = event_log
        self.config = config
        self.warehouse_width = warehouse_width
        self.warehouse_height = warehouse_height
    
    @classmethod
    def from_json_file(cls, json_filepath: str):
        """
        Constructor alternativo que carga eventos desde un archivo raw_events.json.
        
        Args:
            json_filepath: Ruta al archivo raw_events.json
            
        Returns:
            Nueva instancia de AnalyticsEngine cargada desde archivo
        """
        print(f"[ANALYTICS-ENGINE] Cargando eventos desde: {json_filepath}")
        
        try:
            with open(json_filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            if isinstance(data, dict):
                event_log = data.get('events', [])
                config = data.get('config', {})
            elif isinstance(data, list):
                event_log = data
                config = {}
                print("[ANALYTICS-ENGINE] ADVERTENCIA: No se encontró configuración en el archivo JSON")
            else:
                raise ValueError("Formato de archivo JSON no reconocido")
            
            print(f"[ANALYTICS-ENGINE] Cargados {len(event_log)} eventos del archivo")
            return cls(event_log=event_log, config=config)
            
        except FileNotFoundError:
            print(f"[ANALYTICS-ENGINE] ERROR: Archivo no encontrado: {json_filepath}")
            raise
        except json.JSONDecodeError as e:
            print(f"[ANALYTICS-ENGINE] ERROR: Archivo JSON malformado: {e}")
            raise
        except Exception as e:
            print(f"[ANALYTICS-ENGINE] ERROR al cargar archivo: {e}")
            raise
    
    def process_events(self):
        """
        Método principal que procesa todos los eventos y calcula métricas.
        """
        print("[ANALYTICS-ENGINE] Procesando eventos...")
        
        if not self.event_log:
            print("[ANALYTICS-ENGINE] ADVERTENCIA: No hay eventos para procesar")
            return
        
        self.events_df = pd.DataFrame(self.event_log)
        print(f"[ANALYTICS-ENGINE] DataFrame creado con {len(self.events_df)} filas")
        
        self.summary_kpis = self._calculate_summary_kpis()
        self.agent_performance = self._calculate_agent_performance()
        self.heatmap_data = self._calculate_heatmap_data()
        
        print("[ANALYTICS-ENGINE] Procesamiento completado")
    
    def _calculate_summary_kpis(self) -> pd.DataFrame:
        """
        Calcula las métricas clave para la hoja "Resumen Ejecutivo".
        """
        print("[ANALYTICS-ENGINE] Calculando KPIs de resumen ejecutivo...")
        
        if self.events_df is None or self.events_df.empty:
            return pd.DataFrame()
        
        tiempo_inicio = self.events_df['timestamp'].min()
        tiempo_fin = self.events_df['timestamp'].max()
        tiempo_total_sim = tiempo_fin - tiempo_inicio
        tiempo_total_horas = tiempo_total_sim / 3600
        
        work_order_completed_events = self.events_df[self.events_df['tipo'] == 'work_order_completed']
        total_tareas_completadas = len(work_order_completed_events)
        
        if tiempo_total_horas > 0:
            productividad = total_tareas_completadas / tiempo_total_horas
        else:
            productividad = 0
        
        summary_data = {
            'Métrica': [
                'Tiempo Total de Simulación (horas)',
                'Total de Tareas Completadas',
                'Productividad (Tareas/Hora)',
                'Tiempo de Inicio (sim)',
                'Tiempo de Fin (sim)'
            ],
            'Valor': [
                round(tiempo_total_horas, 2),
                total_tareas_completadas,
                round(productividad, 2),
                round(tiempo_inicio, 2),
                round(tiempo_fin, 2)
            ]
        }
        
        summary_df = pd.DataFrame(summary_data)
        print(f"[ANALYTICS-ENGINE] KPIs calculados: {len(summary_df)} métricas")
        return summary_df
    
    def _calculate_agent_performance(self) -> pd.DataFrame:
        """
        Calcula métricas de rendimiento por agente usando eventos reales.
        """
        print("[ANALYTICS-ENGINE] Calculando rendimiento de agentes...")

        if self.events_df is None or self.events_df.empty:
            return pd.DataFrame()

        # Obtener todos los agentes únicos de todos los eventos
        all_agent_ids = set()
        
        # De eventos de estado de agente
        agent_events = self.events_df[self.events_df['tipo'] == 'estado_agente']
        if not agent_events.empty:
            all_agent_ids.update(agent_events['agent_id'].dropna().unique())
        
        # De eventos de work orders completadas
        wo_completed_events = self.events_df[self.events_df['tipo'] == 'work_order_completed']
        if not wo_completed_events.empty:
            all_agent_ids.update(wo_completed_events['agent_id'].dropna().unique())
        
        # De eventos de tareas completadas
        task_completed_events = self.events_df[self.events_df['tipo'] == 'task_completed']
        if not task_completed_events.empty:
            all_agent_ids.update(task_completed_events['agent_id'].dropna().unique())
        
        # Normalizar agent IDs para evitar duplicados
        # Convertir a lista y eliminar duplicados basándose en nombres cortos
        normalized_agents = {}
        for agent_id in all_agent_ids:
            # Verificar si es None o NaN
            if agent_id is None or pd.isna(agent_id):
                continue
            
            # Determinar el nombre corto del agente
            if agent_id.startswith("GroundOperator_"):
                short_name = agent_id.replace("GroundOperator_", "")
            elif agent_id.startswith("Forklift_"):
                short_name = agent_id.replace("Forklift_", "")
            else:
                short_name = agent_id
            
            # Usar el nombre corto como clave para evitar duplicados
            # Pero mantener el nombre largo como valor si está disponible
            if short_name not in normalized_agents:
                # Preferir el nombre largo si está disponible
                if agent_id.startswith("GroundOperator_") or agent_id.startswith("Forklift_"):
                    normalized_agents[short_name] = agent_id
                else:
                    normalized_agents[short_name] = agent_id
        
        # Usar solo los agent IDs únicos normalizados
        all_agent_ids = list(normalized_agents.values())
        print(f"[ANALYTICS-ENGINE] DEBUG: Agent IDs normalizados: {all_agent_ids}")
        
        # Verificar que no hay valores nulos en la lista final
        all_agent_ids = [aid for aid in all_agent_ids if aid is not None and not pd.isna(aid)]
        print(f"[ANALYTICS-ENGINE] DEBUG: Agent IDs finales: {all_agent_ids}")

        if not all_agent_ids:
            print("[ANALYTICS-ENGINE] No se encontraron agentes para procesar")
            return pd.DataFrame()

        # Calcular métricas por agente
        agent_performance = []

        for agent_id in all_agent_ids:
            # Verificar si es None o NaN
            if agent_id is None or pd.isna(agent_id):
                print(f"[ANALYTICS-ENGINE] DEBUG: Saltando agente nulo: {agent_id}")
                continue
                
            print(f"[ANALYTICS-ENGINE] DEBUG: Procesando agente: {agent_id}")
            try:
                # Normalizar nombres de agentes para buscar en todos los tipos de eventos
                # Los eventos tienen diferentes formatos de nombres:
                # estado_agente: "GroundOp-01", "Forklift-01"
                # work_order_completed: "GroundOperator_GroundOp-01", "Forklift_Forklift-01"
                # task_completed: "GroundOp-01", "Forklift-01"
                
                # Crear variaciones del nombre del agente para buscar
                agent_variations = [agent_id]
                
                # Si es GroundOperator_GroundOp-XX, también buscar GroundOp-XX
                if agent_id.startswith("GroundOperator_"):
                    short_name = agent_id.replace("GroundOperator_", "")
                    agent_variations.append(short_name)
                
                # Si es Forklift_Forklift-XX, también buscar Forklift-XX
                if agent_id.startswith("Forklift_"):
                    short_name = agent_id.replace("Forklift_", "")
                    agent_variations.append(short_name)
                
                # Si es GroundOp-XX, también buscar GroundOperator_GroundOp-XX
                if agent_id.startswith("GroundOp-"):
                    long_name = f"GroundOperator_{agent_id}"
                    agent_variations.append(long_name)
                
                # Si es Forklift-XX, también buscar Forklift_Forklift-XX
                if agent_id.startswith("Forklift-") and not agent_id.startswith("Forklift_Forklift-"):
                    long_name = f"Forklift_{agent_id}"
                    agent_variations.append(long_name)
                
                # Filtrar eventos usando todas las variaciones del nombre
                agent_events_filtered = agent_events[agent_events['agent_id'].isin(agent_variations)]
                wo_events_filtered = wo_completed_events[wo_completed_events['agent_id'].isin(agent_variations)]
                task_events_filtered = task_completed_events[task_completed_events['agent_id'].isin(agent_variations)]

                # Calcular métricas básicas
                total_events = len(agent_events_filtered)
                tareas_completadas = len(wo_events_filtered)
                
                # Determinar tipo de agente
                agent_type = "Desconocido"
                if not agent_events_filtered.empty:
                    # Los eventos de estado_agente tienen estructura plana
                    agent_type = agent_events_filtered['agent_type'].iloc[0] if 'agent_type' in agent_events_filtered.columns else "Desconocido"
                
                # NUEVA LÓGICA: Calcular tiempos basados en gaps entre eventos consecutivos
                tiempo_picking = 0
                tiempo_viaje = 0  # incluye moving + lifting
                tiempo_descarga = 0
                tiempo_ocioso = 0
                tiempo_total_activo = 0
                
                if not agent_events_filtered.empty and 'timestamp' in agent_events_filtered.columns:
                    # Ordenar eventos por timestamp
                    eventos_ordenados = agent_events_filtered.sort_values('timestamp').reset_index(drop=True)
                    
                    # Calcular tiempo total activo
                    tiempo_total_activo = eventos_ordenados['timestamp'].max() - eventos_ordenados['timestamp'].min()
                    
                    # Calcular tiempos por gaps entre eventos consecutivos
                    for i in range(len(eventos_ordenados) - 1):
                        evento_actual = eventos_ordenados.iloc[i]
                        evento_siguiente = eventos_ordenados.iloc[i + 1]
                        
                        # Obtener duración del gap
                        duracion = evento_siguiente['timestamp'] - evento_actual['timestamp']
                        
                        # Obtener estado del evento actual
                        # Los eventos de estado_agente tienen estructura plana, no anidada en 'data'
                        estado = evento_actual.get('status', 'unknown')
                        
                        # Mapear estado a tiempo correspondiente
                        if estado == "picking":
                            tiempo_picking += duracion
                        elif estado in ["moving", "lifting"]:
                            tiempo_viaje += duracion
                        elif estado == "unloading":
                            tiempo_descarga += duracion
                        elif estado == "idle":
                            tiempo_ocioso += duracion
                        # Ignorar "working" (estado instantáneo)
                
                # Validar que los tiempos sean consistentes
                tiempo_calculado_total = tiempo_picking + tiempo_viaje + tiempo_descarga + tiempo_ocioso
                if tiempo_calculado_total > 0 and abs(tiempo_calculado_total - tiempo_total_activo) > 1.0:
                    print(f"[ANALYTICS-ENGINE] ADVERTENCIA: Discrepancia en tiempos para {agent_id}")
                    print(f"  Tiempo total activo: {tiempo_total_activo}")
                    print(f"  Tiempo calculado: {tiempo_calculado_total}")
                    print(f"  Diferencia: {abs(tiempo_calculado_total - tiempo_total_activo)}")
                
                # Calcular utilización de capacidad basada en tiempos reales
                utilizacion_capacidad = 0
                if tiempo_total_activo > 0:
                    tiempo_productivo = tiempo_picking + tiempo_viaje + tiempo_descarga
                    utilizacion_capacidad = (tiempo_productivo / tiempo_total_activo) * 100

                print(f"[ANALYTICS-ENGINE] DEBUG: Métricas para {agent_id}:")
                print(f"  Tareas completadas: {tareas_completadas}")
                print(f"  Tiempo total activo: {tiempo_total_activo}")
                print(f"  Tiempo picking: {tiempo_picking}")
                print(f"  Tiempo viaje: {tiempo_viaje}")
                print(f"  Tiempo descarga: {tiempo_descarga}")
                print(f"  Tiempo ocioso: {tiempo_ocioso}")
                print(f"  Utilización: {utilizacion_capacidad}")

                agent_performance.append({
                    'Agent_ID': agent_id,
                    'Tipo_Agente': agent_type,
                    'Tareas_Completadas': tareas_completadas,
                    'Tiempo_Total_Activo': round(tiempo_total_activo, 2),
                    'Tiempo_Picking': round(tiempo_picking, 2),
                    'Tiempo_Viaje': round(tiempo_viaje, 2),
                    'Tiempo_Descarga': round(tiempo_descarga, 2),
                    'Tiempo_Ocioso': round(tiempo_ocioso, 2),
                    'Utilizacion_Capacidad_Promedio_Pct': round(utilizacion_capacidad, 2),
                    'Eventos_Totales': total_events
                })
                
            except Exception as e:
                print(f"[ANALYTICS-ENGINE] Error procesando agente {agent_id}: {e}")
                continue

        # Filtrar entradas con Agent_ID nulo antes de crear el DataFrame
        agent_performance = [entry for entry in agent_performance if entry.get('Agent_ID') is not None and not pd.isna(entry.get('Agent_ID'))]
        
        performance_df = pd.DataFrame(agent_performance)
        print(f"[ANALYTICS-ENGINE] Rendimiento calculado para {len(performance_df)} agentes")
        return performance_df
    
        def _calculate_heatmap_data(self) -> pd.DataFrame:
    
            """
    
            Calcula los datos necesarios para generar heatmaps de actividad en el almacén.
    
            """
    
            print("[ANALYTICS-ENGINE] Calculando datos de heatmap...")
    
            
    
            if self.events_df is None or self.events_df.empty:
    
                return pd.DataFrame()
    
            
    
            # Obtener dimensiones del layout TMX desde la configuración
    
            if self.warehouse_width is None or self.warehouse_height is None:
    
                try:
    
                    from src.subsystems.simulation.layout_manager import LayoutManager
    
                    import os
    
    
    
                    layout_file = self.config.get('layout_file', 'layouts/WH1.tmx')
    
                    
    
                    # Resolver la ruta del archivo TMX
    
                    project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    
                    tmx_path = os.path.join(project_root, layout_file)
    
    
    
                    if not os.path.exists(tmx_path):
    
                        # Fallback a la ruta original si no se encuentra en el proyecto
    
                        tmx_path = layout_file
    
    
    
                    print(f"[ANALYTICS-ENGINE] Cargando dimensiones desde TMX: {tmx_path}")
    
                    
    
                    # Inicializar pygame dummy para pytmx
    
                    import os
    
                    os.environ['SDL_VIDEODRIVER'] = 'dummy'
    
                    import pygame
    
                    pygame.display.init()
    
                    pygame.display.set_mode((1, 1), pygame.NOFRAME)
    
    
    
                    layout = LayoutManager(tmx_path)
    
                    self.warehouse_width = layout.grid_width
    
                    self.warehouse_height = layout.grid_height
    
                    
    
                except (ImportError, FileNotFoundError, RuntimeError, pygame.error) as e:
    
                    print(f"[ANALYTICS-ENGINE] ADVERTENCIA: No se pudo cargar el layout TMX ({e}). Usando dimensiones por defecto.")
    
                    self.warehouse_width = 50  # Ancho estándar del almacén
    
                    self.warehouse_height = 35  # Alto estándar del almacén
    
            
    
            print(f"[ANALYTICS-ENGINE] Dimensiones del almacen: {self.warehouse_width}x{self.warehouse_height}")
    
            
    
            # Crear una grilla de ceros con las dimensiones del almacen
    
            coordenadas = [(x, y) for x in range(self.warehouse_width) for y in range(self.warehouse_height)]
    
            heatmap_df = pd.DataFrame(coordenadas, columns=['x', 'y'])
    
            heatmap_df['tiempo_trabajo'] = 0.0
    
            heatmap_df['tiempo_transito'] = 0.0
    
            
    
            # Procesar eventos de movimiento (estado_agente) para tiempo de tránsito
    
            move_events = self.events_df[self.events_df['tipo'] == 'estado_agente']
    
            transito_count = 0
    
            
    
            for _, event in move_events.iterrows():
    
                try:
    
                    position = event.get('position')
    
                    if position is not None and isinstance(position, list) and len(position) >= 2:
    
                        x, y = int(position[0]), int(position[1])
    
                        if 0 <= x < self.warehouse_width and 0 <= y < self.warehouse_height:
    
                            mask = (heatmap_df['x'] == x) & (heatmap_df['y'] == y)
    
                            heatmap_df.loc[mask, 'tiempo_transito'] += 1.0
    
                            transito_count += 1
    
                except (ValueError, TypeError, IndexError, AttributeError):
    
                    continue
    
            
    
            print(f"[ANALYTICS-ENGINE] Procesados {transito_count} eventos de tránsito")
    
            
    
            # Crear un diccionario para mapear work_order_id a su ubicación
    
            work_order_locations = {}
    
            wo_update_events = self.events_df[self.events_df['tipo'] == 'work_order_update']
    
            for _, event in wo_update_events.iterrows():
    
                if event.get('id') and event.get('location'):
    
                    work_order_locations[event['id']] = event['location']
    
    
    
            picking_times = {}
    
            agent_states = {} # To track the last state of each agent
    
    
    
            for _, event in self.events_df.sort_values('timestamp').iterrows():
    
                if event['tipo'] == 'estado_agente':
    
                    agent_id = event.get('agent_id')
    
                    if agent_id:
    
                        last_state = agent_states.get(agent_id)
    
                        if last_state and last_state['status'] == 'picking':
    
                            task_id = last_state.get('current_task')
    
                            if task_id:
    
                                duration = event['timestamp'] - last_state['timestamp']
    
                                picking_times[task_id] = picking_times.get(task_id, 0) + duration
    
                        agent_states[agent_id] = event
    
    
    
            # Procesar eventos de tareas completadas para tiempo de trabajo
    
            task_events = self.events_df[self.events_df['tipo'] == 'task_completed']
    
            trabajo_count = 0
    
    
    
            for _, event in task_events.iterrows():
    
                try:
    
                    work_order_id = event.get('task_id')
    
                    if work_order_id in work_order_locations:
    
                        ubicacion = work_order_locations[work_order_id]
    
                        tiempo_picking = picking_times.get(work_order_id, 0)
    
    
    
                        if ubicacion and isinstance(ubicacion, list) and len(ubicacion) >= 2 and tiempo_picking > 0:
    
                            x, y = int(ubicacion[0]), int(ubicacion[1])
    
                            if 0 <= x < self.warehouse_width and 0 <= y < self.warehouse_height:
    
                                mask = (heatmap_df['x'] == x) & (heatmap_df['y'] == y)
    
                                heatmap_df.loc[mask, 'tiempo_trabajo'] += tiempo_picking
    
                                trabajo_count += 1
    
                except (ValueError, TypeError, IndexError, AttributeError):
    
                    continue
    
            
    
            print(f"[ANALYTICS-ENGINE] Procesados {trabajo_count} eventos de trabajo")
    
            
    
            heatmap_df['tiempo_total'] = heatmap_df['tiempo_trabajo'] + heatmap_df['tiempo_transito']
    
            
    
            total_tiempo_trabajo = heatmap_df['tiempo_trabajo'].sum()
    
            total_tiempo_transito = heatmap_df['tiempo_transito'].sum()
    
            coordenadas_activas = len(heatmap_df[heatmap_df['tiempo_total'] > 0])
    
            
    
            print(f"[ANALYTICS-ENGINE] Heatmap calculado: {coordenadas_activas} coordenadas activas")
    
            print(f"[ANALYTICS-ENGINE] Tiempo total trabajo: {total_tiempo_trabajo:.2f}, tránsito: {total_tiempo_transito:.2f}")
    
            
    
            return heatmap_df
    
    
    
    def export_to_excel(self, filepath: str):
        """
        Exporta todos los DataFrames a un archivo Excel con múltiples hojas.
        """
        print(f"[ANALYTICS-ENGINE] Exportando reporte a: {filepath}")
        
        try:
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                
                if self.summary_kpis is not None and not self.summary_kpis.empty:
                    self.summary_kpis.to_excel(writer, sheet_name='Resumen Ejecutivo', index=False)
                    print("[ANALYTICS-ENGINE] Hoja 'Resumen Ejecutivo' exportada")
                else:
                    pd.DataFrame({'Mensaje': ['No hay datos de KPIs disponibles']}).to_excel(
                        writer, sheet_name='Resumen Ejecutivo', index=False)
                
                if self.agent_performance is not None and not self.agent_performance.empty:
                    self.agent_performance.to_excel(writer, sheet_name='Rendimiento de Agentes', index=False)
                    print("[ANALYTICS-ENGINE] Hoja 'Rendimiento de Agentes' exportada")
                else:
                    pd.DataFrame({'Mensaje': ['No hay datos de agentes disponibles']}).to_excel(
                        writer, sheet_name='Rendimiento de Agentes', index=False)
                
                # Validar que config sea un diccionario antes de iterar
                if isinstance(self.config, dict) and self.config:
                    config_df = pd.DataFrame([
                        {'Parámetro': key, 'Valor': str(value)}
                        for key, value in self.config.items()
                    ])
                else:
                    config_df = pd.DataFrame([{'Parámetro': 'No config', 'Valor': 'N/A'}])
                config_df.to_excel(writer, sheet_name='Configuracion', index=False)
                print("[ANALYTICS-ENGINE] Hoja 'Configuracion' exportada")
                
                if self.heatmap_data is not None and not self.heatmap_data.empty:
                    self.heatmap_data.to_excel(writer, sheet_name='HeatmapData', index=False)
                    print("[ANALYTICS-ENGINE] Hoja 'HeatmapData' exportada")
                else:
                    pd.DataFrame({'Mensaje': ['No hay datos de heatmap disponibles']}).to_excel(
                        writer, sheet_name='HeatmapData', index=False)
                
                self._add_visual_heatmap_sheet(writer)
            
            print(f"[ANALYTICS-ENGINE] Reporte exportado exitosamente: {filepath}")
            return filepath
            
        except Exception as e:
            print(f"[ANALYTICS-ENGINE] ERROR al exportar: {e}")
            return None
    
    def export_to_json(self, filepath: str):
        """
        Exporta todos los datos a un archivo JSON con la misma información que el Excel.
        """
        print(f"[ANALYTICS-ENGINE] Exportando datos a JSON: {filepath}")
        
        try:
            # Crear estructura de datos equivalente al Excel
            json_data = {
                "metadata": {
                    "timestamp": pd.Timestamp.now().isoformat(),
                    "total_events": len(self.event_log) if self.event_log else 0,
                    "simulation_duration": self._get_simulation_duration()
                },
                "resumen_ejecutivo": {},
                "rendimiento_agentes": [],
                "configuracion": {},
                "heatmap_data": [],
                "visual_heatmap": {},
                "heatmap_csv": ""
            }
            
            # Resumen Ejecutivo
            if self.summary_kpis is not None and not self.summary_kpis.empty:
                for _, row in self.summary_kpis.iterrows():
                    metrica = row.get('Métrica', row.get('Metrica', ''))
                    valor = row.get('Valor', 0)
                    json_data["resumen_ejecutivo"][metrica] = valor
            
            # Rendimiento de Agentes
            if self.agent_performance is not None and not self.agent_performance.empty:
                json_data["rendimiento_agentes"] = self.agent_performance.to_dict(orient='records')
            
            # Configuración
            if isinstance(self.config, dict) and self.config:
                json_data["configuracion"] = self.config
            else:
                json_data["configuracion"] = {"No config": "N/A"}
            
            # Heatmap Data
            if self.heatmap_data is not None and not self.heatmap_data.empty:
                json_data["heatmap_data"] = self.heatmap_data.to_dict(orient='records')
            
            # Visual Heatmap (matriz simplificada)
            if self.heatmap_data is not None and not self.heatmap_data.empty:
                heatmap_matrix = self.heatmap_data.pivot(
                    index='y',
                    columns='x',
                    values='tiempo_total'
                ).fillna(0)
                json_data["visual_heatmap"] = {
                    "dimensions": heatmap_matrix.shape,
                    "matrix": heatmap_matrix.values.tolist()
                }
                
                # Convertir la matriz a formato CSV string
                csv_string = "\n".join([",".join(map(str, row)) for row in heatmap_matrix.values.tolist()])
                json_data["heatmap_csv"] = csv_string
            
            # Escribir archivo JSON
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, indent=2, ensure_ascii=False, default=str)
            
            print(f"[ANALYTICS-ENGINE] Datos JSON exportados exitosamente: {filepath}")
            return filepath
            
        except Exception as e:
            print(f"[ANALYTICS-ENGINE] ERROR al exportar JSON: {e}")
            return None
    
    def _get_simulation_duration(self):
        """Calcula la duración total de la simulación."""
        if not self.event_log:
            return 0
        
        timestamps = [event.get('timestamp', 0) for event in self.event_log if 'timestamp' in event]
        if not timestamps:
            return 0
        
        return max(timestamps) - min(timestamps)
    
    def _add_visual_heatmap_sheet(self, writer):
        """
        Crea una hoja VisualHeatmap con representación gráfica del almacén coloreada.
        """
        print("[ANALYTICS-ENGINE] Creando hoja VisualHeatmap...")
        
        if self.heatmap_data is None or self.heatmap_data.empty:
            pd.DataFrame({'Mensaje': ['No hay datos de heatmap para visualización']}).to_excel(
                writer, sheet_name='VisualHeatmap', index=False)
            print("[ANALYTICS-ENGINE] Hoja 'VisualHeatmap' vacía creada")
            return
        
        heatmap_matrix = self.heatmap_data.pivot(
            index='y',
            columns='x',
            values='tiempo_total'
        )
        
        heatmap_matrix = heatmap_matrix.sort_index(ascending=False)
        heatmap_matrix = heatmap_matrix.sort_index(axis=1)
        heatmap_matrix = heatmap_matrix.fillna(0)
        
        print(f"[ANALYTICS-ENGINE] Matriz de heatmap creada: {heatmap_matrix.shape[0]}x{heatmap_matrix.shape[1]}")
        
        heatmap_matrix.to_excel(
            writer, 
            sheet_name='VisualHeatmap', 
            index=False,
            header=False
        )
        
        worksheet = writer.sheets['VisualHeatmap']
        
        num_rows = heatmap_matrix.shape[0]
        num_cols = heatmap_matrix.shape[1]
        
        from openpyxl.utils import get_column_letter
        start_cell = 'A1'
        end_col_letter = get_column_letter(num_cols)
        end_cell = f'{end_col_letter}{num_rows}'
        cell_range = f'{start_cell}:{end_cell}'
        
        print(f"[ANALYTICS-ENGINE] Aplicando formato condicional al rango: {cell_range}")
        
        color_scale = ColorScaleRule(
            start_type='min', start_color='00FF00',
            mid_type='percentile', mid_value=50, mid_color='FFFF00',
            end_type='max', end_color='FF0000'
        )
        
        worksheet.conditional_formatting.add(cell_range, color_scale)
        
        for col in range(1, num_cols + 1):
            col_letter = get_column_letter(col)
            worksheet.column_dimensions[col_letter].width = 3
        
        for row in range(1, num_rows + 1):
            worksheet.row_dimensions[row].height = 15
        
        print(f"[ANALYTICS-ENGINE] Hoja 'VisualHeatmap' creada con formato condicional aplicado")

def main():
    """Función principal para ejecución independiente del AnalyticsEngine"""
    parser = argparse.ArgumentParser(
        description='Motor de Analíticas para Simulador de Almacén - Procesamiento independiente de eventos',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""Ejemplos de uso:\n  python analytics_engine.py raw_events_20250826_011557.json report_output.xlsx\n  python analytics_engine.py events.json analysis_report.xlsx"""
    )
    
    parser.add_argument('input_json', help='Ruta al archivo raw_events.json de entrada')
    parser.add_argument('output_excel', help='Ruta del archivo Excel de salida para el reporte')
    parser.add_argument('--verbose', '-v',
        action='store_true',
        help='Mostrar información detallada del procesamiento')
    
    args = parser.parse_args()
    
    try:
        print("=" * 70)
        print("ANALYTICS ENGINE - PROCESAMIENTO INDEPENDIENTE")
        print("=" * 70)
        
        engine = AnalyticsEngine.from_json_file(args.input_json)
        engine.process_events()
        result = engine.export_to_excel(args.output_excel)
        
        if result:
            print("=" * 70)
            print("PROCESAMIENTO COMPLETADO EXITOSAMENTE")
            print(f"Reporte generado: {result}")
            print("=" * 70)
        else:
            print("ERROR: Fallo la exportacion del reporte")
            return 1
            
    except FileNotFoundError as e:
        print(f"ERROR: Archivo no encontrado - {e}")
        return 1
    except json.JSONDecodeError as e:
        print(f"ERROR: Archivo JSON malformado - {e}")
        return 1
    except Exception as e:
        print(f"ERROR INESPERADO: {e}")
        if args.verbose:
            import traceback
            traceback.print_exc()
        return 1
    
    return 0

if __name__ == "__main__":
    import sys
    sys.exit(main())

print("[OK] Módulo 'analytics_engine' cargado - Motor de analíticas para reportes ejecutivos.")

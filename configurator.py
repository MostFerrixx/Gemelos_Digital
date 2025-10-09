# -*- coding: utf-8 -*-
"""
Configurador del Simulador de Almacen - Herramienta Autocontenida
Interfaz grafica independiente para configurar parametros de simulacion.
Version refactorizada sin dependencias de submodulos.
"""

import sys
import os
import json
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from PIL import Image, ImageTk, ImageDraw
import io


class ModernIconGenerator:
    """
    Generador de iconos vectoriales modernos para la interfaz.
    Crea iconos SVG profesionales usando Pillow.
    """
    
    def __init__(self):
        self.icon_size = 16
        self.colors = {
            'primary': '#2563eb',      # Azul moderno
            'secondary': '#64748b',   # Gris profesional
            'success': '#059669',      # Verde éxito
            'warning': '#d97706',      # Naranja advertencia
            'danger': '#dc2626',       # Rojo peligro
            'info': '#0891b2',         # Azul información
            'dark': '#1e293b',         # Gris oscuro
            'light': '#f8fafc'         # Blanco suave
        }
    
    def create_save_icon(self):
        """Crea icono de guardar moderno"""
        img = Image.new('RGBA', (self.icon_size, self.icon_size), (0, 0, 0, 0))
        draw = ImageDraw.Draw(img)
        
        # Disco de guardado moderno
        draw.rectangle([2, 4, 14, 12], fill=self.colors['primary'], outline=None)
        draw.rectangle([4, 2, 12, 6], fill=self.colors['light'], outline=None)
        draw.rectangle([6, 8, 10, 10], fill=self.colors['light'], outline=None)
        
        return ImageTk.PhotoImage(img)
    
    def create_load_icon(self):
        """Crea icono de cargar moderno"""
        img = Image.new('RGBA', (self.icon_size, self.icon_size), (0, 0, 0, 0))
        draw = ImageDraw.Draw(img)
        
        # Carpeta de carga moderna
        draw.rectangle([2, 6, 14, 12], fill=self.colors['success'], outline=None)
        draw.rectangle([2, 4, 8, 6], fill=self.colors['success'], outline=None)
        draw.rectangle([4, 8, 6, 10], fill=self.colors['light'], outline=None)
        draw.rectangle([8, 8, 10, 10], fill=self.colors['light'], outline=None)
        
        return ImageTk.PhotoImage(img)
    
    def create_manage_icon(self):
        """Crea icono de gestión moderno"""
        img = Image.new('RGBA', (self.icon_size, self.icon_size), (0, 0, 0, 0))
        draw = ImageDraw.Draw(img)
        
        # Engranaje moderno
        draw.rectangle([6, 2, 10, 4], fill=self.colors['secondary'], outline=None)
        draw.rectangle([6, 12, 10, 14], fill=self.colors['secondary'], outline=None)
        draw.rectangle([2, 6, 4, 10], fill=self.colors['secondary'], outline=None)
        draw.rectangle([12, 6, 14, 10], fill=self.colors['secondary'], outline=None)
        draw.rectangle([6, 6, 10, 10], fill=self.colors['primary'], outline=None)
        
        return ImageTk.PhotoImage(img)
    
    def create_default_icon(self):
        """Crea icono de configuración por defecto moderno"""
        img = Image.new('RGBA', (self.icon_size, self.icon_size), (0, 0, 0, 0))
        draw = ImageDraw.Draw(img)
        
        # Estrella moderna
        points = [(8, 2), (9, 6), (13, 6), (10, 9), (11, 13), (8, 10), (5, 13), (6, 9), (3, 6), (7, 6)]
        draw.polygon(points, fill=self.colors['warning'], outline=None)
        
        return ImageTk.PhotoImage(img)
    
    # Método create_exit_icon removido - usar solo la cruz superior derecha
    
    def create_use_icon(self):
        """Crea icono de usar/aplicar moderno"""
        img = Image.new('RGBA', (self.icon_size, self.icon_size), (0, 0, 0, 0))
        draw = ImageDraw.Draw(img)
        
        # Flecha hacia la derecha con círculo
        draw.ellipse([(2, 2), (14, 14)], fill=self.colors['success'], outline=None)
        draw.polygon([(8, 5), (11, 8), (8, 11)], fill='white')
        
        return ImageTk.PhotoImage(img)
    
    def create_delete_icon(self):
        """Crea icono de eliminar moderno"""
        img = Image.new('RGBA', (self.icon_size, self.icon_size), (0, 0, 0, 0))
        draw = ImageDraw.Draw(img)
        
        # Papelera moderna
        draw.rectangle([4, 4, 12, 12], fill=self.colors['danger'], outline=None)
        draw.rectangle([6, 2, 10, 4], fill=self.colors['danger'], outline=None)
        draw.rectangle([7, 6, 9, 10], fill=self.colors['light'], outline=None)
        
        return ImageTk.PhotoImage(img)
    
    def create_refresh_icon(self):
        """Crea icono de actualizar moderno"""
        img = Image.new('RGBA', (self.icon_size, self.icon_size), (0, 0, 0, 0))
        draw = ImageDraw.Draw(img)
        
        # Flecha circular moderna
        draw.arc([2, 2, 14, 14], 0, 270, fill=self.colors['info'], width=2)
        draw.polygon([(12, 4), (14, 2), (12, 6)], fill=self.colors['info'], outline=None)
        
        return ImageTk.PhotoImage(img)


class VentanaConfiguracion:
    """
    Ventana de configuracion completa con todos los widgets y validaciones.
    Implementacion autocontenida sin dependencias externas.
    """

    def __init__(self, parent):
        """Inicializa la ventana de configuracion"""
        self.parent = parent
        self.available_work_areas = []
        
        # Inicializar generador de iconos modernos
        self.icon_generator = ModernIconGenerator()
        self.icons = {
            'save': self.icon_generator.create_save_icon(),
            'load': self.icon_generator.create_load_icon(),
            'manage': self.icon_generator.create_manage_icon(),
            'default': self.icon_generator.create_default_icon(),
            'use': self.icon_generator.create_use_icon(),
            'delete': self.icon_generator.create_delete_icon(),
            'refresh': self.icon_generator.create_refresh_icon()
        }
        self.assignment_rules = {"GroundOperator": {}, "Forklift": {}}
        self.assignment_widgets = {"GroundOperator": [], "Forklift": []}
        self.fleet_groups = []  # Lista de grupos de flota creados

        # Aplicar tema moderno estilo Google
        self._aplicar_tema_moderno()
        
        # Configurar tema oscuro moderno
        self._setup_modern_dark_theme()
        
        # Variable para controlar el tema actual
        self.dark_mode = False

        # Crear notebook con pestanas
        self.notebook = ttk.Notebook(parent)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)

        # Crear pestanas (5 tabs as per user specification)
        self.tab_carga = ttk.Frame(self.notebook)
        self.tab_estrategias = ttk.Frame(self.notebook)
        self.tab_flota = ttk.Frame(self.notebook)
        self.tab_layout_datos = ttk.Frame(self.notebook)
        self.tab_staging = ttk.Frame(self.notebook)

        self.notebook.add(self.tab_carga, text="Carga de Trabajo")
        self.notebook.add(self.tab_estrategias, text="Estrategias")
        self.notebook.add(self.tab_flota, text="Flota de Agentes")
        self.notebook.add(self.tab_layout_datos, text="Layout y Datos")
        self.notebook.add(self.tab_staging, text="Outbound Staging")

        # Inicializar variables
        self._inicializar_variables()

        # Crear widgets en cada pestana
        self._crear_widgets_carga()
        self._crear_widgets_estrategias()
        self._crear_widgets_flota()
        self._crear_widgets_layout_datos()
        self._crear_widgets_staging()

        # Crear frame de botones en la parte inferior
        self._crear_botones_accion()

        # Cargar Work Areas automaticamente desde el archivo por defecto
        self._cargar_work_areas_inicial()

    def _inicializar_variables(self):
        """Inicializa todas las variables de tkinter"""
        # Carga de trabajo
        self.total_ordenes_var = tk.IntVar(value=300)
        self.pct_pequeno = tk.IntVar(value=60)
        self.pct_mediano = tk.IntVar(value=30)
        self.pct_grande = tk.IntVar(value=10)
        self.vol_pequeno = tk.IntVar(value=5)
        self.vol_mediano = tk.IntVar(value=25)
        self.vol_grande = tk.IntVar(value=80)

        # Recursos
        self.num_operarios_terrestres = tk.IntVar(value=1)
        self.num_montacargas = tk.IntVar(value=1)
        self.capacidad_montacargas = tk.IntVar(value=1000)
        self.capacidad_carro = tk.IntVar(value=150)
        self.tiempo_descarga_por_tarea = tk.IntVar(value=5)

        # Estrategias (2 tipos como especifico el usuario)
        self.dispatch_strategy_var = tk.StringVar(value="Optimizacion Global")
        self.tour_type_var = tk.StringVar(value="Tour Mixto (Multi-Destino)")

        # Layout
        self.layout_path_var = tk.StringVar(value="layouts/WH1.tmx")
        self.sequence_path_var = tk.StringVar(value="layouts/Warehouse_Logic.xlsx")
        self.map_scale_var = tk.DoubleVar(value=1.3)

        # Resolucion
        self.resolution_var = tk.StringVar(value="Pequena (800x800)")

        # Outbound Staging Distribution
        self.outbound_staging_vars = {
            str(i): tk.IntVar(value=100 if i == 1 else 0)
            for i in range(1, 8)
        }

        # Flota de Agentes - Grupos dinamicos
        self.fleet_groups = {
            'GroundOperator': [],  # Lista de diccionarios con info de cada grupo
            'Forklift': []
        }
        self.fleet_group_widgets = {
            'GroundOperator': [],  # Lista de widgets para cada grupo
            'Forklift': []
        }
        self.available_work_areas = []  # Se cargara desde sequence file

    def _cargar_work_areas_inicial(self):
        """Carga Work Areas automaticamente al inicio desde el archivo configurado"""
        sequence_file = self.sequence_path_var.get()
        if sequence_file and os.path.exists(sequence_file):
            try:
                self._cargar_work_areas_automatico(sequence_file)
                # No actualizar dropdowns aqui porque aun no hay grupos creados
                # Los dropdowns se actualizaran cuando se creen grupos o se cargue config
                print(f"[INIT] Work Areas cargadas al inicio: {self.available_work_areas}")
            except Exception as e:
                print(f"[INIT] No se pudieron cargar Work Areas al inicio: {e}")
        else:
            print(f"[INIT] Archivo de secuencia no encontrado: {sequence_file}")

    def _aplicar_tema_moderno(self):
        """Aplica un tema moderno estilo Google Material Design"""
        style = ttk.Style()

        # Configurar tema base
        style.theme_use('clam')

        # Paleta de colores Google Material Design
        COLOR_PRIMARY = '#1a73e8'      # Azul Google
        COLOR_PRIMARY_DARK = '#1557b0' # Azul oscuro
        COLOR_SURFACE = '#ffffff'       # Blanco
        COLOR_BACKGROUND = '#f8f9fa'    # Gris muy claro
        COLOR_BORDER = '#dadce0'        # Gris borde
        COLOR_TEXT = '#202124'          # Texto oscuro
        COLOR_TEXT_SECONDARY = '#5f6368' # Texto secundario

        # Configurar colores de fondo de la ventana principal
        self.parent.configure(bg=COLOR_BACKGROUND)

        # Estilo para TFrame
        style.configure('TFrame', background=COLOR_BACKGROUND)

        # Estilo para TLabelframe (grupos)
        style.configure('TLabelframe',
                       background=COLOR_SURFACE,
                       bordercolor=COLOR_BORDER,
                       borderwidth=1,
                       relief='solid')
        style.configure('TLabelframe.Label',
                       background=COLOR_SURFACE,
                       foreground=COLOR_TEXT,
                       font=('Segoe UI', 10, 'bold'))

        # Estilo para TLabel
        style.configure('TLabel',
                       background=COLOR_SURFACE,
                       foreground=COLOR_TEXT,
                       font=('Segoe UI', 9))

        # Estilo para TEntry
        style.configure('TEntry',
                       fieldbackground=COLOR_SURFACE,
                       bordercolor=COLOR_BORDER,
                       lightcolor=COLOR_PRIMARY,
                       darkcolor=COLOR_PRIMARY)

        # Estilo para TButton - Botones modernos con bordes redondeados
        style.configure('TButton',
                       background=COLOR_PRIMARY,
                       foreground='white',
                       borderwidth=0,
                       focuscolor='none',
                       font=('Segoe UI', 9, 'bold'),
                       padding=(16, 8))
        style.map('TButton',
                 background=[('active', COLOR_PRIMARY_DARK),
                            ('disabled', COLOR_BORDER)],
                 foreground=[('disabled', COLOR_TEXT_SECONDARY)])

        # Estilo para botones secundarios
        style.configure('Secondary.TButton',
                       background=COLOR_SURFACE,
                       foreground=COLOR_PRIMARY,
                       borderwidth=1,
                       bordercolor=COLOR_BORDER,
                       font=('Segoe UI', 9),
                       padding=(16, 8))
        style.map('Secondary.TButton',
                 background=[('active', COLOR_BACKGROUND)])

        # Estilo para Notebook (pestañas)
        style.configure('TNotebook',
                       background=COLOR_BACKGROUND,
                       borderwidth=0)
        style.configure('TNotebook.Tab',
                       background=COLOR_SURFACE,
                       foreground=COLOR_TEXT_SECONDARY,
                       padding=(20, 12),
                       font=('Segoe UI', 10))
        style.map('TNotebook.Tab',
                 background=[('selected', COLOR_SURFACE)],
                 foreground=[('selected', COLOR_PRIMARY)],
                 expand=[('selected', (0, 0, 0, 2))])

        # Estilo para Combobox
        style.configure('TCombobox',
                       fieldbackground=COLOR_SURFACE,
                       background=COLOR_SURFACE,
                       bordercolor=COLOR_BORDER,
                       arrowcolor=COLOR_TEXT)

        # Estilo para Spinbox (si se usa)
        style.configure('TSpinbox',
                       fieldbackground=COLOR_SURFACE,
                       bordercolor=COLOR_BORDER)

    def _setup_modern_dark_theme(self):
        """Configura tema oscuro moderno estilo VS Code/Discord"""
        style = ttk.Style()
        
        # Colores del tema oscuro moderno
        dark_colors = {
            'bg_primary': '#1e1e1e',      # Fondo principal
            'bg_secondary': '#2d2d30',     # Fondo secundario
            'bg_tertiary': '#3c3c3c',      # Fondo terciario
            'text_primary': '#cccccc',     # Texto principal
            'text_secondary': '#969696',    # Texto secundario
            'accent': '#007acc',           # Color de acento
            'success': '#4caf50',          # Verde éxito
            'warning': '#ff9800',          # Naranja advertencia
            'danger': '#f44336',           # Rojo peligro
            'border': '#464647'            # Bordes
        }
        
        # Configurar notebook con tema oscuro
        style.configure('Dark.TNotebook', 
                       background=dark_colors['bg_primary'], 
                       borderwidth=0)
        style.configure('Dark.TNotebook.Tab', 
                       padding=[20, 10], 
                       background=dark_colors['bg_secondary'],
                       foreground=dark_colors['text_primary'])
        style.map('Dark.TNotebook.Tab', 
                 background=[('selected', dark_colors['bg_tertiary'])],
                 foreground=[('selected', dark_colors['text_primary'])])
        
        # Botones con tema oscuro
        style.configure('Dark.TButton', 
                       padding=[15, 8], 
                       relief='flat',
                       background=dark_colors['bg_secondary'],
                       foreground=dark_colors['text_primary'],
                       borderwidth=1,
                       focuscolor='none')
        style.map('Dark.TButton', 
                 background=[('active', dark_colors['bg_tertiary']),
                           ('pressed', dark_colors['accent'])],
                 foreground=[('active', dark_colors['text_primary'])])
        
        # Labels con tema oscuro
        style.configure('Dark.TLabel', 
                       background=dark_colors['bg_primary'], 
                       foreground=dark_colors['text_primary'])
        
        # Frames con tema oscuro
        style.configure('Dark.TFrame', 
                       background=dark_colors['bg_primary'],
                       relief='flat',
                       borderwidth=1)
        
        # LabelFrame con tema oscuro
        style.configure('Dark.TLabelFrame', 
                       background=dark_colors['bg_primary'],
                       foreground=dark_colors['text_primary'],
                       borderwidth=1,
                       relief='solid')
        style.configure('Dark.TLabelFrame.Label', 
                       background=dark_colors['bg_primary'],
                       foreground=dark_colors['accent'])
        
        # Combobox con tema oscuro
        style.configure('Dark.TCombobox', 
                       padding=[8, 4],
                       background=dark_colors['bg_secondary'],
                       foreground=dark_colors['text_primary'],
                       fieldbackground=dark_colors['bg_secondary'],
                       borderwidth=1)
        
        # Entry con tema oscuro
        style.configure('Dark.TEntry', 
                       background=dark_colors['bg_secondary'],
                       foreground=dark_colors['text_primary'],
                       fieldbackground=dark_colors['bg_secondary'],
                       borderwidth=1)
        
        # Checkbutton con tema oscuro
        style.configure('Dark.TCheckbutton', 
                       background=dark_colors['bg_primary'],
                       foreground=dark_colors['text_primary'],
                       focuscolor='none')
        
        print("[VENTANA_CONFIGURACION] Tema oscuro moderno configurado")
    
    def _toggle_theme(self):
        """Alterna entre tema claro y oscuro"""
        self.dark_mode = not self.dark_mode
        
        if self.dark_mode:
            # Aplicar tema oscuro
            self._apply_dark_theme()
            print("[VENTANA_CONFIGURACION] Tema oscuro activado")
        else:
            # Aplicar tema claro
            self._apply_light_theme()
            print("[VENTANA_CONFIGURACION] Tema claro activado")
    
    def _apply_dark_theme(self):
        """Aplica el tema oscuro a todos los widgets"""
        style = ttk.Style()
        
        # Configurar colores del tema oscuro
        dark_colors = {
            'bg_primary': '#1e1e1e',
            'bg_secondary': '#2d2d30',
            'bg_tertiary': '#3c3c3c',
            'text_primary': '#cccccc',
            'text_secondary': '#969696',
            'accent': '#007acc',
            'border': '#464647'
        }
        
        # Configurar fondo principal
        self.parent.configure(bg=dark_colors['bg_primary'])
        
        # Aplicar estilos oscuros a todos los widgets ttk
        style.configure('TFrame', background=dark_colors['bg_primary'])
        style.configure('TLabel', background=dark_colors['bg_primary'], foreground=dark_colors['text_primary'])
        style.configure('TLabelFrame', background=dark_colors['bg_primary'], foreground=dark_colors['text_primary'])
        style.configure('TLabelFrame.Label', background=dark_colors['bg_primary'], foreground=dark_colors['accent'])
        style.configure('TButton', background=dark_colors['bg_secondary'], foreground=dark_colors['text_primary'])
        style.configure('TEntry', fieldbackground=dark_colors['bg_secondary'], foreground=dark_colors['text_primary'])
        style.configure('TCombobox', fieldbackground=dark_colors['bg_secondary'], foreground=dark_colors['text_primary'])
        style.configure('TNotebook', background=dark_colors['bg_primary'])
        style.configure('TNotebook.Tab', background=dark_colors['bg_secondary'], foreground=dark_colors['text_primary'])
        style.configure('TScrollbar', background=dark_colors['bg_secondary'], troughcolor=dark_colors['bg_primary'])
        
        # Aplicar tema oscuro a widgets tk nativos que no heredan de ttk
        self._apply_dark_theme_to_tk_widgets(dark_colors)
        
        # Actualizar el botón de tema
        for widget in self.parent.winfo_children():
            if isinstance(widget, ttk.Button) and widget.cget('text') in ['🌙', '☀️']:
                widget.configure(text='☀️')
                break
    
    def _apply_dark_theme_to_tk_widgets(self, dark_colors):
        """Aplica tema oscuro a widgets tk nativos que no heredan de ttk"""
        try:
            # Configurar colores para widgets tk nativos
            self.parent.option_add('*Entry*background', dark_colors['bg_secondary'])
            self.parent.option_add('*Entry*foreground', dark_colors['text_primary'])
            self.parent.option_add('*Entry*insertBackground', dark_colors['text_primary'])
            self.parent.option_add('*Entry*selectBackground', dark_colors['accent'])
            self.parent.option_add('*Entry*selectForeground', dark_colors['text_primary'])
            
            self.parent.option_add('*Text*background', dark_colors['bg_secondary'])
            self.parent.option_add('*Text*foreground', dark_colors['text_primary'])
            self.parent.option_add('*Text*insertBackground', dark_colors['text_primary'])
            self.parent.option_add('*Text*selectBackground', dark_colors['accent'])
            self.parent.option_add('*Text*selectForeground', dark_colors['text_primary'])
            
            self.parent.option_add('*Canvas*background', dark_colors['bg_secondary'])
            self.parent.option_add('*Canvas*foreground', dark_colors['text_primary'])
            
            self.parent.option_add('*Listbox*background', dark_colors['bg_secondary'])
            self.parent.option_add('*Listbox*foreground', dark_colors['text_primary'])
            self.parent.option_add('*Listbox*selectBackground', dark_colors['accent'])
            self.parent.option_add('*Listbox*selectForeground', dark_colors['text_primary'])
            
            self.parent.option_add('*Scrollbar*background', dark_colors['bg_secondary'])
            self.parent.option_add('*Scrollbar*troughcolor', dark_colors['bg_primary'])
            self.parent.option_add('*Scrollbar*activebackground', dark_colors['bg_tertiary'])
            
            # Aplicar a widgets existentes recursivamente
            self._apply_dark_theme_recursive(self.parent, dark_colors)
            
        except Exception as e:
            print(f"[VENTANA_CONFIGURACION] Error aplicando tema oscuro a widgets tk: {e}")
    
    def _apply_dark_theme_recursive(self, widget, dark_colors):
        """Aplica tema oscuro recursivamente a todos los widgets hijos"""
        try:
            # Solo aplicar a widgets tk nativos específicos, NO a widgets ttk
            widget_class = widget.__class__.__name__
            
            # Verificar si es un widget tk nativo (no ttk)
            if widget_class.startswith('Tk') or widget_class in ['Entry', 'Text', 'Canvas', 'Listbox', 'Scrollbar', 'Toplevel', 'Label', 'Button']:
                if isinstance(widget, tk.Entry):
                    widget.configure(
                        bg=dark_colors['bg_secondary'],
                        fg=dark_colors['text_primary'],
                        insertbackground=dark_colors['text_primary'],
                        selectbackground=dark_colors['accent'],
                        selectforeground=dark_colors['text_primary']
                    )
                elif isinstance(widget, tk.Text):
                    widget.configure(
                        bg=dark_colors['bg_secondary'],
                        fg=dark_colors['text_primary'],
                        insertbackground=dark_colors['text_primary'],
                        selectbackground=dark_colors['accent'],
                        selectforeground=dark_colors['text_primary']
                    )
                elif isinstance(widget, tk.Canvas):
                    widget.configure(
                        bg=dark_colors['bg_secondary'],
                        fg=dark_colors['text_primary']
                    )
                elif isinstance(widget, tk.Listbox):
                    widget.configure(
                        bg=dark_colors['bg_secondary'],
                        fg=dark_colors['text_primary'],
                        selectbackground=dark_colors['accent'],
                        selectforeground=dark_colors['text_primary']
                    )
                elif isinstance(widget, tk.Scrollbar):
                    widget.configure(
                        bg=dark_colors['bg_secondary'],
                        troughcolor=dark_colors['bg_primary'],
                        activebackground=dark_colors['bg_tertiary']
                    )
                elif isinstance(widget, tk.Toplevel):
                    widget.configure(bg=dark_colors['bg_primary'])
                elif isinstance(widget, tk.Label):
                    widget.configure(
                        bg=dark_colors['bg_primary'],
                        fg=dark_colors['text_primary']
                    )
                elif isinstance(widget, tk.Button):
                    widget.configure(
                        bg=dark_colors['bg_secondary'],
                        fg=dark_colors['text_primary'],
                        activebackground=dark_colors['accent'],
                        activeforeground=dark_colors['text_primary']
                    )
            
            # Recursivamente aplicar a widgets hijos
            for child in widget.winfo_children():
                self._apply_dark_theme_recursive(child, dark_colors)
                
        except Exception as e:
            # Silenciar errores de widgets que no soportan configuración
            pass
    
    def _apply_light_theme(self):
        """Aplica el tema claro a todos los widgets"""
        # Reaplicar el tema claro original
        self._aplicar_tema_moderno()
        
        # Restaurar colores claros para widgets tk nativos
        self._restore_light_theme_to_tk_widgets()
        
        # Actualizar el botón de tema
        for widget in self.parent.winfo_children():
            if isinstance(widget, ttk.Button) and widget.cget('text') in ['🌙', '☀️']:
                widget.configure(text='🌙')
                break
    
    def _restore_light_theme_to_tk_widgets(self):
        """Restaura tema claro a widgets tk nativos"""
        try:
            # Restaurar colores claros por defecto
            self.parent.option_clear()
            
            # Aplicar tema claro recursivamente
            self._apply_light_theme_recursive(self.parent)
            
        except Exception as e:
            print(f"[VENTANA_CONFIGURACION] Error restaurando tema claro: {e}")
    
    def _apply_light_theme_recursive(self, widget):
        """Aplica tema claro recursivamente a todos los widgets hijos"""
        try:
            # Solo aplicar a widgets tk nativos específicos, NO a widgets ttk
            widget_class = widget.__class__.__name__
            
            # Verificar si es un widget tk nativo (no ttk)
            if widget_class.startswith('Tk') or widget_class in ['Entry', 'Text', 'Canvas', 'Listbox', 'Scrollbar', 'Toplevel', 'Label', 'Button']:
                if isinstance(widget, tk.Entry):
                    widget.configure(
                        bg='white',
                        fg='black',
                        insertbackground='black',
                        selectbackground='#0078d4',
                        selectforeground='white'
                    )
                elif isinstance(widget, tk.Text):
                    widget.configure(
                        bg='white',
                        fg='black',
                        insertbackground='black',
                        selectbackground='#0078d4',
                        selectforeground='white'
                    )
                elif isinstance(widget, tk.Canvas):
                    widget.configure(
                        bg='white',
                        fg='black'
                    )
                elif isinstance(widget, tk.Listbox):
                    widget.configure(
                        bg='white',
                        fg='black',
                        selectbackground='#0078d4',
                        selectforeground='white'
                    )
                elif isinstance(widget, tk.Scrollbar):
                    widget.configure(
                        bg='#f0f0f0',
                        troughcolor='#e0e0e0',
                        activebackground='#d0d0d0'
                    )
                elif isinstance(widget, tk.Toplevel):
                    widget.configure(bg='#f0f0f0')
                elif isinstance(widget, tk.Label):
                    widget.configure(
                        bg='#f0f0f0',
                        fg='black'
                    )
                elif isinstance(widget, tk.Button):
                    widget.configure(
                        bg='#f0f0f0',
                        fg='black',
                        activebackground='#e0e0e0',
                        activeforeground='black'
                    )
            
            # Recursivamente aplicar a widgets hijos
            for child in widget.winfo_children():
                self._apply_light_theme_recursive(child)
                
        except Exception as e:
            # Silenciar errores de widgets que no soportan configuración
            pass

    def _crear_widgets_carga(self):
        """Crea widgets de la pestana Carga de Trabajo"""
        frame = ttk.LabelFrame(self.tab_carga, text="Configuracion de Ordenes", padding=10)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Total de ordenes
        row = 0
        ttk.Label(frame, text="Total de Ordenes:").grid(row=row, column=0, sticky=tk.W, pady=5)
        ttk.Entry(frame, textvariable=self.total_ordenes_var, width=15).grid(row=row, column=1, sticky=tk.W, pady=5)

        # Distribucion de tipos
        row += 1
        ttk.Label(frame, text="DISTRIBUCION DE TIPOS", font=("Arial", 10, "bold")).grid(row=row, column=0, columnspan=4, pady=10)

        # Pequeno
        row += 1
        ttk.Label(frame, text="Ordenes Pequenas (%):").grid(row=row, column=0, sticky=tk.W, pady=5)
        ttk.Entry(frame, textvariable=self.pct_pequeno, width=15).grid(row=row, column=1, sticky=tk.W, pady=5)
        ttk.Label(frame, text="Volumen:").grid(row=row, column=2, sticky=tk.W, pady=5, padx=(10, 0))
        ttk.Entry(frame, textvariable=self.vol_pequeno, width=15).grid(row=row, column=3, sticky=tk.W, pady=5)

        # Mediano
        row += 1
        ttk.Label(frame, text="Ordenes Medianas (%):").grid(row=row, column=0, sticky=tk.W, pady=5)
        ttk.Entry(frame, textvariable=self.pct_mediano, width=15).grid(row=row, column=1, sticky=tk.W, pady=5)
        ttk.Label(frame, text="Volumen:").grid(row=row, column=2, sticky=tk.W, pady=5, padx=(10, 0))
        ttk.Entry(frame, textvariable=self.vol_mediano, width=15).grid(row=row, column=3, sticky=tk.W, pady=5)

        # Grande
        row += 1
        ttk.Label(frame, text="Ordenes Grandes (%):").grid(row=row, column=0, sticky=tk.W, pady=5)
        ttk.Entry(frame, textvariable=self.pct_grande, width=15).grid(row=row, column=1, sticky=tk.W, pady=5)
        ttk.Label(frame, text="Volumen:").grid(row=row, column=2, sticky=tk.W, pady=5, padx=(10, 0))
        ttk.Entry(frame, textvariable=self.vol_grande, width=15).grid(row=row, column=3, sticky=tk.W, pady=5)

        # Label de validacion
        row += 1
        self.label_validacion = ttk.Label(frame, text="", foreground="green")
        self.label_validacion.grid(row=row, column=0, columnspan=4, pady=10)

        # Validar porcentajes en tiempo real
        self.pct_pequeno.trace_add("write", lambda *args: self.validar_porcentajes())
        self.pct_mediano.trace_add("write", lambda *args: self.validar_porcentajes())
        self.pct_grande.trace_add("write", lambda *args: self.validar_porcentajes())

    def _crear_widgets_estrategias(self):
        """Crea widgets de la pestana Estrategias - Solo 2 estrategias"""
        frame = ttk.LabelFrame(self.tab_estrategias, text="Estrategias de Operacion", padding=10)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        row = 0

        # Estrategia de Despacho
        ttk.Label(frame, text="Estrategia de Despacho:").grid(row=row, column=0, sticky=tk.W, pady=5)
        estrategias_despacho = [
            "Optimizacion Global",
            "Ejecucion de Plan (Filtro por Prioridad)"
        ]
        combo_despacho = ttk.Combobox(frame, textvariable=self.dispatch_strategy_var, values=estrategias_despacho, width=40)
        combo_despacho.grid(row=row, column=1, sticky=tk.W, pady=5)

        # Tipo de Tour de Picking
        row += 1
        ttk.Label(frame, text="Tipo de Tour de Picking:").grid(row=row, column=0, sticky=tk.W, pady=5)
        tipos_tour = [
            "Tour Mixto (Multi-Destino)",
            "Tour Simple (Un Destino)"
        ]
        combo_tour = ttk.Combobox(frame, textvariable=self.tour_type_var, values=tipos_tour, width=40)
        combo_tour.grid(row=row, column=1, sticky=tk.W, pady=5)

    def _crear_widgets_flota(self):
        """Crea widgets de la pestana Flota de Agentes con grupos dinamicos"""
        # Frame principal con scroll
        main_canvas = tk.Canvas(self.tab_flota)
        main_scrollbar = ttk.Scrollbar(self.tab_flota, orient=tk.VERTICAL, command=main_canvas.yview)
        self.flota_scrollable_frame = ttk.Frame(main_canvas)

        self.flota_scrollable_frame.bind(
            "<Configure>",
            lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        )

        main_canvas.create_window((0, 0), window=self.flota_scrollable_frame, anchor=tk.NW)
        main_canvas.configure(yscrollcommand=main_scrollbar.set)

        main_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        main_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Titulo
        ttk.Label(self.flota_scrollable_frame, text="Configuracion de Flota de Agentes",
                 font=("Arial", 12, "bold"), foreground="blue").pack(pady=10)

        # Boton Generar Flota por Defecto
        ttk.Button(self.flota_scrollable_frame, text="Generar Flota por Defecto",
                  command=self._generar_flota_defecto).pack(pady=5)

        # Seccion Operarios Terrestres
        self.ground_operators_frame = ttk.LabelFrame(self.flota_scrollable_frame,
                                                     text="Operarios Terrestres", padding=10)
        self.ground_operators_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        ttk.Button(self.ground_operators_frame, text="+ Anadir Grupo",
                  command=lambda: self._anadir_grupo_flota('GroundOperator')).pack(anchor=tk.W, pady=5)

        self.ground_operators_container = ttk.Frame(self.ground_operators_frame)
        self.ground_operators_container.pack(fill=tk.BOTH, expand=True)

        # Seccion Montacargas
        self.forklifts_frame = ttk.LabelFrame(self.flota_scrollable_frame,
                                             text="Montacargas", padding=10)
        self.forklifts_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        ttk.Button(self.forklifts_frame, text="+ Anadir Grupo",
                  command=lambda: self._anadir_grupo_flota('Forklift')).pack(anchor=tk.W, pady=5)

        self.forklifts_container = ttk.Frame(self.forklifts_frame)
        self.forklifts_container.pack(fill=tk.BOTH, expand=True)

    def _anadir_grupo_flota(self, agent_type):
        """Anade un nuevo grupo de agentes (Operarios Terrestres o Montacargas)"""
        # Determinar contenedor y valores por defecto
        if agent_type == 'GroundOperator':
            container = self.ground_operators_container
            default_capacity = 150
        else:  # Forklift
            container = self.forklifts_container
            default_capacity = 1000

        group_num = len(self.fleet_groups[agent_type]) + 1

        # Frame para el grupo
        group_frame = ttk.LabelFrame(container, text=f"Grupo {group_num}", padding=10)
        group_frame.pack(fill=tk.X, pady=5, padx=5)

        # Frame superior con parametros
        params_frame = ttk.Frame(group_frame)
        params_frame.pack(fill=tk.X)

        # Cantidad
        ttk.Label(params_frame, text="Cantidad:").grid(row=0, column=0, sticky=tk.W, padx=5)
        cantidad_var = tk.IntVar(value=2)
        cantidad_spin = ttk.Spinbox(params_frame, from_=1, to=50, textvariable=cantidad_var, width=10)
        cantidad_spin.grid(row=0, column=1, padx=5)

        # Capacidad
        ttk.Label(params_frame, text="Capacidad (L):").grid(row=0, column=2, sticky=tk.W, padx=5)
        capacidad_var = tk.IntVar(value=default_capacity)
        capacidad_spin = ttk.Spinbox(params_frame, from_=50, to=2000, textvariable=capacidad_var, width=10)
        capacidad_spin.grid(row=0, column=3, padx=5)

        # Tiempo Descarga
        ttk.Label(params_frame, text="Tiempo Descarga (s):").grid(row=0, column=4, sticky=tk.W, padx=5)
        tiempo_var = tk.IntVar(value=5)
        tiempo_spin = ttk.Spinbox(params_frame, from_=1, to=60, textvariable=tiempo_var, width=10)
        tiempo_spin.grid(row=0, column=5, padx=5)

        # Boton eliminar grupo
        ttk.Button(params_frame, text="x", width=3,
                  command=lambda: self._eliminar_grupo_flota(agent_type, group_num-1, group_frame)).grid(
                      row=0, column=6, padx=10)

        # Seccion de Prioridades de Work Area
        priorities_frame = ttk.LabelFrame(group_frame, text="Prioridades de Work Area", padding=5)
        priorities_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        # Boton para agregar prioridad
        ttk.Button(priorities_frame, text="+", width=3,
                  command=lambda: self._anadir_prioridad_wa(agent_type, group_num-1,
                                                            priorities_container)).pack(anchor=tk.W, pady=2)

        # Contenedor de prioridades
        priorities_container = ttk.Frame(priorities_frame)
        priorities_container.pack(fill=tk.BOTH, expand=True)

        # Guardar informacion del grupo
        group_data = {
            'frame': group_frame,
            'cantidad_var': cantidad_var,
            'capacidad_var': capacidad_var,
            'tiempo_var': tiempo_var,
            'priorities_container': priorities_container,
            'priorities': []  # Lista de {wa_var, priority_var, frame}
        }

        self.fleet_groups[agent_type].append(group_data)

    def _anadir_prioridad_wa(self, agent_type, group_idx, priorities_container):
        """Anade una fila de Work Area y Prioridad"""
        priority_frame = ttk.Frame(priorities_container)
        priority_frame.pack(fill=tk.X, pady=2)

        # Work Area dropdown
        ttk.Label(priority_frame, text="Work Area:").grid(row=0, column=0, sticky=tk.W, padx=5)
        wa_var = tk.StringVar()
        wa_combo = ttk.Combobox(priority_frame, textvariable=wa_var, width=15, state='readonly')
        wa_combo['values'] = self.available_work_areas if self.available_work_areas else ["1", "2", "3", "4", "5"]
        wa_combo.grid(row=0, column=1, padx=5)

        # Prioridad spinner
        ttk.Label(priority_frame, text="Prioridad:").grid(row=0, column=2, sticky=tk.W, padx=5)
        priority_var = tk.IntVar(value=1)
        priority_spin = ttk.Spinbox(priority_frame, from_=1, to=10, textvariable=priority_var, width=10)
        priority_spin.grid(row=0, column=3, padx=5)

        # Boton eliminar prioridad
        ttk.Button(priority_frame, text="x", width=3,
                  command=lambda: self._eliminar_prioridad_wa(agent_type, group_idx,
                                                              priority_frame)).grid(row=0, column=4, padx=5)

        # Guardar en el grupo (incluir referencia al combobox)
        priority_data = {
            'frame': priority_frame,
            'wa_var': wa_var,
            'priority_var': priority_var,
            'wa_combo': wa_combo  # Guardar referencia directa al combobox
        }
        self.fleet_groups[agent_type][group_idx]['priorities'].append(priority_data)

    def _eliminar_prioridad_wa(self, agent_type, group_idx, priority_frame):
        """Elimina una fila de prioridad de Work Area"""
        priority_frame.destroy()
        # Remover de la lista
        group = self.fleet_groups[agent_type][group_idx]
        group['priorities'] = [p for p in group['priorities'] if p['frame'] != priority_frame]

    def _eliminar_grupo_flota(self, agent_type, group_idx, group_frame):
        """Elimina un grupo completo de la flota"""
        group_frame.destroy()
        # Remover de la lista
        del self.fleet_groups[agent_type][group_idx]
        # Renumerar grupos restantes
        self._renumerar_grupos_flota(agent_type)

    def _renumerar_grupos_flota(self, agent_type):
        """Renumera los grupos despues de eliminar uno"""
        for idx, group in enumerate(self.fleet_groups[agent_type]):
            group['frame'].config(text=f"Grupo {idx + 1}")

    def _generar_flota_defecto(self):
        """Genera configuracion de flota por defecto"""
        # Limpiar grupos existentes
        for agent_type in ['GroundOperator', 'Forklift']:
            for group in list(self.fleet_groups[agent_type]):
                group['frame'].destroy()
            self.fleet_groups[agent_type].clear()

        # Crear 1 grupo de Operarios Terrestres
        self._anadir_grupo_flota('GroundOperator')

        # Crear 1 grupo de Montacargas
        self._anadir_grupo_flota('Forklift')

        messagebox.showinfo("Flota Generada", "Configuracion de flota por defecto generada exitosamente.")

    def _crear_widgets_layout_datos(self):
        """Crea widgets de la pestana Layout y Datos segun screenshot original"""
        # SECCION 1: Seleccion de Archivo de Layout
        frame_layout = ttk.LabelFrame(self.tab_layout_datos, text="Seleccion de Archivo de Layout", padding=10)
        frame_layout.pack(fill=tk.X, padx=10, pady=(10, 5))

        ttk.Label(frame_layout, text="Archivo TMX:").grid(row=0, column=0, sticky=tk.W, pady=5, padx=5)
        ttk.Entry(frame_layout, textvariable=self.layout_path_var, width=50).grid(row=0, column=1, sticky=tk.W, pady=5, padx=5)
        ttk.Button(frame_layout, text="Seleccionar...", command=self._examinar_tmx).grid(row=0, column=2, padx=5)

        # SECCION 2: Gestion de Datos de Secuencia (CSV)
        frame_secuencia = ttk.LabelFrame(self.tab_layout_datos, text="Gestion de Datos de Secuencia (CSV)", padding=10)
        frame_secuencia.pack(fill=tk.X, padx=10, pady=5)

        # Archivo de Secuencia
        ttk.Label(frame_secuencia, text="Archivo de Secuencia:").grid(row=0, column=0, sticky=tk.W, pady=5, padx=5)
        ttk.Entry(frame_secuencia, textvariable=self.sequence_path_var, width=50).grid(row=0, column=1, sticky=tk.W, pady=5, padx=5)
        ttk.Button(frame_secuencia, text="Seleccionar...", command=self._examinar_sequence).grid(row=0, column=2, padx=5)

        # Botones de accion
        button_frame = ttk.Frame(frame_secuencia)
        button_frame.grid(row=1, column=0, columnspan=3, pady=10)

        ttk.Button(button_frame, text="Generar Plantilla desde TMX",
                  command=self._generar_plantilla_desde_tmx).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Poblar SKUs Aleatorios en CSV",
                  command=self._poblar_skus_aleatorios).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Cargar Work Areas",
                  command=self._cargar_work_areas_manual).pack(side=tk.LEFT, padx=5)

        # Texto explicativo
        help_text = (
            "• Generar Plantilla: Analiza el TMX y crea un CSV con ubicaciones de picking\n"
            "• Poblar SKUs: Rellena el CSV con SKUs y cantidades aleatorias\n"
            "• Cargar Work Areas: Carga WA/WG desde archivo de secuencia para configurar prioridades"
        )
        help_label = ttk.Label(frame_secuencia, text=help_text, foreground="gray", justify=tk.LEFT)
        help_label.grid(row=2, column=0, columnspan=3, sticky=tk.W, pady=5, padx=5)

        # SECCION 3: Configuracion de Ventana
        frame_ventana = ttk.LabelFrame(self.tab_layout_datos, text="Configuracion de Ventana", padding=10)
        frame_ventana.pack(fill=tk.X, padx=10, pady=5)

        # Resolucion con label descriptivo
        res_frame = ttk.Frame(frame_ventana)
        res_frame.pack(fill=tk.X, padx=5, pady=5)

        ttk.Label(res_frame, text="Resolucion de Pantalla:", font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=5)
        resoluciones = ["Pequena (800x800)", "Mediana (1024x768)", "Grande (1280x1024)", "Extra Grande (1920x1080)"]
        combo = ttk.Combobox(res_frame, textvariable=self.resolution_var, values=resoluciones, width=25, state='readonly')
        combo.pack(side=tk.LEFT, padx=5)
        ttk.Label(res_frame, text="(Tamano de ventana del simulador)", foreground="gray").pack(side=tk.LEFT, padx=5)

        # Texto explicativo
        help_res = "Pequena: Rendimiento optimo | Mediana: Balance | Grande: Mejor visualizacion"
        ttk.Label(frame_ventana, text=help_res, foreground="gray").pack(anchor=tk.W, padx=5, pady=(0, 5))

    def _generar_plantilla_desde_tmx(self):
        """Genera Warehouse_Logic.xlsx desde el archivo TMX"""
        tmx_file = self.layout_path_var.get()
        if not tmx_file or not os.path.exists(tmx_file):
            messagebox.showerror("Error", "Debe seleccionar un archivo TMX valido primero.")
            return

        try:
            # TODO: Implementar logica para parsear TMX y crear XLSX
            messagebox.showinfo("Generacion de Plantilla",
                              "Funcionalidad en desarrollo.\n\n"
                              "Creara Warehouse_Logic.xlsx con:\n"
                              "- Columnas de ubicaciones de picking\n"
                              "- Datos por defecto para el simulador\n"
                              "- Estructura lista para modificar")
        except Exception as e:
            messagebox.showerror("Error", f"Error generando plantilla: {e}")

    def _poblar_skus_aleatorios(self):
        """Pobla el CSV con SKUs y cantidades aleatorias"""
        csv_file = self.sequence_path_var.get()
        if not csv_file:
            messagebox.showerror("Error", "Debe seleccionar un archivo de secuencia primero.")
            return

        try:
            # TODO: Implementar logica para poblar CSV con datos aleatorios
            messagebox.showinfo("Poblar SKUs",
                              "Funcionalidad en desarrollo.\n\n"
                              "Rellenara el CSV con:\n"
                              "- SKUs aleatorios\n"
                              "- Cantidades aleatorias")
        except Exception as e:
            messagebox.showerror("Error", f"Error poblando SKUs: {e}")

    def _cargar_work_areas_manual(self):
        """Carga Work Areas manualmente desde el archivo de secuencia seleccionado"""
        sequence_file = self.sequence_path_var.get()

        if not sequence_file:
            messagebox.showerror("Error",
                               "Debe seleccionar un archivo de secuencia primero.")
            return

        if not os.path.exists(sequence_file):
            messagebox.showerror("Error",
                               f"El archivo no existe:\n{sequence_file}")
            return

        try:
            # Usar el metodo automatico existente
            self._cargar_work_areas_automatico(sequence_file)

            # Actualizar todos los dropdowns existentes en Flota de Agentes
            self._actualizar_dropdowns_work_areas()

            if self.available_work_areas:
                messagebox.showinfo("Work Areas Cargadas",
                                  f"Se cargaron {len(self.available_work_areas)} Work Areas:\n\n" +
                                  ", ".join(self.available_work_areas) +
                                  "\n\nYa estan disponibles en la pestana Flota de Agentes.")
            else:
                messagebox.showwarning("Advertencia",
                                     "No se encontraron Work Areas en el archivo.\n"
                                     "Se usaran valores por defecto.")

        except Exception as e:
            messagebox.showerror("Error",
                               f"Error cargando Work Areas:\n{str(e)}")
            import traceback
            traceback.print_exc()

    def _actualizar_dropdowns_work_areas(self):
        """Actualiza los valores de todos los dropdowns de Work Areas en grupos existentes"""
        print(f"[DEBUG] Actualizando dropdowns con Work Areas: {self.available_work_areas}")

        # Actualizar dropdowns en Operarios Terrestres
        for group_idx, group in enumerate(self.fleet_groups['GroundOperator']):
            for priority_idx, priority_data in enumerate(group['priorities']):
                if 'wa_combo' in priority_data:
                    combo = priority_data['wa_combo']
                    current_value = priority_data['wa_var'].get()
                    combo['values'] = self.available_work_areas
                    # Preservar selección si aún es válida
                    if current_value and current_value in self.available_work_areas:
                        priority_data['wa_var'].set(current_value)
                    print(f"[DEBUG] GroundOperator Grupo {group_idx+1}, Prioridad {priority_idx+1}: Actualizado")

        # Actualizar dropdowns en Montacargas
        for group_idx, group in enumerate(self.fleet_groups['Forklift']):
            for priority_idx, priority_data in enumerate(group['priorities']):
                if 'wa_combo' in priority_data:
                    combo = priority_data['wa_combo']
                    current_value = priority_data['wa_var'].get()
                    combo['values'] = self.available_work_areas
                    # Preservar selección si aún es válida
                    if current_value and current_value in self.available_work_areas:
                        priority_data['wa_var'].set(current_value)
                    print(f"[DEBUG] Forklift Grupo {group_idx+1}, Prioridad {priority_idx+1}: Actualizado")

    def _crear_widgets_staging(self):
        """Crea widgets de la pestana Outbound Staging"""
        frame = ttk.LabelFrame(self.tab_staging, text="Distribucion de Staging", padding=10)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        ttk.Label(frame, text="Distribucion porcentual de ordenes por zona de staging:",
                 font=("Arial", 10, "bold")).grid(row=0, column=0, columnspan=3, pady=10)

        for i in range(1, 8):
            row = i
            ttk.Label(frame, text=f"Staging {i}:").grid(row=row, column=0, sticky=tk.W, pady=5)
            ttk.Entry(frame, textvariable=self.outbound_staging_vars[str(i)], width=15).grid(
                row=row, column=1, sticky=tk.W, pady=5)
            ttk.Label(frame, text="%").grid(row=row, column=2, sticky=tk.W, pady=5)

            # Trace para validacion
            self.outbound_staging_vars[str(i)].trace_add("write",
                lambda *args: self._validar_staging_distribution())

        self.staging_validation_label = ttk.Label(frame, text="", foreground="green")
        self.staging_validation_label.grid(row=8, column=0, columnspan=3, pady=10)

    def _crear_botones_accion(self):
        """Crea botones de accion en la parte inferior con diseño mejorado"""
        button_frame = ttk.Frame(self.parent)
        button_frame.pack(fill=tk.X, padx=15, pady=15)

        # ========================================================================
        # SISTEMA DE SLOTS DE CONFIGURACION - FASE 4.2: PULIDO DE UI
        # ========================================================================
        
        # Frame para botones del sistema de slots
        slots_frame = ttk.LabelFrame(button_frame, text="Sistema de Configuraciones", padding=10)
        slots_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        # SLOTS SYSTEM BUTTONS con iconos vectoriales modernos
        save_btn = ttk.Button(slots_frame, text="Save", image=self.icons['save'], 
                             compound=tk.LEFT, width=12,
                             command=self._guardar_como_callback)
        save_btn.pack(side=tk.LEFT, padx=(0, 8))
        
        load_btn = ttk.Button(slots_frame, text="Load", image=self.icons['load'], 
                             compound=tk.LEFT, width=12,
                             command=self._cargar_desde_callback)
        load_btn.pack(side=tk.LEFT, padx=(0, 8))
        
        manage_btn = ttk.Button(slots_frame, text="Manage", image=self.icons['manage'], 
                               compound=tk.LEFT, width=12,
                               command=self._eliminar_configuracion_callback)
        manage_btn.pack(side=tk.LEFT, padx=(0, 8))
        
        default_btn = ttk.Button(slots_frame, text="Default", image=self.icons['default'], 
                               compound=tk.LEFT, width=12,
                               command=self.valores_por_defecto_new)
        default_btn.pack(side=tk.LEFT, padx=(0, 8))
        
        use_btn = ttk.Button(slots_frame, text="Use", image=self.icons['use'], 
                            compound=tk.LEFT, width=12,
                            command=self._use_configuration_callback)
        use_btn.pack(side=tk.LEFT, padx=(0, 8))

        # Frame para botones de control
        control_frame = ttk.Frame(button_frame)
        control_frame.pack(side=tk.RIGHT)
        
        # Botón alternar tema
        theme_btn = ttk.Button(control_frame, text="🌙", width=3,
                              command=self._toggle_theme)
        theme_btn.pack(side=tk.RIGHT, padx=(5, 0))
        
        # Botón Salir removido - usar solo la cruz superior derecha
        
        print("[VENTANA_CONFIGURACION] Botones del sistema de slots con diseño mejorado agregados")

    # ========================================================================
    # SISTEMA DE SLOTS DE CONFIGURACION - CALLBACKS PLACEHOLDER
    # ========================================================================
    
    def _guardar_como_callback(self):
        """Placeholder para callback de Guardar Como - se conecta en ConfiguradorSimulador"""
        print("[VENTANA_CONFIGURACION] Callback _guardar_como_callback llamado")
        if hasattr(self, '_guardar_como_callback_real'):
            self._guardar_como_callback_real()
        else:
            print("[VENTANA_CONFIGURACION] Callback real no conectado")
    
    def _cargar_desde_callback(self):
        """Placeholder para callback de Cargar Desde - se conecta en ConfiguradorSimulador"""
        print("[VENTANA_CONFIGURACION] Callback _cargar_desde_callback llamado")
        if hasattr(self, '_cargar_desde_callback_real'):
            self._cargar_desde_callback_real()
        else:
            print("[VENTANA_CONFIGURACION] Callback real no conectado")
    
    def _eliminar_configuracion_callback(self):
        """Placeholder para callback de Eliminar Configuracion - se conecta en ConfiguradorSimulador"""
        print("[VENTANA_CONFIGURACION] Callback _eliminar_configuracion_callback llamado")
        if hasattr(self, '_eliminar_configuracion_callback_real'):
            self._eliminar_configuracion_callback_real()
        else:
            print("[VENTANA_CONFIGURACION] Callback real no conectado")
    
    def _use_configuration_callback(self):
        """Placeholder para callback de Use - se conecta en ConfiguradorSimulador"""
        print("[VENTANA_CONFIGURACION] Callback _use_configuration_callback llamado")
        if hasattr(self, '_use_configuration_callback_real'):
            self._use_configuration_callback_real()
        else:
            print("[VENTANA_CONFIGURACION] Callback real no conectado")

    # ========================================================================
    # METODOS DE VALIDACION
    # ========================================================================

    def validar_porcentajes(self):
        """Valida que los porcentajes sumen 100%"""
        try:
            total = self.pct_pequeno.get() + self.pct_mediano.get() + self.pct_grande.get()
            if total == 100:
                self.label_validacion.config(text="OK: Suma 100%", foreground="green")
                return True
            else:
                self.label_validacion.config(text=f"ERROR: Suma {total}% (debe ser 100%)", foreground="red")
                return False
        except:
            self.label_validacion.config(text="ERROR: Valores invalidos", foreground="red")
            return False

    def actualizar_total(self):
        """Actualiza el total de recursos - deprecated, kept for compatibility"""
        # Esta funcionalidad fue movida a "Flota de Agentes"
        pass

    def _validar_staging_distribution(self):
        """Valida que la distribucion de staging sume 100%"""
        try:
            total = sum(var.get() for var in self.outbound_staging_vars.values())
            if total == 100:
                self.staging_validation_label.config(text="OK: Suma 100%", foreground="green")
                return True
            else:
                self.staging_validation_label.config(
                    text=f"ERROR: Suma {total}% (debe ser 100%)", foreground="red")
                return False
        except:
            self.staging_validation_label.config(text="ERROR: Valores invalidos", foreground="red")
            return False

    def _validar_configuracion_picking(self, total_ordenes, pct_pequeno, pct_mediano, pct_grande,
                                      vol_pequeno, vol_mediano, vol_grande, capacidad_carro,
                                      op_terrestres, montacargas, total_recursos):
        """Valida la configuracion de picking"""
        # Validacion basica
        if total_ordenes <= 0:
            messagebox.showerror("Error", "El total de ordenes debe ser mayor a 0")
            return False

        if pct_pequeno + pct_mediano + pct_grande != 100:
            messagebox.showerror("Error", "Los porcentajes deben sumar 100%")
            return False

        if capacidad_carro <= 0:
            messagebox.showerror("Error", "La capacidad del carro debe ser mayor a 0")
            return False

        if total_recursos <= 0:
            messagebox.showerror("Error", "Debe haber al menos un recurso")
            return False

        return True

    # ========================================================================
    # METODOS DE ARCHIVOS
    # ========================================================================

    def _examinar_tmx(self):
        """Abre dialogo para seleccionar archivo TMX"""
        filename = filedialog.askopenfilename(
            title="Seleccionar archivo TMX",
            filetypes=[("TMX Files", "*.tmx"), ("All Files", "*.*")]
        )
        if filename:
            self.layout_path_var.set(filename)

    def _examinar_sequence(self):
        """Abre dialogo para seleccionar archivo de secuencia"""
        filename = filedialog.askopenfilename(
            title="Seleccionar archivo de secuencia",
            filetypes=[("Excel Files", "*.xlsx *.xls"), ("CSV Files", "*.csv"), ("All Files", "*.*")]
        )
        if filename:
            self.sequence_path_var.set(filename)
            # Cargar Work Areas automaticamente cuando se selecciona un archivo
            try:
                self._cargar_work_areas_automatico(filename)
                self._actualizar_dropdowns_work_areas()
                print(f"[AUTO-LOAD] Work Areas cargadas automaticamente: {self.available_work_areas}")
            except Exception as e:
                print(f"[AUTO-LOAD] Error al cargar Work Areas: {e}")

    # ========================================================================
    # METODOS DE FLOTA
    # ========================================================================

    def _cargar_work_areas_automatico(self, sequence_file):
        """Carga WorkAreas desde el archivo de secuencia"""
        try:
            if not os.path.exists(sequence_file):
                raise FileNotFoundError(f"Archivo no encontrado: {sequence_file}")

            # Leer WorkAreas del archivo Excel
            import openpyxl
            wb = openpyxl.load_workbook(sequence_file, read_only=True, data_only=True)

            # Buscar hoja PickingLocations (nuevo formato) o Warehouse_Logic (formato antiguo)
            sheet_name = None
            if 'PickingLocations' in wb.sheetnames:
                sheet_name = 'PickingLocations'
            elif 'Warehouse_Logic' in wb.sheetnames:
                sheet_name = 'Warehouse_Logic'

            if sheet_name:
                ws = wb[sheet_name]
                work_areas = set()

                # Buscar columna WorkArea en los encabezados
                headers = [cell.value for cell in ws[1]]
                wa_col_idx = None
                for idx, header in enumerate(headers):
                    if header and 'WorkArea' in str(header):
                        wa_col_idx = idx
                        break

                if wa_col_idx is not None:
                    # Leer WorkAreas desde la columna encontrada
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row and len(row) > wa_col_idx:
                            work_area = row[wa_col_idx]
                            if work_area:
                                work_areas.add(str(work_area))

                    self.available_work_areas = sorted(list(work_areas))
                    print(f"[VENTANA_CONFIG] {len(work_areas)} WorkAreas cargadas desde '{sheet_name}': {self.available_work_areas}")
                else:
                    print(f"[VENTANA_CONFIG] WARNING: Columna 'WorkArea' no encontrada en '{sheet_name}'")
                    self.available_work_areas = ["Area_1", "Area_2", "Area_3"]
            else:
                # Si no hay hoja, usar valores por defecto
                self.available_work_areas = ["Area_1", "Area_2", "Area_3"]
                print(f"[VENTANA_CONFIG] Hojas disponibles: {wb.sheetnames}")
                print("[VENTANA_CONFIG] WorkAreas por defecto cargadas (hoja no encontrada)")

            wb.close()

        except Exception as e:
            print(f"[VENTANA_CONFIG ERROR] Error cargando WorkAreas: {e}")
            import traceback
            traceback.print_exc()
            # Usar valores por defecto en caso de error
            self.available_work_areas = ["Area_1", "Area_2", "Area_3"]

    def _poblar_ui_flota(self, grupos):
        """Pobla la UI de flota con grupos existentes desde config cargado"""
        # Limpiar grupos actuales
        for agent_type in ['GroundOperator', 'Forklift']:
            for group in list(self.fleet_groups[agent_type]):
                group['frame'].destroy()
            self.fleet_groups[agent_type].clear()

        # Recrear grupos desde config
        for grupo in grupos:
            agent_type = grupo['agent_type']
            self._anadir_grupo_flota(agent_type)

            # Obtener el grupo recien creado
            group = self.fleet_groups[agent_type][-1]

            # Setear valores
            group['cantidad_var'].set(grupo['cantidad'])
            group['capacidad_var'].set(grupo['capacidad'])
            group['tiempo_var'].set(grupo['tiempo_descarga'])

            # Agregar prioridades
            for wa, priority in grupo['work_area_priorities'].items():
                self._anadir_prioridad_wa(agent_type, len(self.fleet_groups[agent_type])-1,
                                         group['priorities_container'])
                # Setear valores de la prioridad recien creada
                priority_data = group['priorities'][-1]
                priority_data['wa_var'].set(str(wa))
                priority_data['priority_var'].set(priority)

    # ========================================================================
    # METODOS DE VALORES POR DEFECTO
    # ========================================================================

    def valores_por_defecto_new(self):
        """Carga valores por defecto desde la configuración marcada como default"""
        try:
            print("[VENTANA_CONFIGURACION] Cargando valores por defecto desde configuración...")
            
            # Intentar cargar configuración marcada como default
            if hasattr(self, '_configurador') and hasattr(self._configurador, 'config_manager'):
                default_config = self._configurador.config_manager.get_default_configuration()
                
                if default_config and 'configuration' in default_config:
                    config = default_config['configuration']
                    print(f"[VENTANA_CONFIGURACION] Cargando configuración default: {default_config['metadata']['name']}")
                    
                    # Carga de trabajo
                    self.total_ordenes_var.set(config.get('total_ordenes', 300))
                    
                    # Distribución de tipos
                    distribucion = config.get('distribucion_tipos', {})
                    self.pct_pequeno.set(distribucion.get('pequeno', {}).get('porcentaje', 60))
                    self.pct_mediano.set(distribucion.get('mediano', {}).get('porcentaje', 30))
                    self.pct_grande.set(distribucion.get('grande', {}).get('porcentaje', 10))
                    self.vol_pequeno.set(distribucion.get('pequeno', {}).get('volumen', 5))
                    self.vol_mediano.set(distribucion.get('mediano', {}).get('volumen', 25))
                    self.vol_grande.set(distribucion.get('grande', {}).get('volumen', 80))
                    
                    self.capacidad_carro.set(config.get('capacidad_carro', 150))

                    # Recursos
                    self.num_operarios_terrestres.set(config.get('num_operarios_terrestres', 2))
                    self.num_montacargas.set(config.get('num_montacargas', 1))
                    self.capacidad_montacargas.set(config.get('capacidad_montacargas', 1000))
                    self.tiempo_descarga_por_tarea.set(config.get('tiempo_descarga_por_tarea', 5))

                    # Estrategias
                    self.dispatch_strategy_var.set(config.get('dispatch_strategy', "Optimizacion Global"))
                    self.tour_type_var.set(config.get('tour_type', "Tour Mixto (Multi-Destino)"))

                    # Layout
                    self.layout_path_var.set(config.get('layout_file', "layouts/WH1.tmx"))
                    self.sequence_path_var.set(config.get('sequence_file', "layouts/Warehouse_Logic.xlsx"))
                    self.map_scale_var.set(config.get('map_scale', 1.3))
                    self.resolution_var.set(config.get('selected_resolution_key', "Pequena (800x800)"))

                    # Outbound Staging
                    staging_dist = config.get('outbound_staging_distribution', {})
                    for i in range(1, 8):
                        self.outbound_staging_vars[str(i)].set(staging_dist.get(str(i), 100 if i == 1 else 0))

                    # Cargar configuración de agentes desde agent_types
                    agent_types = config.get('agent_types', [])
                    if agent_types:
                        self._cargar_configuracion_agentes_desde_agent_types(agent_types)
                    else:
                        # Fallback a valores por defecto si no hay agent_types
                        self._generar_asignacion_defecto()

                    # Validaciones
                    self.validar_porcentajes()
                    self.actualizar_total()
                    self._validar_staging_distribution()

                    print("[CONFIGURATOR] Valores por defecto cargados desde configuración")
                    return
                    
            # Fallback: usar valores hardcoded si no hay configuración default
            print("[VENTANA_CONFIGURACION] No hay configuración default, usando valores hardcoded...")
            
            # Carga de trabajo
            self.total_ordenes_var.set(300)
            self.pct_pequeno.set(60)
            self.pct_mediano.set(30)
            self.pct_grande.set(10)
            self.vol_pequeno.set(5)
            self.vol_mediano.set(25)
            self.vol_grande.set(80)
            self.capacidad_carro.set(150)

            # Recursos
            self.num_operarios_terrestres.set(2)
            self.num_montacargas.set(1)
            self.capacidad_montacargas.set(1000)
            self.tiempo_descarga_por_tarea.set(5)

            # Estrategias (solo 2)
            self.dispatch_strategy_var.set("Optimizacion Global")
            self.tour_type_var.set("Tour Mixto (Multi-Destino)")

            # Layout
            self.layout_path_var.set("layouts/WH1.tmx")
            self.sequence_path_var.set("layouts/Warehouse_Logic.xlsx")
            self.map_scale_var.set(1.3)
            self.resolution_var.set("Pequena (800x800)")

            # Outbound Staging
            for i in range(1, 8):
                self.outbound_staging_vars[str(i)].set(100 if i == 1 else 0)

            # Asignacion de recursos - valores por defecto
            self._generar_asignacion_defecto()

            # Validaciones
            self.validar_porcentajes()
            self.actualizar_total()
            self._validar_staging_distribution()

            print("[CONFIGURATOR] Valores por defecto cargados (fallback)")
            
        except Exception as e:
            print(f"[VENTANA_CONFIGURACION ERROR] Error cargando valores por defecto: {e}")

    def _generar_asignacion_defecto(self):
        """Genera asignacion por defecto para los recursos"""
        try:
            # Limpiar grupos existentes
            self.fleet_groups = {"GroundOperator": [], "Forklift": []}
            
            # Crear grupo por defecto de Operarios Terrestres
            self._anadir_grupo_flota("GroundOperator")
            
            # Crear grupo por defecto de Montacargas
            self._anadir_grupo_flota("Forklift")
            
            print("[VENTANA_CONFIGURACION] Asignacion por defecto generada")
            
        except Exception as e:
            print(f"[VENTANA_CONFIGURACION ERROR] Error generando asignacion por defecto: {e}")

    def _generar_config_defecto(self):
        """Genera configuracion por defecto para la flota de agentes"""
        try:
            config_defecto = []
            
            # Grupo por defecto de Operarios Terrestres
            config_defecto.append({
                'agent_type': 'GroundOperator',
                'cantidad': 2,
                'capacidad': 150,
                'tiempo_descarga': 5,
                'priorities': []
            })
            
            # Grupo por defecto de Montacargas
            config_defecto.append({
                'agent_type': 'Forklift',
                'cantidad': 1,
                'capacidad': 1000,
                'tiempo_descarga': 5,
                'priorities': []
            })
            
            print("[VENTANA_CONFIGURACION] Configuracion por defecto generada")
            return config_defecto
            
        except Exception as e:
            print(f"[VENTANA_CONFIGURACION ERROR] Error generando configuracion por defecto: {e}")
            return []

    def _crear_grupo_desde_config(self, group_config):
        """Crea un grupo de flota desde configuracion"""
        try:
            agent_type = group_config['agent_type']
            
            # Anadir grupo
            self._anadir_grupo_flota(agent_type)
            
            # Obtener el grupo recien creado
            group = self.fleet_groups[agent_type][-1]
            
            # Setear valores
            group['cantidad_var'].set(group_config.get('cantidad', 2))
            group['capacidad_var'].set(group_config.get('capacidad', 150 if agent_type == 'GroundOperator' else 1000))
            group['tiempo_var'].set(group_config.get('tiempo_descarga', 5))
            
            # Agregar prioridades si existen
            priorities = group_config.get('priorities', [])
            for priority_data in priorities:
                self._anadir_prioridad_wa(agent_type, len(self.fleet_groups[agent_type]) - 1, group['priorities_container'])
                # Setear valores de la prioridad recien creada
                priority_data_ui = group['priorities'][-1]
                priority_data_ui['wa_var'].set(str(priority_data.get('wa', '')))
                priority_data_ui['priority_var'].set(priority_data.get('priority', 1))
            
            print(f"[VENTANA_CONFIGURACION] Grupo {agent_type} creado desde configuracion")
            
        except Exception as e:
            print(f"[VENTANA_CONFIGURACION ERROR] Error creando grupo desde configuracion: {e}")

    def _limpiar_todos_los_grupos(self):
        """Limpia todos los grupos de flota existentes"""
        try:
            for agent_type in ['GroundOperator', 'Forklift']:
                for group in list(self.fleet_groups[agent_type]):
                    group['frame'].destroy()
                self.fleet_groups[agent_type].clear()
            
            print("[VENTANA_CONFIGURACION] Todos los grupos limpiados")
            
        except Exception as e:
            print(f"[VENTANA_CONFIGURACION ERROR] Error limpiando grupos: {e}")

    def _cargar_configuracion_agentes_desde_agent_types(self, agent_types):
        """Carga configuración de agentes desde la lista agent_types"""
        try:
            print("[VENTANA_CONFIGURACION] Cargando configuración de agentes desde agent_types...")
            
            # Limpiar grupos existentes
            self._limpiar_todos_los_grupos()
            
            # Agrupar agentes por tipo
            agentes_por_tipo = {}
            for agent in agent_types:
                agent_type = agent.get('type', 'GroundOperator')
                if agent_type not in agentes_por_tipo:
                    agentes_por_tipo[agent_type] = []
                agentes_por_tipo[agent_type].append(agent)
            
            # Crear grupos para cada tipo de agente
            for agent_type, agentes in agentes_por_tipo.items():
                if agent_type in ['GroundOperator', 'Forklift']:
                    # Crear grupo
                    self._anadir_grupo_flota(agent_type)
                    group = self.fleet_groups[agent_type][-1]
                    
                    # Configurar valores del grupo
                    group['cantidad_var'].set(len(agentes))
                    
                    # Usar valores del primer agente como referencia
                    if agentes:
                        first_agent = agentes[0]
                        group['capacidad_var'].set(first_agent.get('capacity', 150 if agent_type == 'GroundOperator' else 1000))
                        group['tiempo_var'].set(first_agent.get('discharge_time', 5))
                        
                        # Cargar prioridades de Work Area
                        work_area_priorities = first_agent.get('work_area_priorities', {})
                        for wa, priority in work_area_priorities.items():
                            self._anadir_prioridad_wa(agent_type, len(self.fleet_groups[agent_type]) - 1, group['priorities_container'])
                            priority_data_ui = group['priorities'][-1]
                            priority_data_ui['wa_var'].set(str(wa))
                            priority_data_ui['priority_var'].set(priority)
            
            print("[VENTANA_CONFIGURACION] Configuración de agentes cargada desde agent_types")
            
        except Exception as e:
            print(f"[VENTANA_CONFIGURACION ERROR] Error cargando configuración de agentes: {e}")
            # Fallback a valores por defecto
            self._generar_asignacion_defecto()

    def obtener_configuracion(self):
        """Obtiene la configuracion actual como diccionario"""
        # Sincronizar assignment_rules desde widgets
        for agent_type, widget_rows in self.assignment_widgets.items():
            self.assignment_rules[agent_type] = {}
            for row in widget_rows:
                level = row['level_var'].get()
                priority = row['priority_var'].get()
                self.assignment_rules[agent_type][level] = priority

        # Construir agent_types desde fleet_groups (nueva estructura)
        agent_types = []

        # Procesar Operarios Terrestres
        for group in self.fleet_groups['GroundOperator']:
            cantidad = group['cantidad_var'].get()
            capacidad = group['capacidad_var'].get()
            tiempo_descarga = group['tiempo_var'].get()

            # Construir diccionario de prioridades de Work Area
            work_area_priorities = {}
            for priority_data in group['priorities']:
                wa = priority_data['wa_var'].get()
                priority = priority_data['priority_var'].get()
                if wa:  # Solo agregar si se selecciono un WA
                    work_area_priorities[wa] = priority

            # Crear agentes individuales
            for _ in range(cantidad):
                agent_types.append({
                    'type': 'GroundOperator',
                    'capacity': capacidad,
                    'discharge_time': tiempo_descarga,
                    'work_area_priorities': work_area_priorities.copy()
                })

        # Procesar Montacargas
        for group in self.fleet_groups['Forklift']:
            cantidad = group['cantidad_var'].get()
            capacidad = group['capacidad_var'].get()
            tiempo_descarga = group['tiempo_var'].get()

            # Construir diccionario de prioridades de Work Area
            work_area_priorities = {}
            for priority_data in group['priorities']:
                wa = priority_data['wa_var'].get()
                priority = priority_data['priority_var'].get()
                if wa:  # Solo agregar si se selecciono un WA
                    work_area_priorities[wa] = priority

            # Crear agentes individuales
            for _ in range(cantidad):
                agent_types.append({
                    'type': 'Forklift',
                    'capacity': capacidad,
                    'discharge_time': tiempo_descarga,
                    'work_area_priorities': work_area_priorities.copy()
                })

        config = {
            'total_ordenes': self.total_ordenes_var.get(),
            'distribucion_tipos': {
                'pequeno': {'porcentaje': self.pct_pequeno.get(), 'volumen': self.vol_pequeno.get()},
                'mediano': {'porcentaje': self.pct_mediano.get(), 'volumen': self.vol_mediano.get()},
                'grande': {'porcentaje': self.pct_grande.get(), 'volumen': self.vol_grande.get()}
            },
            'capacidad_carro': self.capacidad_carro.get(),
            'dispatch_strategy': self.dispatch_strategy_var.get(),
            'tour_type': self.tour_type_var.get(),
            'layout_file': self.layout_path_var.get(),
            'sequence_file': self.sequence_path_var.get(),
            'map_scale': self.map_scale_var.get(),
            'selected_resolution_key': self.resolution_var.get(),
            'num_operarios_terrestres': self.num_operarios_terrestres.get(),
            'num_montacargas': self.num_montacargas.get(),
            'num_operarios_total': self.num_operarios_terrestres.get() + self.num_montacargas.get(),
            'capacidad_montacargas': self.capacidad_montacargas.get(),
            'tiempo_descarga_por_tarea': self.tiempo_descarga_por_tarea.get(),
            'assignment_rules': self.assignment_rules.copy(),
            'outbound_staging_distribution': {str(i): self.outbound_staging_vars[str(i)].get() for i in range(1, 8)},
            'agent_types': agent_types,
            'num_operarios': self.num_operarios_terrestres.get() + self.num_montacargas.get()
        }

        return config

    # ========================================================================
    # CALLBACKS (stubs - seran conectados por ConfiguradorSimulador)
    # ========================================================================

    def _guardar_callback(self):
        """Placeholder para guardar - se sobreescribe en ConfiguradorSimulador"""
        print("[VENTANA] Guardar callback")

    def _cargar_callback(self):
        """Placeholder para cargar - se sobreescribe en ConfiguradorSimulador"""
        print("[VENTANA] Cargar callback")

    def _probar_callback(self):
        """Placeholder para probar"""
        print("[VENTANA] Probar callback")

    # Callback _salir_callback removido - usar solo la cruz superior derecha


class ConfiguradorSimulador:
    """Configurador independiente del simulador con funcionalidad de guardado"""

    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Warehouse Simulation Configurator")
        self.root.geometry("900x700")
        self.root.resizable(True, True)

        # Configurar tamaño mínimo
        self.root.minsize(800, 600)

        # Centrar ventana
        self._centrar_ventana()

        # ========================================================================
        # SISTEMA DE SLOTS DE CONFIGURACION - FASE 3.1: INTEGRACION
        # ========================================================================
        
        # Inicializar ConfigurationManager
        self.config_manager = ConfigurationManager()
        print("[CONFIGURATOR] ConfigurationManager inicializado")

        # Crear el configurador principal pasando la ventana raiz
        self.ventana_config = VentanaConfiguracion(self.root)

        # Conectar callbacks INMEDIATAMENTE después de crear la ventana
        self.ventana_config._guardar_callback = self.guardar_configuracion
        self.ventana_config._cargar_callback = self.cargar_configuracion_manual
        # Callback _salir_callback removido - usar solo la cruz superior derecha
        
        # ========================================================================
        # NUEVOS CALLBACKS PARA SISTEMA DE SLOTS
        # ========================================================================
        self.ventana_config._guardar_como_callback_real = self._guardar_como_callback
        self.ventana_config._cargar_desde_callback_real = self._cargar_desde_callback
        self.ventana_config._eliminar_configuracion_callback_real = self._eliminar_configuracion_callback
        self.ventana_config._use_configuration_callback_real = self._use_configuration_callback
        
        # Conectar referencia al configurador para acceso a config_manager
        self.ventana_config._configurador = self
        
        # Conectar referencia a los iconos para los diálogos
        self.ventana_config._icons = self.ventana_config.icons
        
        print("[CONFIGURATOR] Callbacks conectados (incluyendo sistema de slots)")

        # CORRECCION: Diferir carga hasta que UI este completamente lista
        self.root.after(100, self._cargar_configuracion_existente)

        print("[CONFIGURATOR] Configurador independiente inicializado con sistema de slots")

    def _centrar_ventana(self):
        """Centra la ventana en la pantalla"""
        self.root.update_idletasks()
        ancho_ventana = 900
        alto_ventana = 700
        x = (self.root.winfo_screenwidth() // 2) - (ancho_ventana // 2)
        y = (self.root.winfo_screenheight() // 2) - (alto_ventana // 2)
        self.root.geometry(f"{ancho_ventana}x{alto_ventana}+{x}+{y}")

    def _cargar_configuracion_existente(self):
        """Carga configuración marcada como default al iniciar"""
        try:
            print("[CONFIGURATOR] Intentando cargar configuración marcada como default...")
            
            # Intentar cargar configuración marcada como default
            default_config = self.config_manager.get_default_configuration()
            
            if default_config and 'configuration' in default_config:
                config = default_config['configuration']
                print(f"[CONFIGURATOR] Cargando configuración default: {default_config['metadata']['name']}")
                
                # Sanitizar assignment_rules: convertir claves str a int
                if 'assignment_rules' in config and config['assignment_rules']:
                    sanitized_rules = {}
                    for agent_type, rules in config['assignment_rules'].items():
                        sanitized_rules[agent_type] = {int(k): v for k, v in rules.items()}
                    config['assignment_rules'] = sanitized_rules
                    print("[CONFIGURATOR] assignment_rules sanitizadas: claves str -> int")

                self._poblar_ui_desde_config(config)
                print("[CONFIGURATOR] Configuración default cargada exitosamente en UI")
                return
                
            # Fallback: cargar config.json si no hay configuración default
            print("[CONFIGURATOR] No hay configuración default, cargando config.json...")
            config_path = os.path.join(os.path.dirname(__file__), "config.json")

            if os.path.exists(config_path):
                print(f"[CONFIGURATOR] Cargando configuracion desde: {config_path}")
                with open(config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)

                # Sanitizar assignment_rules: convertir claves str a int
                if 'assignment_rules' in config and config['assignment_rules']:
                    sanitized_rules = {}
                    for agent_type, rules in config['assignment_rules'].items():
                        sanitized_rules[agent_type] = {int(k): v for k, v in rules.items()}
                    config['assignment_rules'] = sanitized_rules
                    print("[CONFIGURATOR] assignment_rules sanitizadas: claves str -> int")

                self._poblar_ui_desde_config(config)
                print("[CONFIGURATOR] Configuracion cargada exitosamente en UI")

            else:
                print("[CONFIGURATOR] config.json no encontrado, usando valores por defecto")
                # Cargar valores por defecto usando el método que ya funciona
                self.ventana_config.valores_por_defecto_new()

        except Exception as e:
            print(f"[CONFIGURATOR WARN] Error cargando configuración: {e}")
            print("[CONFIGURATOR] Usando valores por defecto")
            # Cargar valores por defecto usando el método que ya funciona
            self.ventana_config.valores_por_defecto_new()

    def _poblar_ui_desde_config(self, config: dict):
        """Pobla todos los campos de la UI con datos del config"""
        try:
            # Carga de trabajo
            self.ventana_config.total_ordenes_var.set(config.get('total_ordenes', 300))

            distribucion = config.get('distribucion_tipos', {})
            pequeno = distribucion.get('pequeno', {'porcentaje': 60, 'volumen': 5})
            mediano = distribucion.get('mediano', {'porcentaje': 30, 'volumen': 25})
            grande = distribucion.get('grande', {'porcentaje': 10, 'volumen': 80})

            self.ventana_config.pct_pequeno.set(pequeno.get('porcentaje', 60))
            self.ventana_config.pct_mediano.set(mediano.get('porcentaje', 30))
            self.ventana_config.pct_grande.set(grande.get('porcentaje', 10))

            self.ventana_config.vol_pequeno.set(pequeno.get('volumen', 5))
            self.ventana_config.vol_mediano.set(mediano.get('volumen', 25))
            self.ventana_config.vol_grande.set(grande.get('volumen', 80))

            self.ventana_config.capacidad_carro.set(config.get('capacidad_carro', 150))

            # Recursos
            self.ventana_config.num_operarios_terrestres.set(config.get('num_operarios_terrestres', 1))
            self.ventana_config.num_montacargas.set(config.get('num_montacargas', 1))
            self.ventana_config.capacidad_montacargas.set(config.get('capacidad_montacargas', 1000))
            self.ventana_config.tiempo_descarga_por_tarea.set(config.get('tiempo_descarga_por_tarea', 5))

            # Estrategias (2 tipos)
            self.ventana_config.dispatch_strategy_var.set(config.get('dispatch_strategy', 'Optimizacion Global'))
            self.ventana_config.tour_type_var.set(config.get('tour_type', 'Tour Mixto (Multi-Destino)'))

            # Layout y archivos
            self.ventana_config.layout_path_var.set(config.get('layout_file', 'layouts/WH1.tmx'))
            self.ventana_config.sequence_path_var.set(config.get('sequence_file', 'layouts/Warehouse_Logic.xlsx'))
            self.ventana_config.map_scale_var.set(config.get('map_scale', 1.3))
            self.ventana_config.resolution_var.set(config.get('selected_resolution_key', 'Pequena (800x800)'))

            # Asignacion de recursos eliminada - ahora está en "Flota de Agentes"

            # Distribucion de OutboundStaging
            outbound_staging_distribution = config.get('outbound_staging_distribution', {
                "1": 100, "2": 0, "3": 0, "4": 0, "5": 0, "6": 0, "7": 0
            })
            for staging_id, percentage in outbound_staging_distribution.items():
                if staging_id in self.ventana_config.outbound_staging_vars:
                    self.ventana_config.outbound_staging_vars[staging_id].set(percentage)
            self.ventana_config._validar_staging_distribution()

            # Actualizar validaciones y resumenes
            self.ventana_config.validar_porcentajes()
            self.ventana_config.actualizar_total()

            # NUEVA FUNCIONALIDAD: Carga de Flota de Agentes desde config.json
            # Primero intentar cargar agent_types (formato original)
            agent_types = config.get('agent_types', [])
            
            # Si no hay agent_types, intentar convertir desde agent_fleet (formato de slots)
            if not agent_types and 'agent_fleet' in config:
                print(f"[CONFIGURATOR] Convirtiendo agent_fleet a agent_types...")
                agent_fleet = config['agent_fleet']
                agent_types = self._convertir_agent_fleet_a_agent_types(agent_fleet)
                print(f"[CONFIGURATOR] Convertidos {len(agent_types)} agentes desde agent_fleet")
            
            if agent_types:
                print(f"[CONFIGURATOR] Cargando {len(agent_types)} agentes en UI de Flota...")
                grupos_para_ui = self._agrupar_agentes_para_ui(agent_types)
                self.ventana_config._poblar_ui_flota(grupos_para_ui)
                print(f"[CONFIGURATOR] EXITO - Flota cargada: {len(grupos_para_ui)} grupos creados")

            # Inicializacion inteligente
            self._inicializacion_inteligente(config)

        except Exception as e:
            print(f"[CONFIGURATOR ERROR] Error poblando UI desde config: {e}")
            import traceback
            traceback.print_exc()

    def _inicializacion_inteligente(self, config):
        """Inicializacion inteligente: Carga WorkAreas automaticamente"""
        try:
            print("[CONFIGURATOR] Iniciando inicializacion inteligente...")

            # PASO 1: Cargar WorkAreas automaticamente
            sequence_file = self.ventana_config.sequence_path_var.get()
            if sequence_file and os.path.exists(sequence_file):
                try:
                    print(f"[CONFIGURATOR] Cargando WorkAreas desde: {sequence_file}")
                    self.ventana_config._cargar_work_areas_automatico(sequence_file)
                    print(f"[CONFIGURATOR] WorkAreas cargadas: {self.ventana_config.available_work_areas}")
                except Exception as e:
                    print(f"[CONFIGURATOR] ADVERTENCIA - Error cargando WorkAreas: {e}")

            # PASO 2: Generar flota por defecto si no existe
            # Primero intentar cargar agent_types (formato original)
            agent_types = config.get('agent_types', [])
            
            # Si no hay agent_types, intentar convertir desde agent_fleet (formato de slots)
            if not agent_types and 'agent_fleet' in config:
                print(f"[CONFIGURATOR] Convirtiendo agent_fleet a agent_types en inicializacion...")
                agent_fleet = config['agent_fleet']
                agent_types = self._convertir_agent_fleet_a_agent_types(agent_fleet)
                print(f"[CONFIGURATOR] Convertidos {len(agent_types)} agentes desde agent_fleet")
            
            if not agent_types and self.ventana_config.available_work_areas:
                print("[CONFIGURATOR] Generando flota por defecto silenciosa...")
                try:
                    config_defecto = self.ventana_config._generar_config_defecto()
                    self.ventana_config._limpiar_todos_los_grupos()
                    for group_config in config_defecto:
                        self.ventana_config._crear_grupo_desde_config(group_config)
                    print("[CONFIGURATOR] Flota por defecto generada")
                except Exception as e:
                    print(f"[CONFIGURATOR] Error generando flota: {e}")

        except Exception as e:
            print(f"[CONFIGURATOR ERROR] Error en inicializacion inteligente: {e}")

    def _convertir_agent_fleet_a_agent_types(self, agent_fleet):
        """Convierte agent_fleet (formato de slots) a agent_types (formato original)"""
        agent_types = []
        
        for group in agent_fleet:
            agent_type = group['agent_type']
            cantidad = group['cantidad']
            capacidad = group['capacidad']
            tiempo_descarga = group['tiempo_descarga']
            
            # Convertir priorities a work_area_priorities
            work_area_priorities = {}
            for priority_data in group['priorities']:
                wa = priority_data['wa']
                priority = priority_data['priority']
                work_area_priorities[wa] = priority
            
            # Crear agentes individuales
            for _ in range(cantidad):
                agent_types.append({
                    'type': agent_type,
                    'capacity': capacidad,
                    'discharge_time': tiempo_descarga,
                    'work_area_priorities': work_area_priorities.copy()
                })
        
        return agent_types

    def _agrupar_agentes_para_ui(self, agent_types_list):
        """Agrupa agent_types del JSON en formato para _poblar_ui_flota"""
        grupos = {}

        for agent in agent_types_list:
            agent_type = agent.get('type', 'GroundOperator')
            capacity = agent.get('capacity', 150)
            discharge_time = agent.get('discharge_time', 5)
            work_area_priorities = agent.get('work_area_priorities', {})

            # Crear clave unica para agrupar agentes similares
            priority_key = tuple(sorted(work_area_priorities.items()))
            group_key = (agent_type, capacity, discharge_time, priority_key)

            if group_key not in grupos:
                grupos[group_key] = {
                    'agent_type': agent_type,
                    'cantidad': 0,
                    'capacidad': capacity,
                    'tiempo_descarga': discharge_time,
                    'work_area_priorities': work_area_priorities
                }

            grupos[group_key]['cantidad'] += 1

        return list(grupos.values())

    def guardar_configuracion(self):
        """Guarda la configuracion actual en config.json"""
        try:
            # Validar configuracion antes de guardar
            if not self._validar_configuracion_actual():
                return

            # Obtener configuracion de la UI
            config = self._obtener_configuracion_ui()

            # Guardar en archivo JSON
            config_path = os.path.join(os.path.dirname(__file__), "config.json")

            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=4, ensure_ascii=False)

            messagebox.showinfo(
                "Configuracion Guardada",
                f"La configuracion se ha guardado exitosamente en:\n{config_path}\n\n"
                f"Ahora puedes ejecutar 'python run_simulator.py' directamente."
            )
            print(f"[CONFIGURATOR] Configuracion guardada en: {config_path}")

        except Exception as e:
            messagebox.showerror(
                "Error al Guardar",
                f"No se pudo guardar la configuracion:\n{str(e)}"
            )
            print(f"[CONFIGURATOR ERROR] Error guardando configuracion: {e}")

    def cargar_configuracion_manual(self):
        """Carga la configuracion desde config.json manualmente (por solicitud del usuario)"""
        print("[CONFIGURATOR] *** BOTON CARGAR PRESIONADO ***")
        config_path = os.path.join(os.path.dirname(__file__), "config.json")
        print(f"[CONFIGURATOR] Buscando archivo en: {config_path}")

        if not os.path.exists(config_path):
            print("[CONFIGURATOR] ERROR: Archivo no encontrado")
            messagebox.showwarning(
                "Archivo No Encontrado",
                f"No se encontro el archivo de configuracion:\n{config_path}\n\n"
                "Use 'Guardar Configuracion' primero para crear el archivo."
            )
            return

        try:
            print(f"[CONFIGURATOR] Cargando configuracion desde: {config_path}")
            with open(config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)

            # Sanitizar assignment_rules: convertir claves str a int
            if 'assignment_rules' in config and config['assignment_rules']:
                sanitized_rules = {}
                for agent_type, rules in config['assignment_rules'].items():
                    sanitized_rules[agent_type] = {int(k): v for k, v in rules.items()}
                config['assignment_rules'] = sanitized_rules

            # Poblar UI
            self._poblar_ui_desde_config(config)

            messagebox.showinfo(
                "Configuracion Cargada",
                f"La configuracion se ha cargado exitosamente desde:\n{config_path}"
            )
            print("[CONFIGURATOR] Configuracion cargada exitosamente en UI")

        except (json.JSONDecodeError, KeyError, Exception) as e:
            messagebox.showerror(
                "Error al Cargar",
                f"No se pudo cargar la configuracion:\n{str(e)}\n\n"
                "Verifique que el archivo config.json sea valido."
            )
            print(f"[CONFIGURATOR ERROR] Error cargando config.json: {e}")

    def cargar_defaults(self):
        """Carga los valores por defecto en la UI"""
        try:
            self.ventana_config.valores_por_defecto_new()
            messagebox.showinfo(
                "Valores por Defecto",
                "Los valores por defecto han sido cargados exitosamente."
            )
            print("[CONFIGURATOR] Valores por defecto cargados")

        except Exception as e:
            messagebox.showerror(
                "Error al Cargar Defaults",
                f"No se pudieron cargar los valores por defecto:\n{str(e)}"
            )
            print(f"[CONFIGURATOR ERROR] Error cargando defaults: {e}")

    # Método salir removido - usar solo la cruz superior derecha

    def _validar_configuracion_actual(self) -> bool:
        """Valida la configuracion actual de la UI"""
        try:
            # Validaciones basicas
            total_ordenes = self.ventana_config.total_ordenes_var.get()
            pct_pequeno = self.ventana_config.pct_pequeno.get()
            pct_mediano = self.ventana_config.pct_mediano.get()
            pct_grande = self.ventana_config.pct_grande.get()
            vol_pequeno = self.ventana_config.vol_pequeno.get()
            vol_mediano = self.ventana_config.vol_mediano.get()
            vol_grande = self.ventana_config.vol_grande.get()
            capacidad_carro = self.ventana_config.capacidad_carro.get()
            op_terrestres = self.ventana_config.num_operarios_terrestres.get()
            montacargas = self.ventana_config.num_montacargas.get()

            # Validar configuracion de picking
            picking_valid = self.ventana_config._validar_configuracion_picking(
                total_ordenes, pct_pequeno, pct_mediano, pct_grande,
                vol_pequeno, vol_mediano, vol_grande, capacidad_carro,
                op_terrestres, montacargas, op_terrestres + montacargas
            )

            # Validar distribucion de OutboundStaging
            staging_valid = self.ventana_config._validar_staging_distribution()

            if not staging_valid:
                messagebox.showerror("Error de Configuracion",
                                   "La distribucion de Outbound Staging debe sumar exactamente 100%.")

            return picking_valid and staging_valid
        except Exception as e:
            messagebox.showerror("Error de Validacion", f"Error validando configuracion: {str(e)}")
            return False

    def _obtener_configuracion_ui(self) -> dict:
        """Obtiene la configuracion actual de la UI como diccionario"""
        # Obtener configuracion desde VentanaConfiguracion
        config = self.ventana_config.obtener_configuracion()

        # Anadir campos de compatibilidad
        config['tareas_zona_a'] = 0
        config['tareas_zona_b'] = 0

        # Convertir rutas a relativas
        config['layout_file'] = self._make_relative_path(config['layout_file'])
        config['sequence_file'] = self._make_relative_path(config['sequence_file'])

        return config

    def _make_relative_path(self, file_path: str) -> str:
        """Convierte rutas absolutas a relativas"""
        if not file_path:
            return file_path

        try:
            project_root = os.path.dirname(os.path.abspath(__file__))
            abs_path = os.path.abspath(file_path)

            if abs_path.startswith(project_root):
                relative_path = os.path.relpath(abs_path, project_root)
                return relative_path.replace('\\', '/')
            else:
                return abs_path.replace('\\', '/')

        except (ValueError, OSError):
            return file_path

    # ========================================================================
    # SISTEMA DE SLOTS DE CONFIGURACION - CALLBACKS
    # ========================================================================
    
    def _guardar_como_callback(self):
        """Callback para 'Save' - Abre dialogo de seleccion de modo"""
        try:
            print("[CONFIGURATOR] *** BOTON SAVE PRESIONADO ***")
            
            # Validar configuracion actual
            if not self._validar_configuracion_actual():
                return
            
            # Obtener configuracion de la UI
            config_data = self._obtener_configuracion_ui()
            
            # Abrir dialogo de seleccion de modo (New/Update)
            dialog = ConfigurationSaveModeDialog(
                parent=self.root,
                config_manager=self.config_manager,
                config_data=config_data
            )
            
            # Esperar resultado
            self.root.wait_window(dialog)
            
            if dialog.result:
                print(f"[CONFIGURATOR] Configuracion guardada con ID: {dialog.result}")
                messagebox.showinfo("Exito", "Configuracion guardada exitosamente en el sistema de slots")
            
        except Exception as e:
            print(f"[CONFIGURATOR ERROR] Error en guardar como: {e}")
            messagebox.showerror("Error", f"No se pudo guardar la configuracion: {e}")
    
    def _cargar_desde_callback(self):
        """Callback para 'Cargar Desde...' - Abre dialogo de carga"""
        try:
            print("[CONFIGURATOR] *** BOTON CARGAR DESDE PRESIONADO ***")
            
            # Abrir dialogo de carga
            dialog = ConfigurationDialog(
                parent=self.root,
                config_manager=self.config_manager,
                config_data={},  # No necesario para modo load
                mode="load"
            )
            
            # Esperar resultado
            self.root.wait_window(dialog)
            
            if dialog.result:
                # Cargar configuracion en la UI
                config_data = dialog.result["configuration"]
                self._poblar_ui_desde_config(config_data)
                
                print(f"[CONFIGURATOR] Configuracion cargada: {dialog.result['metadata']['name']}")
                messagebox.showinfo("Exito", f"Configuracion '{dialog.result['metadata']['name']}' cargada exitosamente")
            
        except Exception as e:
            print(f"[CONFIGURATOR ERROR] Error en cargar desde: {e}")
            messagebox.showerror("Error", f"No se pudo cargar la configuracion: {e}")
    
    def _eliminar_configuracion_callback(self):
        """Callback para 'Eliminar Configuracion' - Abre dialogo de gestion"""
        try:
            print("[CONFIGURATOR] *** BOTON ELIMINAR CONFIGURACION PRESIONADO ***")
            
            # Obtener iconos desde VentanaConfiguracion
            icons = getattr(self.ventana_config, '_icons', {})
            
            # Abrir dialogo de gestion
            dialog = ConfigurationManagerDialog(
                parent=self.root,
                config_manager=self.config_manager,
                icons=icons
            )
            
            # Esperar resultado (dialog se cierra automaticamente)
            self.root.wait_window(dialog)
            
            print("[CONFIGURATOR] Dialogo de gestion cerrado")
            
        except Exception as e:
            print(f"[CONFIGURATOR ERROR] Error en eliminar configuracion: {e}")
            messagebox.showerror("Error", f"No se pudo abrir el gestor de configuraciones: {e}")

    def _use_configuration_callback(self):
        """Callback para 'Use' - Selecciona configuración de slots y la aplica a config.json"""
        try:
            print("[CONFIGURATOR] *** BOTON USE CONFIGURACION PRESIONADO ***")
            
            # Obtener iconos desde VentanaConfiguracion
            icons = getattr(self.ventana_config, '_icons', {})
            
            # Abrir dialogo de selección (similar a Load pero para aplicar)
            dialog = ConfigurationDialog(
                parent=self.root,
                config_manager=self.config_manager,
                config_data=None,
                mode="use"
            )
            
            # Esperar resultado
            self.root.wait_window(dialog)
            
            if dialog.result:
                # Aplicar la configuración seleccionada a config.json
                config_data = dialog.result
                config_name = config_data['metadata']['name']
                
                # Guardar configuración en config.json
                self._guardar_configuracion_directa(config_data['configuration'])
                
                print(f"[CONFIGURATOR] Configuracion '{config_name}' aplicada a config.json")
                messagebox.showinfo("Exito", f"Configuracion '{config_name}' aplicada exitosamente a config.json")
            
        except Exception as e:
            print(f"[CONFIGURATOR ERROR] Error en use configuracion: {e}")
            messagebox.showerror("Error", f"No se pudo aplicar la configuracion: {e}")

    def _guardar_configuracion_directa(self, config_data):
        """Guarda configuración directamente en config.json"""
        try:
            config_path = "config.json"
            
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(config_data, f, indent=4, ensure_ascii=False)
            
            print(f"[CONFIGURATOR] Configuracion guardada en: {config_path}")
            
        except Exception as e:
            print(f"[CONFIGURATOR ERROR] Error guardando configuracion directa: {e}")
            raise

    def ejecutar(self):
        """Ejecuta el configurador"""
        print("[CONFIGURATOR] Iniciando configurador independiente...")
        print("[CONFIGURATOR] Use 'Use' para aplicar configuracion de slots a config.json")
        print("[CONFIGURATOR] Use 'Save' para guardar en sistema de slots")
        print("[CONFIGURATOR] Use 'Load' para cargar desde sistema de slots")

        # Callbacks ya conectados en __init__
        self.root.mainloop()


# ========================================================================
# SISTEMA DE SLOTS DE CONFIGURACION - FASE 2: INFRAESTRUCTURA
# ========================================================================

class ConfigurationManager:
    """
    Gestor principal del sistema de slots de configuracion.
    Maneja la logica de negocio para gestionar configuraciones nombradas.
    Optimizado con cache para mejor rendimiento.
    """
    
    def __init__(self, config_dir="configurations"):
        """Inicializa el ConfigurationManager con cache optimizado"""
        self.config_dir = config_dir
        self.index_file = os.path.join(config_dir, "index.json")
        self.backup_dir = os.path.join(config_dir, "backups")
        self.storage = ConfigurationStorage(config_dir)
        
        # Cache para optimizar rendimiento
        self._config_cache = {}  # Cache de configuraciones cargadas
        self._index_cache = None  # Cache del indice
        self._cache_timestamp = 0  # Timestamp de ultima actualizacion del cache
        
        self._ensure_directories()
        self._load_index_cache()  # Cargar cache inicial
        print(f"[CONFIGURATION_MANAGER] Inicializado con cache optimizado en directorio: {config_dir}")
    
    def _ensure_directories(self):
        """Crea los directorios necesarios si no existen"""
        os.makedirs(self.config_dir, exist_ok=True)
        os.makedirs(self.backup_dir, exist_ok=True)
        print(f"[CONFIGURATION_MANAGER] Directorios creados: {self.config_dir}, {self.backup_dir}")
    
    def _load_index_cache(self):
        """Carga el indice en cache para optimizar rendimiento"""
        try:
            if os.path.exists(self.index_file):
                with open(self.index_file, 'r', encoding='utf-8') as f:
                    self._index_cache = json.load(f)
                self._cache_timestamp = os.path.getmtime(self.index_file)
                print(f"[CONFIGURATION_MANAGER] Cache de indice cargado: {len(self._index_cache.get('configurations', []))} configuraciones")
            else:
                self._index_cache = {"version": "1.0", "configurations": []}
                print("[CONFIGURATION_MANAGER] Cache de indice inicializado vacio")
        except Exception as e:
            print(f"[CONFIGURATION_MANAGER] Error cargando cache de indice: {e}")
            self._index_cache = {"version": "1.0", "configurations": []}
    
    def _invalidate_cache(self):
        """Invalida el cache cuando se realizan cambios"""
        self._config_cache.clear()
        self._load_index_cache()
        print("[CONFIGURATION_MANAGER] Cache invalidado y recargado")
    
    def _get_cached_configuration(self, config_id):
        """Obtiene una configuracion del cache si esta disponible"""
        if config_id in self._config_cache:
            return self._config_cache[config_id]
        return None
    
    def save_configuration(self, name, description, tags, config_data, is_default=False, overwrite=False, target_config_id=None):
        """
        Guarda una nueva configuracion con metadatos
        
        Args:
            name (str): Nombre de la configuracion
            description (str): Descripcion de la configuracion
            tags (list): Lista de tags
            config_data (dict): Datos de configuracion
            is_default (bool): Si es configuracion por defecto
            overwrite (bool): Si sobrescribir configuracion existente
            target_config_id (str): ID de configuracion especifica para sobrescribir
        
        Returns:
            str: ID de la configuracion creada
        """
        try:
            # Validar nombre unico (solo si no se permite sobrescribir)
            if self._name_exists(name) and not overwrite:
                raise ValueError(f"Ya existe una configuracion con el nombre '{name}'")
            
            # Si se especifica target_config_id, usar ese ID para sobrescribir
            if target_config_id and overwrite:
                config_id = target_config_id
                # Eliminar configuracion existente
                self.delete_configuration(config_id, create_backup=True)
                print(f"[CONFIGURATION_MANAGER] Configuracion '{config_id}' eliminada para sobrescribir")
            elif self._name_exists(name) and overwrite:
                # Si existe y se permite sobrescribir, eliminar la anterior
                existing_configs = self.list_configurations()
                for config in existing_configs:
                    if config['name'].lower() == name.lower():
                        self.delete_configuration(config['id'], create_backup=True)
                        print(f"[CONFIGURATION_MANAGER] Configuracion existente '{name}' eliminada para sobrescribir")
                        break
                # Crear ID unico
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                config_id = f"config_{name.lower().replace(' ', '_')}_{timestamp}"
            else:
                # Crear ID unico
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                config_id = f"config_{name.lower().replace(' ', '_')}_{timestamp}"
            
            # Crear metadatos
            metadata = {
                "id": config_id,
                "name": name,
                "description": description,
                "created_at": datetime.now().isoformat(),
                "modified_at": datetime.now().isoformat(),
                "version": "1.0",
                "tags": tags if isinstance(tags, list) else [tags] if tags else [],
                "author": "Usuario",
                "is_default": is_default
            }
            
            # Crear estructura completa
            full_config = {
                "metadata": metadata,
                "configuration": config_data
            }
            
            # Guardar archivo
            config_file = os.path.join(self.config_dir, f"{config_id}.json")
            self.storage._save_config_file(config_file, full_config)
            
            # Actualizar indice
            self._update_index(metadata, config_file)
            
            # Si es por defecto, desmarcar otros
            if is_default:
                self._set_as_default(config_id)
            
            # Invalidar cache para reflejar cambios
            self._invalidate_cache()
            
            print(f"[CONFIGURATION_MANAGER] Configuracion guardada: {config_id}")
            return config_id
            
        except Exception as e:
            print(f"[CONFIGURATION_MANAGER ERROR] Error guardando configuracion: {e}")
            raise
    
    def load_configuration(self, config_id):
        """
        Carga una configuracion especifica usando cache optimizado
        
        Args:
            config_id (str): ID de la configuracion
        
        Returns:
            dict: Configuracion completa con metadatos
        """
        try:
            # Verificar cache primero
            cached_config = self._get_cached_configuration(config_id)
            if cached_config:
                print(f"[CONFIGURATION_MANAGER] Configuracion '{config_id}' cargada desde cache")
                return cached_config
            
            # Buscar en el indice (usar cache si esta disponible)
            config_info = None
            if self._index_cache is not None:
                configurations = self._index_cache.get("configurations", [])
            else:
                index = self.storage._load_index()
                configurations = index.get("configurations", [])
            
            # Buscar configuracion
            for config in configurations:
                if config["id"] == config_id:
                    config_info = config
                    break
            
            if not config_info:
                raise ValueError(f"Configuracion '{config_id}' no encontrada")
            
            # Cargar archivo
            config_file = config_info["file_path"]
            if not os.path.exists(config_file):
                raise FileNotFoundError(f"Archivo de configuracion no encontrado: {config_file}")
            
            with open(config_file, 'r', encoding='utf-8') as f:
                full_config = json.load(f)
                
            # Guardar en cache para futuras consultas
            self._config_cache[config_id] = full_config
            print(f"[CONFIGURATION_MANAGER] Configuracion '{config_id}' cargada y guardada en cache")
            
            print(f"[CONFIGURATION_MANAGER] Configuracion cargada: {config_id}")
            return full_config
            
        except Exception as e:
            print(f"[CONFIGURATION_MANAGER ERROR] Error cargando configuracion: {e}")
            raise
    
    def list_configurations(self):
        """
        Lista todas las configuraciones disponibles usando cache optimizado
        
        Returns:
            list: Lista de configuraciones con metadatos
        """
        try:
            # Usar cache si esta disponible
            if self._index_cache is not None:
                configurations = self._index_cache.get("configurations", [])
                
                # Ordenar por fecha de modificacion (mas reciente primero)
                configurations.sort(key=lambda x: x.get("modified_at", ""), reverse=True)
                
                print(f"[CONFIGURATION_MANAGER] Listando {len(configurations)} configuraciones desde cache")
                return configurations
            else:
                # Fallback a carga directa
                index = self.storage._load_index()
                configurations = index.get("configurations", [])
                configurations.sort(key=lambda x: x.get("modified_at", ""), reverse=True)
                
                print(f"[CONFIGURATION_MANAGER] Listando {len(configurations)} configuraciones desde archivo")
                return configurations
            
        except Exception as e:
            print(f"[CONFIGURATION_MANAGER ERROR] Error listando configuraciones: {e}")
            return []
    
    def delete_configuration(self, config_id, create_backup=True):
        """
        Elimina una configuracion
        
        Args:
            config_id (str): ID de la configuracion
            create_backup (bool): Si crear backup antes de eliminar
        
        Returns:
            bool: True si se elimino exitosamente
        """
        try:
            index = self.storage._load_index()
            config_info = None
            
            # Buscar en el indice
            for i, config in enumerate(index.get("configurations", [])):
                if config["id"] == config_id:
                    config_info = config
                    break
            
            if not config_info:
                raise ValueError(f"Configuracion '{config_id}' no encontrada")
            
            # Crear backup si se solicita
            if create_backup:
                self._create_backup(config_id)
            
            # Eliminar archivo
            config_file = config_info["file_path"]
            if os.path.exists(config_file):
                os.remove(config_file)
                print(f"[CONFIGURATION_MANAGER] Archivo eliminado: {config_file}")
            
            # Actualizar indice
            index["configurations"].pop(i)
            self.storage._save_index(index)
            
            # Invalidar cache para reflejar cambios
            self._invalidate_cache()
            
            print(f"[CONFIGURATION_MANAGER] Configuracion eliminada: {config_id}")
            return True
            
        except Exception as e:
            print(f"[CONFIGURATION_MANAGER ERROR] Error eliminando configuracion: {e}")
            raise
    
    def search_configurations(self, query):
        """
        Busca configuraciones por nombre, descripcion o tags
        
        Args:
            query (str): Termino de busqueda
        
        Returns:
            list: Lista de configuraciones que coinciden
        """
        try:
            configurations = self.list_configurations()
            query_lower = query.lower()
            results = []
            
            for config in configurations:
                # Buscar en nombre
                if query_lower in config.get("name", "").lower():
                    results.append(config)
                    continue
                
                # Buscar en descripcion
                if query_lower in config.get("description", "").lower():
                    results.append(config)
                    continue
                
                # Buscar en tags
                tags = config.get("tags", [])
                if any(query_lower in tag.lower() for tag in tags):
                    results.append(config)
                    continue
            
            print(f"[CONFIGURATION_MANAGER] Busqueda '{query}': {len(results)} resultados")
            return results
            
        except Exception as e:
            print(f"[CONFIGURATION_MANAGER ERROR] Error en busqueda: {e}")
            return []
    
    def _name_exists(self, name):
        """Verifica si ya existe una configuracion con ese nombre"""
        configurations = self.list_configurations()
        return any(config.get("name", "").lower() == name.lower() for config in configurations)
    
    def _update_index(self, metadata, config_file):
        """Actualiza el indice con una nueva configuracion"""
        index = self.storage._load_index()
        
        # Agregar entrada al indice
        index_entry = {
            "id": metadata["id"],
            "name": metadata["name"],
            "description": metadata["description"],
            "created_at": metadata["created_at"],
            "modified_at": metadata["modified_at"],
            "version": metadata["version"],
            "tags": metadata["tags"],
            "is_default": metadata["is_default"],
            "file_path": config_file
        }
        
        index["configurations"].append(index_entry)
        index["last_updated"] = datetime.now().isoformat()
        
        self.storage._save_index(index)
    
    def _set_as_default(self, config_id):
        """Marca una configuracion como por defecto y desmarca las demas"""
        index = self.storage._load_index()
        
        for config in index.get("configurations", []):
            config["is_default"] = (config["id"] == config_id)
        
        self.storage._save_index(index)
    
    def get_default_configuration(self):
        """
        Obtiene la configuracion marcada como por defecto
        
        Returns:
            dict: Configuracion completa con metadata y configuration, o None si no hay default
        """
        try:
            index = self.storage._load_index()
            configurations = index.get("configurations", [])
            
            # Buscar configuracion marcada como default
            for config in configurations:
                if config.get("is_default", False):
                    # Cargar configuracion completa
                    full_config = self.load_configuration(config["id"])
                    print(f"[CONFIGURATION_MANAGER] Configuracion por defecto encontrada: {config['name']}")
                    return full_config
            
            print("[CONFIGURATION_MANAGER] No hay configuracion por defecto")
            return None
            
        except Exception as e:
            print(f"[CONFIGURATION_MANAGER ERROR] Error obteniendo configuracion por defecto: {e}")
            return None
    
    def _create_backup(self, config_id):
        """Crea un backup de una configuracion"""
        try:
            config = self.load_configuration(config_id)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_file = os.path.join(self.backup_dir, f"{config_id}_backup_{timestamp}.json")
            
            with open(backup_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=4, ensure_ascii=False)
            
            print(f"[CONFIGURATION_MANAGER] Backup creado: {backup_file}")
            
        except Exception as e:
            print(f"[CONFIGURATION_MANAGER ERROR] Error creando backup: {e}")


class ConfigurationStorage:
    """
    Manejo de archivos del sistema de configuraciones.
    Serializacion/deserializacion JSON y gestion de indices.
    """
    
    def __init__(self, config_dir):
        """Inicializa ConfigurationStorage"""
        self.config_dir = config_dir
        self.index_file = os.path.join(config_dir, "index.json")
        self.backup_dir = os.path.join(config_dir, "backups")
        self._ensure_directories()
        print(f"[CONFIGURATION_STORAGE] Inicializado en: {config_dir}")
    
    def _ensure_directories(self):
        """Crea los directorios necesarios si no existen"""
        os.makedirs(self.config_dir, exist_ok=True)
        os.makedirs(self.backup_dir, exist_ok=True)
    
    def _load_index(self):
        """Carga el indice de configuraciones"""
        if os.path.exists(self.index_file):
            try:
                with open(self.index_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                print(f"[CONFIGURATION_STORAGE ERROR] Error cargando indice: {e}")
        
        # Crear indice vacio si no existe
        return {
            "version": "1.0",
            "last_updated": datetime.now().isoformat(),
            "configurations": []
        }
    
    def _save_index(self, index_data):
        """Guarda el indice de configuraciones"""
        try:
            with open(self.index_file, 'w', encoding='utf-8') as f:
                json.dump(index_data, f, indent=4, ensure_ascii=False)
            print(f"[CONFIGURATION_STORAGE] Indice guardado: {self.index_file}")
        except Exception as e:
            print(f"[CONFIGURATION_STORAGE ERROR] Error guardando indice: {e}")
            raise
    
    def _save_config_file(self, file_path, config_data):
        """Guarda un archivo de configuracion"""
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(config_data, f, indent=4, ensure_ascii=False)
            print(f"[CONFIGURATION_STORAGE] Archivo guardado: {file_path}")
        except Exception as e:
            print(f"[CONFIGURATION_STORAGE ERROR] Error guardando archivo: {e}")
            raise
    
    def _validate_configuration(self, config_data):
        """Valida la estructura de una configuracion"""
        try:
            # Validar estructura basica
            if not isinstance(config_data, dict):
                return False, "Configuracion debe ser un diccionario"
            
            if "metadata" not in config_data:
                return False, "Falta seccion 'metadata'"
            
            if "configuration" not in config_data:
                return False, "Falta seccion 'configuration'"
            
            metadata = config_data["metadata"]
            required_fields = ["id", "name", "description", "created_at", "version"]
            
            for field in required_fields:
                if field not in metadata:
                    return False, f"Falta campo requerido en metadata: {field}"
            
            return True, "Configuracion valida"
            
        except Exception as e:
            return False, f"Error validando configuracion: {e}"


# ========================================================================
# SISTEMA DE SLOTS DE CONFIGURACION - FASE 2.3: CONFIGURATION UI
# ========================================================================

class ConfigurationSaveModeDialog(tk.Toplevel):
    """
    Dialogo para seleccionar modo de guardado: New o Update
    """
    
    def __init__(self, parent, config_manager, config_data):
        """
        Inicializa el dialogo de seleccion de modo
        
        Args:
            parent: Ventana padre
            config_manager: Instancia de ConfigurationManager
            config_data: Datos de configuracion a guardar
        """
        super().__init__(parent)
        self.parent = parent
        self.config_manager = config_manager
        self.config_data = config_data
        self.result = None
        
        # Configurar ventana
        self.title("Seleccionar Modo de Guardado")
        self.geometry("400x200")
        self.resizable(False, False)
        
        # Centrar ventana
        self.transient(parent)
        self.grab_set()
        
        # Crear UI
        self._create_ui()
        
        # Aplicar tema oscuro si está activo
        self._apply_dialog_theme()
        
        # Centrar ventana en pantalla
        self._center_window()
        
        print("[CONFIGURATION_SAVE_MODE_DIALOG] Dialogo de modo de guardado creado")
    
    def _create_ui(self):
        """Crea la interfaz de usuario"""
        # Frame principal
        main_frame = ttk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Titulo
        title_label = ttk.Label(main_frame, text="Seleccionar Modo de Guardado", 
                               font=("Arial", 14, "bold"))
        title_label.pack(pady=(0, 20))
        
        # Descripcion
        desc_label = ttk.Label(main_frame, 
                              text="¿Como desea guardar la configuracion actual?",
                              font=("Arial", 10))
        desc_label.pack(pady=(0, 30))
        
        # Frame para botones
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=20)
        
        # Boton New
        new_button = ttk.Button(button_frame, text="New", 
                               command=self._on_new_clicked,
                               style="Accent.TButton")
        new_button.pack(side=tk.LEFT, padx=(0, 10), fill=tk.X, expand=True)
        
        # Boton Update
        update_button = ttk.Button(button_frame, text="Update", 
                                  command=self._on_update_clicked,
                                  style="Accent.TButton")
        update_button.pack(side=tk.LEFT, padx=(10, 0), fill=tk.X, expand=True)
        
        # Boton Cancelar
        cancel_button = ttk.Button(main_frame, text="Cancelar", 
                                  command=self._on_cancel_clicked)
        cancel_button.pack(pady=(20, 0))
    
    def _center_window(self):
        """Centra la ventana en la pantalla"""
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f"{width}x{height}+{x}+{y}")
    
    def _on_new_clicked(self):
        """Maneja click en boton New"""
        try:
            print("[CONFIGURATION_SAVE_MODE_DIALOG] Modo New seleccionado")
            
            # Abrir dialogo de guardado normal (New)
            dialog = ConfigurationDialog(
                parent=self.parent,
                config_manager=self.config_manager,
                config_data=self.config_data,
                mode="save"
            )
            
            # Esperar resultado
            self.parent.wait_window(dialog)
            
            if dialog.result:
                self.result = dialog.result
                self.destroy()
            
        except Exception as e:
            print(f"[CONFIGURATION_SAVE_MODE_DIALOG ERROR] Error en modo New: {e}")
            messagebox.showerror("Error", f"No se pudo abrir el dialogo de guardado: {e}")
    
    def _on_update_clicked(self):
        """Maneja click en boton Update"""
        try:
            print("[CONFIGURATION_SAVE_MODE_DIALOG] Modo Update seleccionado")
            
            # Abrir dialogo de sobrescritura
            dialog = ConfigurationOverwriteDialog(
                parent=self.parent,
                config_manager=self.config_manager,
                config_data=self.config_data
            )
            
            # Esperar resultado
            self.parent.wait_window(dialog)
            
            if dialog.result:
                self.result = dialog.result
                self.destroy()
            
        except Exception as e:
            print(f"[CONFIGURATION_SAVE_MODE_DIALOG ERROR] Error en modo Update: {e}")
            messagebox.showerror("Error", f"No se pudo abrir el dialogo de sobrescritura: {e}")
    
    def _on_cancel_clicked(self):
        """Maneja click en boton Cancelar"""
        print("[CONFIGURATION_SAVE_MODE_DIALOG] Operacion cancelada")
        self.destroy()
    
    def _apply_dialog_theme(self):
        """Aplica tema oscuro al diálogo si está activo en la ventana principal"""
        try:
            # Verificar si el tema oscuro está activo en la ventana principal
            if hasattr(self.parent, 'dark_mode') and self.parent.dark_mode:
                dark_colors = {
                    'bg_primary': '#1e1e1e',
                    'bg_secondary': '#2d2d30',
                    'text_primary': '#cccccc',
                    'accent': '#007acc'
                }
                
                # Aplicar tema oscuro al diálogo
                self.configure(bg=dark_colors['bg_primary'])
                
                # Aplicar recursivamente a todos los widgets
                self._apply_dark_theme_recursive(self, dark_colors)
                
        except Exception as e:
            print(f"[CONFIGURATION_SAVE_MODE_DIALOG] Error aplicando tema: {e}")
    
    def _apply_dark_theme_recursive(self, widget, dark_colors):
        """Aplica tema oscuro recursivamente a widgets del diálogo"""
        try:
            # Solo aplicar a widgets tk nativos específicos, NO a widgets ttk
            widget_class = widget.__class__.__name__
            
            # Verificar si es un widget tk nativo (no ttk)
            if widget_class.startswith('Tk') or widget_class in ['Entry', 'Text', 'Canvas', 'Listbox', 'Scrollbar', 'Toplevel', 'Label', 'Button']:
                if isinstance(widget, tk.Entry):
                    widget.configure(
                        bg=dark_colors['bg_secondary'],
                        fg=dark_colors['text_primary'],
                        insertbackground=dark_colors['text_primary'],
                        selectbackground=dark_colors['accent'],
                        selectforeground=dark_colors['text_primary']
                    )
                elif isinstance(widget, tk.Text):
                    widget.configure(
                        bg=dark_colors['bg_secondary'],
                        fg=dark_colors['text_primary'],
                        insertbackground=dark_colors['text_primary'],
                        selectbackground=dark_colors['accent'],
                        selectforeground=dark_colors['text_primary']
                    )
                elif isinstance(widget, tk.Label):
                    widget.configure(
                        bg=dark_colors['bg_primary'],
                        fg=dark_colors['text_primary']
                    )
                elif isinstance(widget, tk.Button):
                    widget.configure(
                        bg=dark_colors['bg_secondary'],
                        fg=dark_colors['text_primary'],
                        activebackground=dark_colors['accent'],
                        activeforeground=dark_colors['text_primary']
                    )
            
            # Recursivamente aplicar a widgets hijos
            for child in widget.winfo_children():
                self._apply_dark_theme_recursive(child, dark_colors)
                
        except Exception as e:
            # Silenciar errores de widgets que no soportan configuración
            pass


class ConfigurationDialog(tk.Toplevel):
    """
    Dialogo para guardar configuracion con nombre personalizado.
    Reemplaza el boton "Valores por Defecto" con funcionalidad avanzada.
    """
    
    def __init__(self, parent, config_manager, config_data, mode="save"):
        """
        Inicializa el dialogo de configuracion
        
        Args:
            parent: Ventana padre
            config_manager: Instancia de ConfigurationManager
            config_data: Datos de configuracion a guardar
            mode: Modo del dialogo ("save" o "load")
        """
        super().__init__(parent)
        self.parent = parent
        self.config_manager = config_manager
        self.config_data = config_data
        self.mode = mode
        self.result = None
        
        self._setup_window()
        self._create_widgets()
        self._center_window()
        self._apply_dialog_theme()
        
        # Hacer modal
        self.transient(parent)
        self.grab_set()
        
        print(f"[CONFIGURATION_DIALOG] Dialogo {mode} inicializado")
    
    def _setup_window(self):
        """Configura la ventana del dialogo"""
        if self.mode == "save":
            self.title("Guardar Configuracion")
            self.geometry("500x400")
        else:
            self.title("Cargar Configuracion")
            self.geometry("600x500")
        
        self.resizable(True, True)
    
    def _create_widgets(self):
        """Crea los widgets del dialogo"""
        if self.mode == "save":
            self._create_save_widgets()
        else:
            self._create_load_widgets()
    
    def _create_save_widgets(self):
        """Crea widgets para modo guardar con diseño mejorado"""
        # Frame principal con padding mejorado
        main_frame = ttk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=25, pady=25)
        
        # Titulo con icono
        title_frame = ttk.Frame(main_frame)
        title_frame.pack(fill=tk.X, pady=(0, 25))
        
        title_label = ttk.Label(title_frame, text="💾 Guardar Configuracion")
        title_label.pack(side=tk.LEFT)
        
        # Subtitulo descriptivo
        subtitle_label = ttk.Label(main_frame, text="Guarda la configuracion actual con un nombre personalizado y metadatos",
                                  foreground="gray")
        subtitle_label.pack(pady=(0, 20))
        
        # Campos de entrada con diseño mejorado
        fields_frame = ttk.LabelFrame(main_frame, text="📝 Informacion de la Configuracion", padding=15)
        fields_frame.pack(fill=tk.X, pady=(0, 20))
        
        # Configurar grid para mejor espaciado
        fields_frame.grid_columnconfigure(1, weight=1)
        
        # Nombre con icono
        name_label = ttk.Label(fields_frame, text="🏷️ Nombre:")
        name_label.grid(row=0, column=0, sticky=tk.W, padx=(0, 15), pady=8)
        self.name_var = tk.StringVar()
        self.name_entry = ttk.Entry(fields_frame, textvariable=self.name_var, width=45)
        self.name_entry.grid(row=0, column=1, padx=(0, 15), pady=8, sticky=tk.W+tk.E)
        
        # Descripcion con icono
        desc_label = ttk.Label(fields_frame, text="📄 Descripcion:")
        desc_label.grid(row=1, column=0, sticky=tk.W, padx=(0, 15), pady=8)
        self.description_var = tk.StringVar()
        self.description_entry = ttk.Entry(fields_frame, textvariable=self.description_var, width=45)
        self.description_entry.grid(row=1, column=1, padx=(0, 15), pady=8, sticky=tk.W+tk.E)
        
        # Tags con icono
        tags_label = ttk.Label(fields_frame, text="🏷️ Tags:")
        tags_label.grid(row=2, column=0, sticky=tk.W, padx=(0, 15), pady=8)
        self.tags_var = tk.StringVar()
        self.tags_entry = ttk.Entry(fields_frame, textvariable=self.tags_var, width=45)
        self.tags_entry.grid(row=2, column=1, padx=(0, 15), pady=8, sticky=tk.W+tk.E)
        
        # Label de ayuda para tags
        tags_help = ttk.Label(fields_frame, text="(separados por comas, ej: produccion, optimizado)", 
                             foreground="gray")
        tags_help.grid(row=3, column=1, sticky=tk.W, padx=(0, 15), pady=(0, 8))
        
        # Checkbox por defecto con icono
        self.is_default_var = tk.BooleanVar()
        self.is_default_check = ttk.Checkbutton(fields_frame, text="⭐ Marcar como configuracion por defecto",
                                               variable=self.is_default_var)
        self.is_default_check.grid(row=4, column=0, columnspan=2, padx=(0, 15), pady=15, sticky=tk.W)
        
        # Boton sobrescribir con icono
        self.overwrite_button = ttk.Button(fields_frame, text="🔄 Seleccionar Configuracion para Sobrescribir",
                                          command=self._open_overwrite_dialog, width=50)
        self.overwrite_button.grid(row=5, column=0, columnspan=2, padx=(0, 15), pady=10, sticky=tk.W+tk.E)
        
        # Label para mostrar configuracion seleccionada
        self.overwrite_label = ttk.Label(fields_frame, text="", foreground="blue")
        self.overwrite_label.grid(row=5, column=0, columnspan=2, padx=10, pady=5, sticky=tk.W)
        
        # Variable para almacenar configuracion seleccionada
        self.selected_config_id = None
        
        # Configurar grid
        fields_frame.columnconfigure(1, weight=1)
        
        # Vista previa
        preview_frame = ttk.LabelFrame(main_frame, text="Vista Previa")
        preview_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        
        self.preview_text = tk.Text(preview_frame, height=8, wrap=tk.WORD)
        preview_scrollbar = ttk.Scrollbar(preview_frame, orient=tk.VERTICAL, command=self.preview_text.yview)
        self.preview_text.configure(yscrollcommand=preview_scrollbar.set)
        
        self.preview_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10, 0), pady=10)
        preview_scrollbar.pack(side=tk.RIGHT, fill=tk.Y, pady=10)
        
        # Mostrar vista previa
        self._update_preview()
        
        # Botones
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X)
        
        ttk.Button(button_frame, text="Guardar", command=self._save_configuration).pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(button_frame, text="Cancelar", command=self._cancel).pack(side=tk.RIGHT)
    
    def _create_load_widgets(self):
        """Crea widgets para modo cargar"""
        # Frame principal
        main_frame = ttk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Titulo
        title_label = ttk.Label(main_frame, text="Cargar Configuracion", 
                               font=("Arial", 14, "bold"))
        title_label.pack(pady=(0, 20))
        
        # Busqueda
        search_frame = ttk.Frame(main_frame)
        search_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(search_frame, text="Buscar:").pack(side=tk.LEFT, padx=(0, 5))
        self.search_var = tk.StringVar()
        self.search_entry = ttk.Entry(search_frame, textvariable=self.search_var, width=30)
        self.search_entry.pack(side=tk.LEFT, padx=(0, 10))
        self.search_entry.bind('<KeyRelease>', self._on_search_change)
        
        # Lista de configuraciones
        list_frame = ttk.LabelFrame(main_frame, text="Configuraciones Disponibles")
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        
        # Treeview para lista
        columns = ("Nombre", "Descripcion", "Fecha", "Tags")
        self.config_tree = ttk.Treeview(list_frame, columns=columns, show="tree headings", height=10)
        
        # Configurar columnas
        self.config_tree.heading("#0", text="")
        self.config_tree.column("#0", width=0, stretch=False)
        
        for col in columns:
            self.config_tree.heading(col, text=col)
            if col == "Nombre":
                self.config_tree.column(col, width=150)
            elif col == "Descripcion":
                self.config_tree.column(col, width=200)
            elif col == "Fecha":
                self.config_tree.column(col, width=120)
            else:
                self.config_tree.column(col, width=100)
        
        # Scrollbar para lista
        list_scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.config_tree.yview)
        self.config_tree.configure(yscrollcommand=list_scrollbar.set)
        
        self.config_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10, 0), pady=10)
        list_scrollbar.pack(side=tk.RIGHT, fill=tk.Y, pady=10)
        
        # Bind seleccion
        self.config_tree.bind('<<TreeviewSelect>>', self._on_config_select)
        
        # Vista previa
        preview_frame = ttk.LabelFrame(main_frame, text="Vista Previa")
        preview_frame.pack(fill=tk.X, pady=(0, 20))
        
        self.preview_text = tk.Text(preview_frame, height=6, wrap=tk.WORD)
        preview_scrollbar = ttk.Scrollbar(preview_frame, orient=tk.VERTICAL, command=self.preview_text.yview)
        self.preview_text.configure(yscrollcommand=preview_scrollbar.set)
        
        self.preview_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10, 0), pady=10)
        preview_scrollbar.pack(side=tk.RIGHT, fill=tk.Y, pady=10)
        
        # Botones
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X)
        
        ttk.Button(button_frame, text="Cargar", command=self._load_configuration).pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(button_frame, text="Cancelar", command=self._cancel).pack(side=tk.RIGHT)
        
        # Cargar configuraciones
        self._load_configurations_list()
    
    def _open_overwrite_dialog(self):
        """Abre el dialogo para seleccionar configuracion a sobrescribir"""
        try:
            # Verificar si hay configuraciones existentes
            existing_configs = self.config_manager.list_configurations()
            if not existing_configs:
                messagebox.showinfo("Informacion", "No hay configuraciones existentes para sobrescribir")
                return
            
            # Abrir dialogo de seleccion
            dialog = ConfigurationOverwriteDialog(self, self.config_manager)
            self.wait_window(dialog)
            
            if dialog.result:
                self.selected_config_id = dialog.result
                # Actualizar label para mostrar configuracion seleccionada
                config_name = None
                for config in existing_configs:
                    if config['id'] == dialog.result:
                        config_name = config['name']
                        break
                
                if config_name:
                    self.overwrite_label.config(text=f"Configuracion seleccionada: {config_name}")
                    self.overwrite_button.config(text="Cambiar Configuracion Seleccionada")
                else:
                    self.overwrite_label.config(text="Error: Configuracion no encontrada")
            
        except Exception as e:
            print(f"[CONFIGURATION_DIALOG ERROR] Error abriendo dialogo de sobrescritura: {e}")
            messagebox.showerror("Error", f"Error abriendo dialogo de sobrescritura: {e}")
    
    def _update_preview(self):
        """Actualiza la vista previa de la configuracion"""
        try:
            preview_text = json.dumps(self.config_data, indent=2, ensure_ascii=False)
            self.preview_text.delete(1.0, tk.END)
            self.preview_text.insert(1.0, preview_text)
        except Exception as e:
            self.preview_text.delete(1.0, tk.END)
            self.preview_text.insert(1.0, f"Error generando vista previa: {e}")
    
    def _load_configurations_list(self):
        """Carga la lista de configuraciones disponibles"""
        try:
            configurations = self.config_manager.list_configurations()
            
            # Limpiar lista
            for item in self.config_tree.get_children():
                self.config_tree.delete(item)
            
            # Agregar configuraciones
            for config in configurations:
                # Formatear fecha
                created_at = config.get("created_at", "")
                if created_at:
                    try:
                        from datetime import datetime
                        dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                        formatted_date = dt.strftime("%Y-%m-%d %H:%M")
                    except:
                        formatted_date = created_at[:16]
                else:
                    formatted_date = "N/A"
                
                # Formatear tags
                tags = config.get("tags", [])
                tags_text = ", ".join(tags) if tags else "Sin tags"
                
                # Insertar en treeview
                self.config_tree.insert("", tk.END, values=(
                    config.get("name", "Sin nombre"),
                    config.get("description", "Sin descripcion"),
                    formatted_date,
                    tags_text
                ), tags=(config["id"],))
            
            print(f"[CONFIGURATION_DIALOG] Cargadas {len(configurations)} configuraciones")
            
        except Exception as e:
            print(f"[CONFIGURATION_DIALOG ERROR] Error cargando lista: {e}")
            messagebox.showerror("Error", f"No se pudieron cargar las configuraciones: {e}")
    
    def _on_search_change(self, event):
        """Maneja cambios en el campo de busqueda"""
        query = self.search_var.get().strip()
        
        if not query:
            self._load_configurations_list()
            return
        
        try:
            results = self.config_manager.search_configurations(query)
            
            # Limpiar lista
            for item in self.config_tree.get_children():
                self.config_tree.delete(item)
            
            # Agregar resultados
            for config in results:
                created_at = config.get("created_at", "")
                if created_at:
                    try:
                        from datetime import datetime
                        dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                        formatted_date = dt.strftime("%Y-%m-%d %H:%M")
                    except:
                        formatted_date = created_at[:16]
                else:
                    formatted_date = "N/A"
                
                tags = config.get("tags", [])
                tags_text = ", ".join(tags) if tags else "Sin tags"
                
                self.config_tree.insert("", tk.END, values=(
                    config.get("name", "Sin nombre"),
                    config.get("description", "Sin descripcion"),
                    formatted_date,
                    tags_text
                ), tags=(config["id"],))
            
        except Exception as e:
            print(f"[CONFIGURATION_DIALOG ERROR] Error en busqueda: {e}")
    
    def _on_config_select(self, event):
        """Maneja seleccion de configuracion"""
        selection = self.config_tree.selection()
        if not selection:
            return
        
        try:
            item = self.config_tree.item(selection[0])
            config_id = item["tags"][0]
            
            # Cargar configuracion completa
            full_config = self.config_manager.load_configuration(config_id)
            
            # Mostrar vista previa
            preview_text = json.dumps(full_config["configuration"], indent=2, ensure_ascii=False)
            self.preview_text.delete(1.0, tk.END)
            self.preview_text.insert(1.0, preview_text)
            
            # Guardar configuracion seleccionada
            self.selected_config = full_config
            
        except Exception as e:
            print(f"[CONFIGURATION_DIALOG ERROR] Error cargando configuracion: {e}")
            messagebox.showerror("Error", f"No se pudo cargar la configuracion: {e}")
    
    def _save_configuration(self):
        """Guarda la configuracion"""
        try:
            # Validar campos requeridos
            name = self.name_var.get().strip()
            if not name:
                messagebox.showerror("Error", "El nombre es requerido")
                return
            
            description = self.description_var.get().strip()
            tags_text = self.tags_var.get().strip()
            tags = [tag.strip() for tag in tags_text.split(",") if tag.strip()] if tags_text else []
            is_default = self.is_default_var.get()
            
            # Verificar si se selecciono una configuracion para sobrescribir
            if self.selected_config_id:
                # Sobrescribir configuracion existente
                config_id = self.config_manager.save_configuration(
                    name=name,
                    description=description,
                    tags=tags,
                    config_data=self.config_data,
                    is_default=is_default,
                    overwrite=True,
                    target_config_id=self.selected_config_id
                )
                messagebox.showinfo("Exito", f"Configuracion '{name}' sobrescrita exitosamente")
            else:
                # Verificar si existe una configuracion con el mismo nombre
                existing_configs = self.config_manager.list_configurations()
                existing_names = [config['name'].lower() for config in existing_configs]
                
                if name.lower() in existing_names:
                    # Preguntar si quiere sobrescribir
                    result = messagebox.askyesno("Configuracion Existente", 
                                               f"Ya existe una configuracion con el nombre '{name}'.\n"
                                               f"¿Desea sobrescribirla?")
                    if not result:
                        return
                    
                    # Encontrar y sobrescribir la configuracion existente
                    for config in existing_configs:
                        if config['name'].lower() == name.lower():
                            self.config_manager.delete_configuration(config['id'], create_backup=True)
                            break
                
                # Guardar nueva configuracion
                config_id = self.config_manager.save_configuration(
                    name=name,
                    description=description,
                    tags=tags,
                    config_data=self.config_data,
                    is_default=is_default,
                    overwrite=False
                )
                messagebox.showinfo("Exito", f"Configuracion '{name}' guardada exitosamente")
            
            self.result = config_id
            self.destroy()
            
        except Exception as e:
            print(f"[CONFIGURATION_DIALOG ERROR] Error guardando: {e}")
            messagebox.showerror("Error", f"No se pudo guardar la configuracion: {e}")
    
    def _load_configuration(self):
        """Carga la configuracion seleccionada"""
        try:
            if not hasattr(self, 'selected_config'):
                messagebox.showwarning("Advertencia", "Seleccione una configuracion primero")
                return
            
            self.result = self.selected_config
            messagebox.showinfo("Exito", f"Configuracion '{self.selected_config['metadata']['name']}' cargada")
            self.destroy()
            
        except Exception as e:
            print(f"[CONFIGURATION_DIALOG ERROR] Error cargando: {e}")
            messagebox.showerror("Error", f"No se pudo cargar la configuracion: {e}")
    
    def _cancel(self):
        """Cancela el dialogo"""
        self.result = None
        self.destroy()
    
    def _center_window(self):
        """Centra la ventana en la pantalla"""
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f"{width}x{height}+{x}+{y}")
    
    def _apply_dialog_theme(self):
        """Aplica tema oscuro al diálogo si está activo en la ventana principal"""
        try:
            # Verificar si el tema oscuro está activo en la ventana principal
            if hasattr(self.parent, 'dark_mode') and self.parent.dark_mode:
                dark_colors = {
                    'bg_primary': '#1e1e1e',
                    'bg_secondary': '#2d2d30',
                    'text_primary': '#cccccc',
                    'accent': '#007acc'
                }
                
                # Aplicar tema oscuro al diálogo
                self.configure(bg=dark_colors['bg_primary'])
                
                # Aplicar recursivamente a todos los widgets
                self._apply_dark_theme_recursive(self, dark_colors)
                
        except Exception as e:
            print(f"[CONFIGURATION_DIALOG] Error aplicando tema: {e}")
    
    def _apply_dark_theme_recursive(self, widget, dark_colors):
        """Aplica tema oscuro recursivamente a widgets del diálogo"""
        try:
            # Solo aplicar a widgets tk nativos específicos, NO a widgets ttk
            widget_class = widget.__class__.__name__
            
            # Verificar si es un widget tk nativo (no ttk)
            if widget_class.startswith('Tk') or widget_class in ['Entry', 'Text', 'Canvas', 'Listbox', 'Scrollbar', 'Toplevel', 'Label', 'Button']:
                if isinstance(widget, tk.Entry):
                    widget.configure(
                        bg=dark_colors['bg_secondary'],
                        fg=dark_colors['text_primary'],
                        insertbackground=dark_colors['text_primary'],
                        selectbackground=dark_colors['accent'],
                        selectforeground=dark_colors['text_primary']
                    )
                elif isinstance(widget, tk.Text):
                    widget.configure(
                        bg=dark_colors['bg_secondary'],
                        fg=dark_colors['text_primary'],
                        insertbackground=dark_colors['text_primary'],
                        selectbackground=dark_colors['accent'],
                        selectforeground=dark_colors['text_primary']
                    )
                elif isinstance(widget, tk.Label):
                    widget.configure(
                        bg=dark_colors['bg_primary'],
                        fg=dark_colors['text_primary']
                    )
                elif isinstance(widget, tk.Button):
                    widget.configure(
                        bg=dark_colors['bg_secondary'],
                        fg=dark_colors['text_primary'],
                        activebackground=dark_colors['accent'],
                        activeforeground=dark_colors['text_primary']
                    )
            
            # Recursivamente aplicar a widgets hijos
            for child in widget.winfo_children():
                self._apply_dark_theme_recursive(child, dark_colors)
                
        except Exception as e:
            # Silenciar errores de widgets que no soportan configuración
            pass


class ConfigurationOverwriteDialog(tk.Toplevel):
    """
    Dialogo para seleccionar configuracion existente para sobrescribir.
    Muestra todas las configuraciones con metadatos completos.
    """
    
    def __init__(self, parent, config_manager, config_data=None):
        super().__init__(parent)
        self.parent = parent
        self.config_manager = config_manager
        self.config_data = config_data
        self.result = None
        
        self._setup_window()
        self._create_widgets()
        self._center_window()
        self._apply_dialog_theme()
        
        self.transient(parent)
        self.grab_set()
        
        print("[CONFIGURATION_OVERWRITE_DIALOG] Dialogo de sobrescritura inicializado")
    
    def _setup_window(self):
        """Configura la ventana"""
        self.title("Seleccionar Configuracion para Sobrescribir")
        self.geometry("800x600")
        self.resizable(True, True)
    
    def _create_widgets(self):
        """Crea los widgets del dialogo con diseño mejorado"""
        # Frame principal con padding mejorado
        main_frame = ttk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=25, pady=25)
        
        # Titulo con icono
        title_label = ttk.Label(main_frame, text="🔄 Seleccionar Configuracion para Sobrescribir")
        title_label.pack(pady=(0, 15))
        
        # Descripcion mejorada
        desc_label = ttk.Label(main_frame, 
                              text="Selecciona una configuracion existente para sobrescribirla con la nueva configuracion.\nSe creara un backup automatico antes de sobrescribir.",
                              foreground="gray", justify=tk.CENTER)
        desc_label.pack(pady=(0, 20))
        
        # Frame para la lista con diseño mejorado
        list_frame = ttk.LabelFrame(main_frame, text="📋 Configuraciones Existentes", padding=10)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        
        # Treeview para mostrar configuraciones con estilo mejorado
        columns = ("Nombre", "Descripcion", "Tags", "Fecha Creacion", "Es Default")
        self.tree = ttk.Treeview(list_frame, columns=columns, show="headings", height=10)
        
        # Configurar columnas con iconos
        self.tree.heading("Nombre", text="🏷️ Nombre")
        self.tree.heading("Descripcion", text="📄 Descripcion")
        self.tree.heading("Tags", text="🏷️ Tags")
        self.tree.heading("Fecha Creacion", text="📅 Fecha Creacion")
        self.tree.heading("Es Default", text="⭐ Es Default")
        
        # Configurar anchos optimizados
        self.tree.column("Nombre", width=180, minwidth=150)
        self.tree.column("Descripcion", width=250, minwidth=200)
        self.tree.column("Tags", width=180, minwidth=150)
        self.tree.column("Fecha Creacion", width=140, minwidth=120)
        self.tree.column("Es Default", width=100, minwidth=80)
        
        # Scrollbar con estilo mejorado
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        # Pack treeview y scrollbar con padding
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Bind seleccion
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        
        # Frame para información de selección
        info_frame = ttk.Frame(main_frame)
        info_frame.pack(fill=tk.X, pady=(0, 15))
        
        self.selection_info = ttk.Label(info_frame, text="Selecciona una configuracion de la lista", 
                                       foreground="blue")
        self.selection_info.pack(side=tk.LEFT)
        
        # Frame de botones
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(10, 0))
        
        # Botones
        ttk.Button(button_frame, text="Sobrescribir Seleccionada",
                  command=self._overwrite_selected).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Cancelar",
                  command=self._cancel).pack(side=tk.RIGHT, padx=5)
        
        # Cargar configuraciones
        self._load_configurations()
    
    def _load_configurations(self):
        """Carga las configuraciones existentes"""
        try:
            configs = self.config_manager.list_configurations()
            
            # Limpiar treeview
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            # Agregar configuraciones
            for config in configs:
                tags_str = ", ".join(config.get('tags', []))
                fecha = config.get('created_at', 'N/A')
                if fecha != 'N/A':
                    try:
                        from datetime import datetime
                        fecha_obj = datetime.fromisoformat(fecha.replace('Z', '+00:00'))
                        fecha = fecha_obj.strftime("%Y-%m-%d %H:%M")
                    except:
                        pass
                
                es_default = "Si" if config.get('is_default', False) else "No"
                
                self.tree.insert("", "end", values=(
                    config['name'],
                    config.get('description', 'Sin descripcion'),
                    tags_str,
                    fecha,
                    es_default
                ), tags=(config['id'],))
            
            print(f"[CONFIGURATION_OVERWRITE_DIALOG] Cargadas {len(configs)} configuraciones")
            
        except Exception as e:
            print(f"[CONFIGURATION_OVERWRITE_DIALOG ERROR] Error cargando configuraciones: {e}")
    
    def _on_select(self, event):
        """Maneja la seleccion de una configuracion"""
        selection = self.tree.selection()
        if selection:
            item = self.tree.item(selection[0])
            config_id = item['tags'][0]
            print(f"[CONFIGURATION_OVERWRITE_DIALOG] Configuracion seleccionada: {config_id}")
    
    def _overwrite_selected(self):
        """Sobrescribe la configuracion seleccionada"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Debe seleccionar una configuracion para sobrescribir")
            return
        
        item = self.tree.item(selection[0])
        config_id = item['tags'][0]
        config_name = item['values'][0]
        
        # Confirmar sobrescritura
        result = messagebox.askyesno("Confirmar Sobrescritura", 
                                   f"¿Esta seguro de que desea sobrescribir la configuracion '{config_name}'?\n\n"
                                   f"Esta accion creara un backup automatico.")
        
        if result:
            try:
                # Realizar la sobrescritura usando config_data
                if self.config_data:
                    # Obtener metadatos de la configuracion existente
                    existing_config = self.config_manager.load_configuration(config_id)
                    if existing_config:
                        # Mantener metadatos existentes pero actualizar la configuracion
                        result_id = self.config_manager.save_configuration(
                            name=existing_config['metadata'].get('name', config_name),
                            description=existing_config['metadata'].get('description', ''),
                            tags=existing_config['metadata'].get('tags', []),
                            config_data=self.config_data,
                            is_default=existing_config['metadata'].get('is_default', False),
                            overwrite=True,
                            target_config_id=config_id
                        )
                        
                        if result_id:
                            print(f"[CONFIGURATION_OVERWRITE_DIALOG] Configuracion sobrescrita: {result_id}")
                            self.result = result_id
                            self.destroy()
                        else:
                            messagebox.showerror("Error", "No se pudo sobrescribir la configuracion")
                    else:
                        messagebox.showerror("Error", "No se pudo cargar la configuracion existente")
                else:
                    messagebox.showerror("Error", "No hay datos de configuracion para sobrescribir")
                    
            except Exception as e:
                print(f"[CONFIGURATION_OVERWRITE_DIALOG ERROR] Error sobrescribiendo: {e}")
                messagebox.showerror("Error", f"No se pudo sobrescribir la configuracion: {e}")
    
    def _cancel(self):
        """Cancela el dialogo"""
        self.result = None
        self.destroy()
    
    def _center_window(self):
        """Centra la ventana en la pantalla"""
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f"{width}x{height}+{x}+{y}")
    
    def _apply_dialog_theme(self):
        """Aplica tema oscuro al diálogo si está activo en la ventana principal"""
        try:
            # Verificar si el tema oscuro está activo en la ventana principal
            if hasattr(self.parent, 'dark_mode') and self.parent.dark_mode:
                dark_colors = {
                    'bg_primary': '#1e1e1e',
                    'bg_secondary': '#2d2d30',
                    'text_primary': '#cccccc',
                    'accent': '#007acc'
                }
                
                # Aplicar tema oscuro al diálogo
                self.configure(bg=dark_colors['bg_primary'])
                
                # Aplicar recursivamente a todos los widgets
                self._apply_dark_theme_recursive(self, dark_colors)
                
        except Exception as e:
            print(f"[CONFIGURATION_OVERWRITE_DIALOG] Error aplicando tema: {e}")
    
    def _apply_dark_theme_recursive(self, widget, dark_colors):
        """Aplica tema oscuro recursivamente a widgets del diálogo"""
        try:
            # Solo aplicar a widgets tk nativos específicos, NO a widgets ttk
            widget_class = widget.__class__.__name__
            
            # Verificar si es un widget tk nativo (no ttk)
            if widget_class.startswith('Tk') or widget_class in ['Entry', 'Text', 'Canvas', 'Listbox', 'Scrollbar', 'Toplevel', 'Label', 'Button']:
                if isinstance(widget, tk.Entry):
                    widget.configure(
                        bg=dark_colors['bg_secondary'],
                        fg=dark_colors['text_primary'],
                        insertbackground=dark_colors['text_primary'],
                        selectbackground=dark_colors['accent'],
                        selectforeground=dark_colors['text_primary']
                    )
                elif isinstance(widget, tk.Text):
                    widget.configure(
                        bg=dark_colors['bg_secondary'],
                        fg=dark_colors['text_primary'],
                        insertbackground=dark_colors['text_primary'],
                        selectbackground=dark_colors['accent'],
                        selectforeground=dark_colors['text_primary']
                    )
                elif isinstance(widget, tk.Label):
                    widget.configure(
                        bg=dark_colors['bg_primary'],
                        fg=dark_colors['text_primary']
                    )
                elif isinstance(widget, tk.Button):
                    widget.configure(
                        bg=dark_colors['bg_secondary'],
                        fg=dark_colors['text_primary'],
                        activebackground=dark_colors['accent'],
                        activeforeground=dark_colors['text_primary']
                    )
            
            # Recursivamente aplicar a widgets hijos
            for child in widget.winfo_children():
                self._apply_dark_theme_recursive(child, dark_colors)
                
        except Exception as e:
            # Silenciar errores de widgets que no soportan configuración
            pass

class ConfigurationManagerDialog(tk.Toplevel):
    """
    Dialogo para gestionar configuraciones (eliminar, etc.).
    Funcionalidad avanzada de administracion.
    """
    
    def __init__(self, parent, config_manager, icons=None):
        """
        Inicializa el dialogo de gestion
        
        Args:
            parent: Ventana padre
            config_manager: Instancia de ConfigurationManager
            icons: Diccionario de iconos vectoriales
        """
        super().__init__(parent)
        self.parent = parent
        self.config_manager = config_manager
        self.icons = icons or {}
        
        self._setup_window()
        self._create_widgets()
        self._center_window()
        self._apply_dialog_theme()
        
        # Hacer modal
        self.transient(parent)
        self.grab_set()
        
        print("[CONFIGURATION_MANAGER_DIALOG] Dialogo de gestion inicializado")
    
    def _setup_window(self):
        """Configura la ventana del dialogo"""
        self.title("Gestionar Configuraciones")
        self.geometry("700x500")
        self.resizable(True, True)
    
    def _create_widgets(self):
        """Crea los widgets del dialogo con diseño mejorado"""
        # Frame principal con padding mejorado
        main_frame = ttk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=25, pady=25)
        
        # Titulo con icono
        title_label = ttk.Label(main_frame, text="⚙️ Gestionar Configuraciones")
        title_label.pack(pady=(0, 15))
        
        # Subtitulo descriptivo
        subtitle_label = ttk.Label(main_frame, text="Administra tus configuraciones guardadas: eliminar, marcar como defecto, etc.",
                                  foreground="gray")
        subtitle_label.pack(pady=(0, 20))
        
        # Lista de configuraciones con diseño mejorado
        list_frame = ttk.LabelFrame(main_frame, text="📋 Configuraciones Disponibles", padding=10)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
        
        # Treeview con estilo mejorado
        columns = ("Nombre", "Descripcion", "Fecha", "Por Defecto", "Tags")
        self.config_tree = ttk.Treeview(list_frame, columns=columns, show="headings", height=10)
        
        # Configurar columnas con iconos
        self.config_tree.heading("Nombre", text="🏷️ Nombre")
        self.config_tree.heading("Descripcion", text="📄 Descripcion")
        self.config_tree.heading("Fecha", text="📅 Fecha")
        self.config_tree.heading("Por Defecto", text="⭐ Por Defecto")
        self.config_tree.heading("Tags", text="🏷️ Tags")
        
        # Configurar anchos optimizados
        self.config_tree.column("Nombre", width=180, minwidth=150)
        self.config_tree.column("Descripcion", width=250, minwidth=200)
        self.config_tree.column("Fecha", width=140, minwidth=120)
        self.config_tree.column("Por Defecto", width=100, minwidth=80)
        self.config_tree.column("Tags", width=180, minwidth=150)
        
        # Scrollbar con estilo mejorado
        list_scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.config_tree.yview)
        self.config_tree.configure(yscrollcommand=list_scrollbar.set)
        
        self.config_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        list_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Frame para información de selección
        info_frame = ttk.Frame(main_frame)
        info_frame.pack(fill=tk.X, pady=(0, 15))
        
        self.selection_info = ttk.Label(info_frame, text="Selecciona una configuracion para gestionarla", 
                                       foreground="blue")
        self.selection_info.pack(side=tk.LEFT)
        
        # Botones de accion con diseño mejorado
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X)
        
        # Botones principales con iconos vectoriales modernos (si están disponibles)
        delete_btn = ttk.Button(button_frame, text="Eliminar", 
                               image=self.icons.get('delete'), 
                               compound=tk.LEFT, width=15,
                               command=self._delete_configuration)
        delete_btn.pack(side=tk.LEFT, padx=(0, 8))
        
        default_btn = ttk.Button(button_frame, text="Marcar como Defecto", 
                                image=self.icons.get('default'), 
                                compound=tk.LEFT, width=20,
                                command=self._set_as_default)
        default_btn.pack(side=tk.LEFT, padx=(0, 8))
        
        refresh_btn = ttk.Button(button_frame, text="Actualizar Lista", 
                                image=self.icons.get('refresh'), 
                                compound=tk.LEFT, width=15,
                                command=self._refresh_list)
        refresh_btn.pack(side=tk.LEFT, padx=(0, 8))
        
        # Botón cerrar a la derecha
        close_btn = ttk.Button(button_frame, text="Cerrar", 
                              image=self.icons.get('exit'), 
                              compound=tk.LEFT, width=12,
                              command=self.destroy)
        close_btn.pack(side=tk.RIGHT)
        
        # Cargar lista inicial
        self._refresh_list()
    
    def _refresh_list(self):
        """Actualiza la lista de configuraciones"""
        try:
            configurations = self.config_manager.list_configurations()
            
            # Limpiar lista
            for item in self.config_tree.get_children():
                self.config_tree.delete(item)
            
            # Agregar configuraciones
            for config in configurations:
                created_at = config.get("created_at", "")
                if created_at:
                    try:
                        from datetime import datetime
                        dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                        formatted_date = dt.strftime("%Y-%m-%d %H:%M")
                    except:
                        formatted_date = created_at[:16]
                else:
                    formatted_date = "N/A"
                
                is_default = "Si" if config.get("is_default", False) else "No"
                tags = config.get("tags", [])
                tags_text = ", ".join(tags) if tags else "Sin tags"
                
                self.config_tree.insert("", tk.END, values=(
                    config.get("name", "Sin nombre"),
                    config.get("description", "Sin descripcion"),
                    formatted_date,
                    is_default,
                    tags_text
                ), tags=(config["id"],))
            
            print(f"[CONFIGURATION_MANAGER_DIALOG] Lista actualizada: {len(configurations)} configuraciones")
            
        except Exception as e:
            print(f"[CONFIGURATION_MANAGER_DIALOG ERROR] Error actualizando lista: {e}")
            messagebox.showerror("Error", f"No se pudo actualizar la lista: {e}")
    
    def _delete_configuration(self):
        """Elimina la configuracion seleccionada"""
        selection = self.config_tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Seleccione una configuracion para eliminar")
            return
        
        try:
            item = self.config_tree.item(selection[0])
            config_id = item["tags"][0]
            config_name = item["values"][0]
            
            # Confirmar eliminacion
            if messagebox.askyesno("Confirmar Eliminacion", 
                                 f"Esta seguro de eliminar la configuracion '{config_name}'?\n\n"
                                 "Se creara un backup automaticamente."):
                
                self.config_manager.delete_configuration(config_id, create_backup=True)
                messagebox.showinfo("Exito", f"Configuracion '{config_name}' eliminada exitosamente")
                self._refresh_list()
            
        except Exception as e:
            print(f"[CONFIGURATION_MANAGER_DIALOG ERROR] Error eliminando: {e}")
            messagebox.showerror("Error", f"No se pudo eliminar la configuracion: {e}")
    
    def _set_as_default(self):
        """Marca la configuracion seleccionada como por defecto"""
        selection = self.config_tree.selection()
        if not selection:
            messagebox.showwarning("Advertencia", "Seleccione una configuracion para marcar como defecto")
            return
        
        try:
            item = self.config_tree.item(selection[0])
            config_id = item["tags"][0]
            config_name = item["values"][0]
            
            # Confirmar cambio
            if messagebox.askyesno("Confirmar Cambio", 
                                 f"Marcar '{config_name}' como configuracion por defecto?\n\n"
                                 "Esto desmarcara cualquier otra configuracion por defecto."):
                
                # Usar metodo interno del manager
                self.config_manager._set_as_default(config_id)
                messagebox.showinfo("Exito", f"'{config_name}' marcada como configuracion por defecto")
                self._refresh_list()
            
        except Exception as e:
            print(f"[CONFIGURATION_MANAGER_DIALOG ERROR] Error marcando como defecto: {e}")
            messagebox.showerror("Error", f"No se pudo marcar como defecto: {e}")
    
    def _center_window(self):
        """Centra la ventana en la pantalla"""
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f"{width}x{height}+{x}+{y}")
    
    def _apply_dialog_theme(self):
        """Aplica tema oscuro al diálogo si está activo en la ventana principal"""
        try:
            # Verificar si el tema oscuro está activo en la ventana principal
            if hasattr(self.parent, 'dark_mode') and self.parent.dark_mode:
                dark_colors = {
                    'bg_primary': '#1e1e1e',
                    'bg_secondary': '#2d2d30',
                    'text_primary': '#cccccc',
                    'accent': '#007acc'
                }
                
                # Aplicar tema oscuro al diálogo
                self.configure(bg=dark_colors['bg_primary'])
                
                # Aplicar recursivamente a todos los widgets
                self._apply_dark_theme_recursive(self, dark_colors)
                
        except Exception as e:
            print(f"[CONFIGURATION_MANAGER_DIALOG] Error aplicando tema: {e}")
    
    def _apply_dark_theme_recursive(self, widget, dark_colors):
        """Aplica tema oscuro recursivamente a widgets del diálogo"""
        try:
            # Solo aplicar a widgets tk nativos específicos, NO a widgets ttk
            widget_class = widget.__class__.__name__
            
            # Verificar si es un widget tk nativo (no ttk)
            if widget_class.startswith('Tk') or widget_class in ['Entry', 'Text', 'Canvas', 'Listbox', 'Scrollbar', 'Toplevel', 'Label', 'Button']:
                if isinstance(widget, tk.Entry):
                    widget.configure(
                        bg=dark_colors['bg_secondary'],
                        fg=dark_colors['text_primary'],
                        insertbackground=dark_colors['text_primary'],
                        selectbackground=dark_colors['accent'],
                        selectforeground=dark_colors['text_primary']
                    )
                elif isinstance(widget, tk.Text):
                    widget.configure(
                        bg=dark_colors['bg_secondary'],
                        fg=dark_colors['text_primary'],
                        insertbackground=dark_colors['text_primary'],
                        selectbackground=dark_colors['accent'],
                        selectforeground=dark_colors['text_primary']
                    )
                elif isinstance(widget, tk.Label):
                    widget.configure(
                        bg=dark_colors['bg_primary'],
                        fg=dark_colors['text_primary']
                    )
                elif isinstance(widget, tk.Button):
                    widget.configure(
                        bg=dark_colors['bg_secondary'],
                        fg=dark_colors['text_primary'],
                        activebackground=dark_colors['accent'],
                        activeforeground=dark_colors['text_primary']
                    )
            
            # Recursivamente aplicar a widgets hijos
            for child in widget.winfo_children():
                self._apply_dark_theme_recursive(child, dark_colors)
                
        except Exception as e:
            # Silenciar errores de widgets que no soportan configuración
            pass


# Importar datetime para timestamps
from datetime import datetime


def main():
    """Funcion principal del configurador"""
    print("="*60)
    print("CONFIGURADOR DE SIMULACION - GEMELO DIGITAL")
    print("Herramienta independiente de configuracion")
    print("="*60)
    print()

    configurador = ConfiguradorSimulador()
    configurador.ejecutar()


if __name__ == "__main__":
    main()

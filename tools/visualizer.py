# -*- coding: utf-8 -*-
"""
Generador de Imagen de Heatmap Geométrico
Fusiona la geometría del almacén (TMX) con datos de actividad (Excel) para crear visualizaciones PNG.
"""

import argparse
import sys
import os
from typing import Dict, Tuple, Optional
import pandas as pd
from openpyxl import load_workbook
import pytmx
from PIL import Image, ImageDraw


def load_heatmap_data(excel_path: str) -> Dict[Tuple[int, int], float]:
    """
    Carga datos del heatmap desde la hoja VisualHeatmap de un archivo Excel.
    
    Args:
        excel_path: Ruta al archivo Excel con datos de heatmap
        
    Returns:
        Diccionario con coordenadas (x,y) como claves e intensidad como valores
        
    Raises:
        FileNotFoundError: Si el archivo Excel no existe
        ValueError: Si la hoja VisualHeatmap no existe o está vacía
    """
    print(f"[VISUALIZER] Cargando datos de heatmap desde: {excel_path}")
    
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Archivo Excel no encontrado: {excel_path}")
    
    try:
        wb = load_workbook(excel_path, data_only=True)
        
        if 'VisualHeatmap' not in wb.sheetnames:
            raise ValueError(f"Hoja 'VisualHeatmap' no encontrada en {excel_path}")
        
        heatmap_sheet = wb['VisualHeatmap']
        heatmap_data = {}
        
        # Convertir matriz Excel a estructura de datos (x,y,intensidad)
        for row in range(1, heatmap_sheet.max_row + 1):
            for col in range(1, heatmap_sheet.max_column + 1):
                cell_value = heatmap_sheet.cell(row, col).value
                
                if cell_value is not None and isinstance(cell_value, (int, float)) and cell_value > 0:
                    # Excel: col=1->x=0, row=1->y=0 (convertir a 0-indexed)
                    # Invertir Y para coincidir con coordenadas de TMX (y=0 arriba)
                    x = col - 1
                    y = heatmap_sheet.max_row - row  # Invertir Y
                    heatmap_data[(x, y)] = float(cell_value)
        
        print(f"[VISUALIZER] Heatmap cargado: {len(heatmap_data)} coordenadas con actividad")
        return heatmap_data
        
    except Exception as e:
        raise ValueError(f"Error procesando archivo Excel: {e}")


def load_tmx_geometry(tmx_path: str, layer_name: str = "walls") -> Tuple[pytmx.TiledMap, Optional[pytmx.TiledTileLayer]]:
    """
    Carga geometría del almacén desde un archivo TMX.
    
    Args:
        tmx_path: Ruta al archivo TMX
        layer_name: Nombre de la capa que contiene geometría de muros
        
    Returns:
        Tupla (TiledMap, geometry_layer) donde geometry_layer puede ser None si no se encuentra
        
    Raises:
        FileNotFoundError: Si el archivo TMX no existe
        ValueError: Si el archivo TMX no puede ser parseado
    """
    print(f"[VISUALIZER] Cargando geometría TMX desde: {tmx_path}")
    
    if not os.path.exists(tmx_path):
        raise FileNotFoundError(f"Archivo TMX no encontrado: {tmx_path}")
    
    try:
        tm = pytmx.TiledMap(tmx_path)
        print(f"[VISUALIZER] TMX cargado: {tm.width}x{tm.height} tiles ({tm.tilewidth}x{tm.tileheight}px)")
        
        # Buscar capa de geometría
        geometry_layer = None
        available_layers = []
        
        for layer in tm.layers:
            layer_name_attr = getattr(layer, 'name', 'sin_nombre')
            available_layers.append(layer_name_attr)
            
            # Buscar capa por nombre específico o patrones comunes
            if (layer_name_attr and 
                (layer_name.lower() in layer_name_attr.lower() or
                 'muro' in layer_name_attr.lower() or 
                 'wall' in layer_name_attr.lower() or
                 'patron' in layer_name_attr.lower())):
                geometry_layer = layer
                print(f"[VISUALIZER] Capa de geometría encontrada: {layer_name_attr}")
                break
        
        if not geometry_layer:
            print(f"[VISUALIZER] ADVERTENCIA: Capa '{layer_name}' no encontrada")
            print(f"[VISUALIZER] Capas disponibles: {available_layers}")
            print(f"[VISUALIZER] Continuando sin geometría de muros")
        
        return tm, geometry_layer
        
    except Exception as e:
        raise ValueError(f"Error procesando archivo TMX: {e}")


def get_heatmap_color(intensity: float, max_intensity: float) -> Tuple[int, int, int]:
    """
    Convierte intensidad de actividad a color RGB usando gradiente Verde-Amarillo-Rojo.
    
    Args:
        intensity: Valor de intensidad de actividad
        max_intensity: Valor máximo de intensidad para normalización
        
    Returns:
        Tupla RGB (r, g, b) con valores 0-255
    """
    if intensity <= 0 or max_intensity <= 0:
        return (255, 255, 255)  # Blanco para sin actividad
    
    # Normalizar intensidad 0-1
    ratio = min(intensity / max_intensity, 1.0)
    
    if ratio < 0.5:
        # Verde (0,255,0) a Amarillo (255,255,0)
        r = int(255 * ratio * 2)
        g = 255
        b = 0
    else:
        # Amarillo (255,255,0) a Rojo (255,0,0)
        r = 255
        g = int(255 * (2 - ratio * 2))
        b = 0
    
    return (r, g, b)


def generate_heatmap_image(tm: pytmx.TiledMap, 
                          geometry_layer: Optional[pytmx.TiledTileLayer],
                          heatmap_data: Dict[Tuple[int, int], float],
                          pixel_scale: int = 16,
                          output_path: str = "heatmap_output.png") -> str:
    """
    Genera imagen PNG fusionando geometría TMX con datos de heatmap.
    
    Args:
        tm: Mapa TMX cargado
        geometry_layer: Capa de geometría (puede ser None)
        heatmap_data: Datos de actividad por coordenada
        pixel_scale: Píxeles por tile del almacén
        output_path: Ruta del archivo de salida
        
    Returns:
        Ruta del archivo generado
    """
    print(f"[VISUALIZER] Generando imagen de heatmap...")
    
    # Crear imagen base
    img_width = tm.width * pixel_scale
    img_height = tm.height * pixel_scale
    img = Image.new('RGB', (img_width, img_height), color=(240, 240, 240))  # Fondo gris claro
    draw = ImageDraw.Draw(img)
    
    # Colores de referencia
    WALL_COLOR = (60, 60, 60)         # Gris oscuro para muros/racks
    EMPTY_COLOR = (255, 255, 255)     # Blanco para espacios vacíos
    FLOOR_COLOR = (240, 240, 240)     # Gris claro para suelo
    PICKING_COLOR = (200, 255, 200)   # Verde claro para puntos picking
    PARKING_COLOR = (200, 200, 255)   # Azul claro para estacionamiento
    DEPOT_COLOR = (255, 200, 200)     # Rosa claro para depot
    INBOUND_COLOR = (255, 255, 200)   # Amarillo claro para inbound
    
    # Calcular intensidad máxima para normalización
    max_intensity = max(heatmap_data.values()) if heatmap_data else 1.0
    print(f"[VISUALIZER] Intensidad máxima detectada: {max_intensity}")
    
    # Contadores para estadísticas
    wall_count = 0
    activity_count = 0
    empty_count = 0
    
    # Iterar sobre cada celda del almacén
    for x in range(tm.width):
        for y in range(tm.height):
            # Calcular coordenadas de píxeles en la imagen
            pixel_x = x * pixel_scale
            pixel_y = y * pixel_scale
            pixel_rect = [pixel_x, pixel_y, pixel_x + pixel_scale, pixel_y + pixel_scale]
            
            # Obtener GID del tile y definir color base
            base_color = EMPTY_COLOR
            gid = 0
            
            if geometry_layer and hasattr(geometry_layer, 'data'):
                try:
                    # Acceso como matriz 2D: geometry_layer.data[y][x]
                    gid = geometry_layer.data[y][x]
                    
                    # Asignar color base según tipo de tile (MAPEO REAL TMX)
                    if gid == 1:    # parking (estacionamiento)
                        base_color = PARKING_COLOR
                    elif gid == 2:  # depot (zona depot)
                        base_color = DEPOT_COLOR
                    elif gid == 3:  # floor (suelo navegable)
                        base_color = FLOOR_COLOR
                    elif gid == 4:  # rack (racks/estantes - MUROS)
                        base_color = WALL_COLOR
                        wall_count += 1
                    elif gid == 5:  # picking_location (puntos de picking)
                        base_color = PICKING_COLOR
                    elif gid == 6:  # inbound (zona entrada)
                        base_color = INBOUND_COLOR
                        
                except (IndexError, TypeError):
                    # Si falla el acceso, usar color por defecto
                    pass
            
            # Si es muro, usar color base sin modificar
            if gid == 4:  # Racks son muros sólidos (GID 4 según TMX real)
                draw.rectangle(pixel_rect, fill=base_color)
            else:
                # Para tiles navegables, superponer actividad de heatmap si existe
                intensity = heatmap_data.get((x, y), 0)
                if intensity > 0:
                    # Combinar color base con heatmap usando transparencia
                    heatmap_color = get_heatmap_color(intensity, max_intensity)
                    # Mezclar 70% heatmap + 30% base
                    mixed_color = (
                        int(heatmap_color[0] * 0.7 + base_color[0] * 0.3),
                        int(heatmap_color[1] * 0.7 + base_color[1] * 0.3),
                        int(heatmap_color[2] * 0.7 + base_color[2] * 0.3)
                    )
                    draw.rectangle(pixel_rect, fill=mixed_color)
                    activity_count += 1
                else:
                    # Sin actividad, usar color base
                    draw.rectangle(pixel_rect, fill=base_color)
                    empty_count += 1
    
    # Guardar imagen
    img.save(output_path)
    
    # Estadísticas
    total_tiles = tm.width * tm.height
    print(f"[VISUALIZER] Imagen generada: {output_path}")
    print(f"[VISUALIZER] Dimensiones: {img_width}x{img_height} píxeles ({tm.width}x{tm.height} tiles)")
    print(f"[VISUALIZER] Estadísticas:")
    print(f"  • Muros: {wall_count}/{total_tiles} ({wall_count/total_tiles*100:.1f}%)")
    print(f"  • Con actividad: {activity_count}/{total_tiles} ({activity_count/total_tiles*100:.1f}%)")
    print(f"  • Espacios vacíos: {empty_count}/{total_tiles} ({empty_count/total_tiles*100:.1f}%)")
    
    return output_path


def main():
    """Función principal para ejecución independiente del visualizador de heatmaps"""
    parser = argparse.ArgumentParser(
        description='Generador de Imagen de Heatmap Geometrico - Fusiona TMX + Excel -> PNG',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos de uso:
  python visualizer.py --excel report.xlsx --tmx layouts/WH1.tmx --output heatmap.png
  python visualizer.py --excel visual_heatmap_report.xlsx --tmx layouts/WH1.tmx --output warehouse_heatmap.png --pixel_scale 20
  python visualizer.py --excel data.xlsx --tmx map.tmx --output result.png --layer_name "Muros" --pixel_scale 12
        """
    )
    
    parser.add_argument(
        '--excel_path', 
        required=True,
        help='Ruta al archivo de reporte Excel con hoja VisualHeatmap'
    )
    parser.add_argument(
        '--tmx_path', 
        required=True,
        help='Ruta al archivo TMX con geometría del almacén'
    )
    parser.add_argument(
        '--output_path', 
        required=True,
        help='Nombre del archivo PNG de salida (ej: heatmap_geometrico.png)'
    )
    parser.add_argument(
        '--layer_name', 
        default='walls',
        help='Nombre de la capa TMX con geometría de muros (default: "walls")'
    )
    parser.add_argument(
        '--pixel_scale', 
        type=int, 
        default=16,
        help='Píxeles por tile del almacén (default: 16)'
    )
    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help='Mostrar información detallada del procesamiento'
    )
    
    args = parser.parse_args()
    
    try:
        print("=" * 70)
        print("GENERADOR DE HEATMAP GEOMETRICO")
        print("=" * 70)
        
        # Paso 1: Cargar datos del heatmap desde Excel
        heatmap_data = load_heatmap_data(args.excel_path)
        
        # Paso 2: Cargar geometría del mapa TMX
        tm, geometry_layer = load_tmx_geometry(args.tmx_path, args.layer_name)
        
        # Paso 3 & 4: Generar imagen fusionada
        output_file = generate_heatmap_image(
            tm=tm,
            geometry_layer=geometry_layer,
            heatmap_data=heatmap_data,
            pixel_scale=args.pixel_scale,
            output_path=args.output_path
        )
        
        print("=" * 70)
        print("PROCESAMIENTO COMPLETADO EXITOSAMENTE")
        print(f"Imagen generada: {output_file}")
        print("=" * 70)
        
        return 0
        
    except FileNotFoundError as e:
        print(f"ERROR: Archivo no encontrado - {e}")
        return 1
    except ValueError as e:
        print(f"ERROR: Datos inválidos - {e}")
        return 1
    except Exception as e:
        print(f"ERROR INESPERADO: {e}")
        if args.verbose:
            import traceback
            traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())


print("[OK] Módulo 'visualizer' cargado - Generador de heatmap geométrico en imagen PNG.")
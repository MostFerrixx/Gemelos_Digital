# -*- coding: utf-8 -*-
"""
AnalyticsExporter V2 - Version mejorada con arquitectura robusta.
Fase 2 del refactor - Implementa SimulationContext y ExportResult patterns.
Mantiene compatibilidad con funcionalidad existente mientras mejora robustez.
"""

import os
import sys
import json
import subprocess
import time
from datetime import datetime
from typing import Optional

from engines.analytics_engine import AnalyticsEngine
from subsystems.utils.helpers import mostrar_metricas_consola
from .context import SimulationContext, ExportResult


class AnalyticsExporter:
    """
    Gestor robusto para exportacion de analiticas y reportes de simulacion.

    Version 2.0 con mejoras arquitectonicas:
    - SimulationContext para encapsulacion de datos
    - ExportResult para manejo estructurado de resultados
    - Rollback automatico en caso de fallo
    - Logging y debugging mejorado
    """

    def __init__(self, simulation_context: SimulationContext):
        """
        Inicializa AnalyticsExporter con contexto de simulacion.

        Args:
            simulation_context: Contexto unificado con todos los datos necesarios
        """
        # Validar contexto
        simulation_context.validate()

        self.context = simulation_context

        print(f"[ANALYTICS-EXPORTER V2] Inicializado con contexto de simulacion")
        print(f"[ANALYTICS-EXPORTER V2] Session: {self.context.session_timestamp}")
        print(f"[ANALYTICS-EXPORTER V2] Output: {self.context.session_output_dir}")

    @classmethod
    def from_legacy_params(cls, almacen, configuracion, session_timestamp=None, session_output_dir=None):
        """
        Factory method para compatibilidad con API legacy.

        Permite usar AnalyticsExporter V2 con parametros del V1.

        Args:
            almacen: Objeto AlmacenMejorado
            configuracion: Dict de configuracion
            session_timestamp: Timestamp de sesion (opcional)
            session_output_dir: Directorio de salida (opcional)

        Returns:
            AnalyticsExporter: Instancia configurada con contexto
        """
        context = SimulationContext(
            almacen=almacen,
            configuracion=configuracion,
            session_timestamp=session_timestamp,
            session_output_dir=session_output_dir
        )
        return cls(context)

    def export_complete_analytics(self) -> ExportResult:
        """
        Pipeline completo de exportacion de analiticas con manejo robusto.

        Mejoras sobre V1:
        - Retorna ExportResult estructurado
        - Rollback automatico en caso de fallo
        - Timing automatico de ejecucion
        - Error handling mejorado

        Returns:
            ExportResult: Resultado detallado de la exportacion
        """
        start_time = time.time()
        result = ExportResult()

        # Configurar metadatos de exportacion
        result.export_metadata = self.context.get_export_metadata()
        result.rollback_info['output_dir'] = self.context.session_output_dir

        print("\n" + "="*70)
        print("SIMULACION COMPLETADA - INICIANDO PIPELINE DE ANALITICAS V2")
        print("="*70)

        if not self.context.almacen:
            result.add_error("No hay datos del almacen para procesar")
            return result

        try:
            # Mostrar metricas basicas
            mostrar_metricas_consola(self.context.almacen)

            # Crear estructura de directorios organizados
            output_dir = self.context.session_output_dir
            timestamp = self.context.session_timestamp

            try:
                os.makedirs(output_dir, exist_ok=True)
                print(f"[SETUP] Directorio de salida creado: {output_dir}")
            except Exception as e:
                result.add_warning(f"No se pudo crear directorio de salida: {e}")
                # Fallback: usar directorio actual
                output_dir = "."
                result.rollback_info['output_dir'] = output_dir

            # MEJ-BOTTLENECK: purgados los 3 archivos que nadie leia
            # (simulacion_completada_*.json placeholder "SKELETON",
            # raw_events_*.json ~4.5MB duplicado del .jsonl con esquema viejo,
            # simulation_report_*.json version JSON del Excel sin consumidores).
            # El PNG del heatmap ahora cuelga de archivo_excel (antes colgaba
            # del JSON eliminado, un acople accidental).

            # 1. PIPELINE AUTOMATIZADO: AnalyticsEngine -> Excel
            print("[1/2] Simulacion completada. Generando reporte de Excel...")
            try:
                # Usar el metodo __init__ original con eventos y configuracion en memoria
                analytics_engine = AnalyticsEngine(self.context.event_log, self.context.configuracion)
                analytics_engine.process_events()

                # Generar archivo Excel con ruta organizada
                excel_filename = os.path.join(output_dir, f"simulation_report_{timestamp}.xlsx")
                archivo_excel = analytics_engine.export_to_excel(excel_filename)

                if archivo_excel:
                    result.add_file(archivo_excel)
                    print(f"[1/2] Reporte de Excel generado: {archivo_excel}")
                    # INIT-5: anexar hoja 'Nivel de servicio' (backorders) al reporte.
                    self._append_service_level_sheet(archivo_excel)
                    # INIT-4b: anexar hoja 'Cumplimiento SLA' (due_time) al reporte.
                    self._append_sla_summary_sheet(archivo_excel)
                    # MEJ-BOTTLENECK: anexar hoja 'Cuellos de Botella' al reporte.
                    self._append_bottleneck_sheet(archivo_excel)

                    # 2. PIPELINE AUTOMATIZADO: Visualizer -> PNG
                    print("[2/2] Reporte de Excel generado. Creando imagen de heatmap...")
                    try:
                        png_file = self._ejecutar_visualizador(archivo_excel, timestamp, output_dir)
                        if png_file:
                            result.add_file(png_file)
                    except Exception as e:
                        result.add_warning(f"Error generando heatmap PNG: {e}")

                else:
                    result.add_error("No se pudo generar el reporte de Excel")

            except Exception as e:
                result.add_error(f"Error en pipeline de analiticas: {e}")

            # Timing final
            end_time = time.time()
            result.set_execution_time(start_time, end_time)

            # Resumen final
            print("\n" + "="*70)
            print("PROCESO COMPLETADO V2")
            print("="*70)
            print(result.get_summary())
            print("="*70)

            return result

        except Exception as e:
            # Error critico - rollback automatico
            result.add_error(f"Error critico en exportacion: {e}")
            result.set_execution_time(start_time, time.time())

            print(f"[ROLLBACK] Error critico detectado: {e}")
            if result.rollback_on_failure():
                print("[ROLLBACK] Rollback completado exitosamente")
            else:
                print("[ROLLBACK] Rollback tuvo problemas - verificar manualmente")

            return result

    def _append_service_level_sheet(self, excel_path):
        """INIT-5: anexa una hoja 'Nivel de servicio' (backorders) al reporte Excel ya
        generado, reabriendolo (desacoplado de AnalyticsEngine). Solo con pedidos reales
        (modo determinista); en estocastico anota N/A. No rompe el resto del reporte."""
        try:
            import openpyxl
            from core.replay_utils import build_service_level_summary
            svc = build_service_level_summary(self.context.almacen)
            wb = openpyxl.load_workbook(excel_path)
            if 'Nivel de servicio' in wb.sheetnames:
                del wb['Nivel de servicio']
            ws = wb.create_sheet('Nivel de servicio')
            if not svc.get('available'):
                ws.append(['Nivel de servicio'])
                ws.append(['Modo estocastico: sin validacion de stock -> N/A'])
                wb.save(excel_path)
                return
            ws.append(['Nivel de servicio (backorders)'])
            ws.append([])
            ws.append(['Total pedido (u)', svc['total_requested']])
            ws.append(['Total servido (u)', svc['total_served']])
            ws.append(['Faltante / deuda (u)', svc['total_unfilled']])
            ws.append(['Fill-rate (%)', svc['fill_rate_pct']])
            ws.append(['Pedidos cortos', svc['orders_short']])
            ws.append(['Items en backorder', svc['backorder_items']])
            ws.append([])
            ws.append(['Order ID', 'SKU', 'Pedido', 'Servido', 'Faltante', 'Cumplimiento %'])
            for u in svc.get('unfilled', []):
                req = u.get('qty_requested', 0) or 0
                got = u.get('qty_allocated', 0) or 0
                miss = u.get('qty_unfilled', max(req - got, 0))
                pct = round(got / req * 100, 1) if req else 100.0
                ws.append([u.get('order_id'), u.get('sku_id'), req, got, miss, pct])
            ws.append([])
            ws.append(['TOTAL', '', svc['total_requested'], svc['total_served'],
                       svc['total_unfilled'], svc['fill_rate_pct']])
            wb.save(excel_path)
            print("[1/2] Hoja 'Nivel de servicio' anexada al Excel")
        except Exception as e:
            print(f"[1/2] WARNING: no se pudo anexar hoja de nivel de servicio: {e}")

    def _append_sla_summary_sheet(self, excel_path):
        """INIT-4b: anexa una hoja 'Cumplimiento SLA' (due_time) al reporte Excel
        ya generado, reabriendolo (mismo patron que _append_service_level_sheet).
        Solo con pedidos que tengan due_time (INIT-4 C2 activo); si ninguno lo
        tiene, anota N/A. No rompe el resto del reporte."""
        try:
            import openpyxl
            from core.replay_utils import build_sla_summary
            sla = build_sla_summary(self.context.almacen)
            wb = openpyxl.load_workbook(excel_path)
            if 'Cumplimiento SLA' in wb.sheetnames:
                del wb['Cumplimiento SLA']
            ws = wb.create_sheet('Cumplimiento SLA')
            if not sla.get('available'):
                ws.append(['Cumplimiento SLA'])
                ws.append(['Ningun pedido completado trae due_time (INIT-4 C2 desactivado) -> N/A'])
                wb.save(excel_path)
                return
            ws.append(['Cumplimiento SLA (due_time)'])
            ws.append([])
            ws.append(['Pedidos con SLA', sla['total_orders_with_sla']])
            ws.append(['A tiempo', sla['orders_on_time']])
            ws.append(['Vencidos', sla['orders_late']])
            ws.append(['Cumplimiento (%)', sla['on_time_pct']])
            ws.append([])
            ws.append(['Order ID', 'Due time (s)', 'Completado (s)', 'Retraso (s)'])
            for o in sla.get('late_orders', []):
                ws.append([o.get('order_id'), o.get('due_time'), o.get('completion_time'),
                           o.get('delay_seconds')])
            wb.save(excel_path)
            print("[1/2] Hoja 'Cumplimiento SLA' anexada al Excel")
        except Exception as e:
            print(f"[1/2] WARNING: no se pudo anexar hoja de cumplimiento SLA: {e}")

    def _append_bottleneck_sheet(self, excel_path):
        """MEJ-BOTTLENECK: anexa una hoja 'Cuellos de Botella' al reporte Excel
        ya generado (mismo patron que las hojas de INIT-5/INIT-4b). Consolida
        los hotspots de congestion, las metricas del planner espacio-temporal
        y las esperas del muelle (outbound) que antes quedaban repartidos en
        JSONs sueltos por corrida. Con los subsistemas apagados anota N/A."""
        try:
            import openpyxl
            from core.replay_utils import build_bottleneck_summary
            bn = build_bottleneck_summary(self.context.almacen)
            wb = openpyxl.load_workbook(excel_path)
            if 'Cuellos de Botella' in wb.sheetnames:
                del wb['Cuellos de Botella']
            ws = wb.create_sheet('Cuellos de Botella')
            ws.append(['Cuellos de Botella (donde se pierde tiempo)'])
            ws.append([])
            if not bn.get('available'):
                ws.append(['Congestion y outbound desactivados en esta corrida -> N/A'])
                wb.save(excel_path)
                return

            cong = bn.get('congestion')
            ws.append(['CONGESTION DE PASILLOS (agentes que coinciden en la misma celda)'])
            if cong:
                ws.append(['Co-ocupaciones totales', cong['cooccupation_events_total']])
                ws.append(['Celdas distintas con co-ocupacion', cong['distinct_cells_with_cooccupation']])
                ws.append(['Max agentes simultaneos en una celda', cong['max_concurrent_any_cell']])
                ws.append([])
                ws.append(['Celda (x,y)', 'Co-ocupaciones', 'Max simultaneos'])
                for h in cong.get('top_hotspots', []):
                    cell = h.get('cell', ['?', '?'])
                    ws.append(['(%s, %s)' % (cell[0], cell[1]),
                               h.get('cooccupations', 0), h.get('max_concurrent', 0)])
            else:
                ws.append(['Capa de congestion desactivada -> N/A'])
            ws.append([])

            pl = bn.get('planner')
            ws.append(['PLANNER ANTI-COLISION (ruteo espacio-temporal)'])
            if pl:
                ws.append(['Tramos planificados', pl['segments_planned']])
                ws.append(['Planes encontrados / fallidos', '%s / %s' % (pl['plans_found'], pl['plans_failed'])])
                ws.append(['Esperas insertadas por plan (promedio)', pl['avg_waits_per_plan']])
                ws.append(['Solapes en reservas (debe ser 0)', pl['table_overlap_violations']])
            else:
                ws.append(['Planner desactivado -> N/A'])
            ws.append([])

            ob = bn.get('outbound')
            ws.append(['MUELLE / STAGING (esperas de descarga y despacho)'])
            if ob:
                ws.append(['Pallets depositados / despachados',
                           '%s / %s' % (ob['pallets_staged'], ob['pallets_shipped'])])
                ws.append(['Esperas por slot lleno (eventos)', ob['slot_wait_events']])
                ws.append(['Tiempo total esperando slot (s)', round(ob['slot_wait_time'], 1)])
                ws.append(['Espera maxima por slot (s)', round(ob['max_slot_wait'], 1)])
                ws.append(['Esperas por carril lleno (eventos)', ob['lane_full_wait_events']])
                ws.append(['Tiempo total esperando carril (s)', round(ob['lane_full_wait_time'], 1)])
                peak = ob.get('peak_occupancy') or {}
                if peak:
                    ws.append([])
                    ws.append(['Zona de staging', 'Ocupacion pico (pallets)'])
                    for sid in sorted(peak, key=lambda x: int(x)):
                        ws.append(['Staging %s' % sid, peak[sid]])
            else:
                ws.append(['Outbound desactivado -> N/A'])
            wb.save(excel_path)
            print("[1/2] Hoja 'Cuellos de Botella' anexada al Excel")
        except Exception as e:
            print(f"[1/2] WARNING: no se pudo anexar hoja de cuellos de botella: {e}")

    # (PODA 2026-07-07: aqui vivian export_complete_analytics_with_buffer y
    #  _exportar_eventos_crudos_organizado -- 0 callers, y ademas escribian los
    #  3 archivos ya purgados del pipeline vivo. Ver docs/CHANGELOG.md.)

    def _ejecutar_visualizador(self, excel_path: str, timestamp: str, output_dir: str) -> Optional[str]:
        """
        Ejecuta visualizer.py automaticamente usando subprocess.

        Version mejorada con mejor error handling y retorno de path generado.

        Args:
            excel_path: Path del archivo Excel de entrada
            timestamp: Timestamp para nombre de archivo de salida
            output_dir: Directorio de salida

        Returns:
            Optional[str]: Path del archivo PNG generado o None si falla
        """
        output_filename = os.path.join(output_dir, f"warehouse_heatmap_{timestamp}.png")

        try:
            # Construir rutas dinamicamente - los archivos están en el directorio raíz del proyecto
            script_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            visualizer_path = os.path.join(script_dir, "visualizer.py")
            tmx_path = os.path.join(script_dir, self.context.configuracion.get('layout_file', 'layouts/WH1.tmx'))

            # Verificar que los archivos existen
            missing_files = []
            if not os.path.exists(visualizer_path):
                missing_files.append(f"visualizer.py: {visualizer_path}")
            if not os.path.exists(tmx_path):
                missing_files.append(f"TMX file: {tmx_path}")
            if not os.path.exists(excel_path):
                missing_files.append(f"Excel file: {excel_path}")

            if missing_files:
                for missing in missing_files:
                    print(f"[ERROR V2] Archivo no encontrado: {missing}")
                return None

            # Construir comando para visualizer.py
            cmd = [
                sys.executable,  # Python ejecutable actual
                visualizer_path,
                "--excel_path", excel_path,
                "--tmx_path", tmx_path,
                "--output_path", output_filename,
                "--layer_name", "Capa de patrones 1",  # Nombre de capa TMX conocido
                "--pixel_scale", "16"  # Escala por defecto
            ]

            print(f"[VISUALIZER V2] Ejecutando comando: {' '.join(cmd)}")

            # Ejecutar visualizer.py de forma robusta
            result = subprocess.run(
                cmd,
                cwd=script_dir,  # Ejecutar en el directorio del script
                capture_output=True,
                text=True,
                timeout=300  # Timeout de 5 minutos
            )

            if result.returncode == 0:
                print(f"[4/4] Imagen de heatmap generada exitosamente: {output_filename}")
                # Mostrar output del visualizer si es util
                if result.stdout:
                    # Filtrar solo las lineas importantes del output
                    lines = result.stdout.split('\n')
                    for line in lines:
                        if '[VISUALIZER]' in line or 'PROCESAMIENTO COMPLETADO' in line:
                            print(f"  {line}")
                return output_filename
            else:
                print(f"[ERROR V2] visualizer.py fallo con codigo: {result.returncode}")
                if result.stderr:
                    print(f"[ERROR V2] Error del visualizer: {result.stderr}")
                return None

        except subprocess.TimeoutExpired:
            print("[ERROR V2] visualizer.py tomo demasiado tiempo (>5 min) - proceso terminado")
            return None
        except Exception as e:
            print(f"[ERROR V2] Error ejecutando visualizer.py: {e}")
            return None
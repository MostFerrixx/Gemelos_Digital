# -*- coding: utf-8 -*-
"""Endpoints del CONFIGURADOR web (config.json, presets, upload de ordenes).
REFACTOR 2026-07-07: extraido verbatim del monolito server.py."""
import os
from typing import Any, Dict, List

from fastapi import APIRouter, File, HTTPException, UploadFile
from pydantic import BaseModel

from web_prototype.app_state import PROJECT_ROOT, config_manager

router = APIRouter()


class ConfigData(BaseModel):
    """Model for configuration data"""
    config: Dict[str, Any]


class SaveConfigurationRequest(BaseModel):
    """Model for saving a configuration preset"""
    name: str
    description: str = ""
    config: Dict[str, Any]
    is_default: bool = False


@router.get("/api/configurator/config")
def get_config():
    """Load current config.json"""
    try:
        config = config_manager.load_config()
        return {"success": True, "config": config}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/configurator/config")
def save_config(data: ConfigData):
    """Save/update config.json"""
    try:
        success, errors = config_manager.save_config(data.config)
        
        if success:
            return {"success": True, "message": "Configuration saved successfully"}
        else:
            return {"success": False, "errors": errors}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/configurator/work-areas")
def get_work_areas(sequence_file: str):
    """Extract work areas from sequence file"""
    try:
        work_areas = config_manager.extract_work_areas(sequence_file)
        return {"success": True, "work_areas": work_areas}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/configurator/configurations")
def list_configurations():
    """List all saved configuration presets"""
    try:
        configs = config_manager.list_configurations()
        return {"success": True, "configurations": configs}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/configurator/configurations")
def save_configuration(request: SaveConfigurationRequest):
    """Save a new configuration preset"""
    try:
        success, config_id, errors = config_manager.save_configuration(
            request.name,
            request.description,
            request.config,
            request.is_default
        )
        
        if success:
            return {
                "success": True,
                "config_id": config_id,
                "message": f"Configuration '{request.name}' saved successfully"
            }
        else:
            return {"success": False, "errors": errors}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/configurator/configurations/{config_id}")
def get_configuration(config_id: str):
    """Load a specific configuration preset"""
    try:
        config = config_manager.load_configuration(config_id)
        
        if config is None:
            raise HTTPException(status_code=404, detail="Configuration not found")
        
        return {"success": True, "config": config}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/api/configurator/configurations/{config_id}")
def delete_configuration(config_id: str):
    """Delete a configuration preset"""
    try:
        success, errors = config_manager.delete_configuration(config_id)
        
        if success:
            return {"success": True, "message": "Configuration deleted successfully"}
        else:
            return {"success": False, "errors": errors}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/api/configurator/configurations/{config_id}/set-default")
def set_default_configuration(config_id: str):
    """Set a configuration as the default"""
    try:
        success, errors = config_manager.set_default_configuration(config_id)
        
        if success:
            return {"success": True, "message": "Default configuration set successfully"}
        else:
            return {"success": False, "errors": errors}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/configurator/default")
def get_default_configuration():
    """Get the default configuration"""
    try:
        config = config_manager.get_default_configuration()
        
        if config is None:
            # Return hardcoded default if no default is set
            config = config_manager._get_default_config()
        
        return {"success": True, "config": config}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ========================================================================
# DETERMINISTIC ORDER FILE UPLOAD
# ========================================================================

class OrderFileUploadResponse(BaseModel):
    """Response model for order file upload"""
    success: bool
    message: str
    summary: Dict[str, Any] = {}
    exclusions: List[Dict[str, Any]] = []
    file_path: str = ""

@router.post("/api/upload-orders")
async def upload_orders_file(
    file: UploadFile = File(...),
    fulfillment_policy: str = "ship_partial"
):
    """
    Upload an order file (JSON or CSV) for deterministic simulation mode.
    
    Args:
        file: Order file (JSON or CSV format)
        fulfillment_policy: "ship_partial" or "fill_or_kill"
        
    Returns:
        Validation report with summary and exclusions
    """
    import json
    import csv
    from io import StringIO
    
    try:
        # Validate file extension
        filename = file.filename.lower()
        if not (filename.endswith('.json') or filename.endswith('.csv')):
            raise HTTPException(
                status_code=400,
                detail="Solo se permiten archivos .json o .csv"
            )
        
        # Read file content
        content = await file.read()
        content_str = content.decode('utf-8')
        
        # Parse based on file type
        orders = []
        total_items = 0
        
        if filename.endswith('.json'):
            # Parse JSON
            try:
                data = json.loads(content_str)
                orders_data = data.get('orders', data) if isinstance(data, dict) else data
                
                if not isinstance(orders_data, list):
                    raise HTTPException(status_code=400, detail="JSON debe contener un array 'orders'")
                
                for order in orders_data:
                    items = order.get('items', [])
                    orders.append({
                        'order_id': order.get('order_id', ''),
                        'items': items,
                        'staging_id': order.get('staging_id')
                    })
                    total_items += len(items)
                    
            except json.JSONDecodeError as e:
                raise HTTPException(status_code=400, detail=f"Error parsing JSON: {str(e)}")
                
        else:  # CSV
            # Parse CSV with smart grouping by order_id
            try:
                orders_dict = {}
                reader = csv.DictReader(StringIO(content_str))
                
                for row in reader:
                    order_id = row.get('order_id', '').strip()
                    if not order_id:
                        continue
                    
                    if order_id not in orders_dict:
                        staging_id = row.get('staging_id', '').strip()
                        orders_dict[order_id] = {
                            'order_id': order_id,
                            'items': [],
                            'staging_id': int(staging_id) if staging_id else None
                        }
                    
                    sku_id = row.get('sku_id', '').strip()
                    quantity_str = row.get('quantity', '1').strip()
                    
                    if sku_id:
                        try:
                            quantity = int(quantity_str) if quantity_str else 1
                        except ValueError:
                            quantity = 1
                        
                        orders_dict[order_id]['items'].append({
                            'sku_id': sku_id,
                            'quantity': quantity,
                            'work_area': row.get('work_area', '').strip() or None
                        })
                        total_items += 1
                
                orders = list(orders_dict.values())
                
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Error parsing CSV: {str(e)}")
        
        # Validate orders exist
        if not orders:
            raise HTTPException(status_code=400, detail="No se encontraron ordenes válidas en el archivo")
        
        # Save file for later use by simulation
        uploads_dir = os.path.join(PROJECT_ROOT, "uploads")
        os.makedirs(uploads_dir, exist_ok=True)
        
        # REVIEW 2026-07-07: basename() -- un filename artesanal con separadores
        # escapaba de uploads/ (misma clase que el fix de upload_replay).
        saved_file_path = os.path.join(uploads_dir, f"orders_{os.path.basename(file.filename)}")
        with open(saved_file_path, 'wb') as f:
            f.write(content)
        
        # Load current catalog to validate SKUs - NOW USES SQLite!
        catalog_skus = set()
        try:
            import sqlite3
            db_path = os.path.join(PROJECT_ROOT, 'warehouse.db')
            
            if os.path.exists(db_path):
                # USE SQLITE - Fast and uses real SKUs
                conn = sqlite3.connect(db_path)
                cursor = conn.execute("SELECT sku_code FROM sku_catalog")
                catalog_skus = set(row[0] for row in cursor.fetchall())
                conn.close()
                print(f"[UPLOAD-ORDERS] Loaded {len(catalog_skus)} SKUs from SQLite")
            else:
                # FALLBACK: Read from Excel (legacy)
                print("[UPLOAD-ORDERS] No SQLite DB, using Excel fallback for SKU validation")
                try:
                    import openpyxl
                    config = config_manager.load_config()
                    sequence_file = config.get('sequence_file', 'layouts/Warehouse_Logic.xlsx')
                    excel_path = os.path.join(PROJECT_ROOT, sequence_file)
                    
                    if os.path.exists(excel_path):
                        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
                        if 'PickingLocations' in wb.sheetnames:
                            sheet = wb['PickingLocations']
                            # Find sku_initial column
                            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
                            sku_col = None
                            for i, h in enumerate(headers):
                                if h and 'sku' in str(h).lower():
                                    sku_col = i
                                    break
                            
                            if sku_col is not None:
                                for row in sheet.iter_rows(min_row=2, values_only=True):
                                    if row and sku_col < len(row) and row[sku_col]:
                                        catalog_skus.add(str(row[sku_col]))
                        wb.close()
                        print(f"[UPLOAD-ORDERS] Loaded {len(catalog_skus)} SKUs from Excel")
                except Exception as excel_err:
                    print(f"[UPLOAD-ORDERS] Excel fallback failed: {excel_err}")
        except Exception as db_err:
            print(f"[UPLOAD-ORDERS] SKU catalog load error: {db_err}")
        
        # Validate orders against catalog
        valid_orders = 0
        valid_items = 0
        exclusions = []
        skus_found = set()
        skus_missing = set()
        
        for order in orders:
            order_has_valid_items = False
            
            for item in order['items']:
                sku_id = item.get('sku_id', '')
                
                if catalog_skus and sku_id not in catalog_skus:
                    skus_missing.add(sku_id)
                    exclusions.append({
                        'order_id': order['order_id'],
                        'item_sku': sku_id,
                        'reason': f"SKU '{sku_id}' no encontrado en catálogo",
                        'action': 'excluded'
                    })
                else:
                    skus_found.add(sku_id)
                    valid_items += 1
                    order_has_valid_items = True
            
            if order_has_valid_items or not catalog_skus:
                valid_orders += 1
        
        # Apply policy logic for summary
        if fulfillment_policy == 'fill_or_kill':
            orders_with_exclusions = set(e['order_id'] for e in exclusions)
            final_orders = valid_orders - len(orders_with_exclusions)
        else:
            final_orders = valid_orders
        
        return {
            "success": True,
            "message": f"Archivo '{file.filename}' procesado exitosamente",
            "file_path": saved_file_path,
            "summary": {
                "total_orders_input": len(orders),
                "total_items_input": total_items,
                "total_orders_output": final_orders,
                "total_items_output": valid_items,
                "skus_found": len(skus_found),
                "skus_missing": list(skus_missing),
                "fulfillment_policy": fulfillment_policy
            },
            "exclusions": exclusions[:20]  # Limit to first 20 for UI
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error processing order file: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


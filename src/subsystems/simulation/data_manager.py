# -*- coding: utf-8 -*-
"""
Data Manager Module - Central Data Loading for Warehouse Simulation
Digital Twin Warehouse Simulator

Responsibilities:
- Load warehouse data from SQLite (preferred) or Excel (fallback)
- Create and configure LayoutManager
- Validate data consistency between data source and TMX
- Expose processed data to simulation components

Author: Digital Twin Warehouse Team
Version: V12 - SQLite Dual-Mode Support
"""

import os
import sqlite3
from typing import List, Dict, Tuple, Optional, Any
from .layout_manager import LayoutManager


class DataManagerError(Exception):
    """Custom exception for DataManager errors"""
    pass


class DataManager:
    """
    Central data loader for warehouse simulation

    Loads data from SQLite database (preferred) or Excel file (fallback),
    validates consistency, and exposes processed data.

    This class acts as the single source of truth for all
    warehouse static data (picking points, staging areas, layout).
    
    V12 UPGRADE: Now supports dual-mode loading:
    - If warehouse.db exists -> Load from SQLite (fast, validated)
    - If not -> Fallback to Excel loading (legacy compatibility)
    """

    def __init__(self, tmx_file_path: str, excel_file_path: str,
                 configuracion: Optional[Dict[str, Any]] = None, headless: bool = False):
        """
        Initialize DataManager and load all warehouse data

        Args:
            tmx_file_path: Path to TMX map file (e.g., 'layouts/WH1.tmx')
            excel_file_path: Path to Warehouse_Logic.xlsx (used as fallback)
            configuracion: Optional configuration dict (for future extensions)

        Raises:
            DataManagerError: If data loading or validation fails
            FileNotFoundError: If required files don't exist
        """
        print(f"[DATA-MANAGER] Iniciando carga de TMX '{tmx_file_path}' "
              f"y datos de warehouse...")

        self.tmx_file_path = tmx_file_path
        self.configuracion = configuracion or {}

        # Resolve project root for database path
        project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))
        self.db_path = os.path.join(project_root, 'warehouse.db')

        # Resolve Excel path (for fallback)
        if not os.path.isabs(excel_file_path):
            excel_file_path = os.path.join(project_root, excel_file_path)
        self.excel_file_path = os.path.abspath(excel_file_path)

        # Create LayoutManager first (needed for validation)
        try:
            self.layout_manager = LayoutManager(tmx_file_path, headless=headless)
        except Exception as e:
            raise DataManagerError(f"Failed to create LayoutManager: {e}")

        # Initialize data structures
        self.puntos_de_picking_ordenados: List[Dict[str, Any]] = []
        self.outbound_staging_locations: Dict[int, Tuple[int, int]] = {}
        self.sku_catalog: Dict[str, Dict[str, Any]] = {}  # NEW: SKU catalog cache

        # Load data from SQLite (preferred) or Excel (fallback)
        if os.path.exists(self.db_path):
            print(f"[DATA-MANAGER] SQLite database found: {self.db_path}")
            self._load_from_database()
        else:
            print(f"[DATA-MANAGER] No SQLite database found, using Excel fallback")
            self._load_excel_data()

        # Validate consistency with TMX
        self._validate_data_consistency()

        print(f"[DATA-MANAGER] Plan maestro cargado y ordenado con "
              f"{len(self.puntos_de_picking_ordenados)} puntos.")

        if self.puntos_de_picking_ordenados:
            primer = self.puntos_de_picking_ordenados[0]
            ultimo = self.puntos_de_picking_ordenados[-1]
            print(f"  - Primer punto en secuencia: {primer}")
            print(f"  - Ultimo punto en secuencia: {ultimo}")

    # ========================================================================
    # PUBLIC INTERFACE - Data Access Methods
    # ========================================================================

    def get_picking_points(self) -> List[Dict[str, Any]]:
        """
        Return ordered list of picking points with all metadata

        Each point is a dict with keys:
        - x, y: Grid coordinates
        - ubicacion_grilla: Tuple (x, y)
        - pick_sequence: Order in master plan
        - equipment_required: Agent type (e.g., 'GroundOperator')
        - sku_initial: Initial SKU at this location
        - qty_initial: Initial quantity
        - WorkGroup: Work group identifier (e.g., 'WG_A')
        - WorkArea: Work area identifier (e.g., 'Area_Ground')
        - location_id: Logical location identifier (NEW in V12)

        Returns:
            List of dicts, sorted by pick_sequence
        """
        return self.puntos_de_picking_ordenados

    def get_outbound_staging_locations(self) -> Dict[int, Tuple[int, int]]:
        """
        Return dict mapping staging_id -> (x, y) grid coordinates

        Returns:
            Dict {staging_id: (x, y)} for all outbound staging areas
        """
        return self.outbound_staging_locations

    def get_layout_manager(self) -> LayoutManager:
        """
        Return configured LayoutManager instance

        Returns:
            LayoutManager with loaded TMX data
        """
        return self.layout_manager

    def get_pathfinder(self):
        """
        Return pathfinder instance (convenience method)

        Creates pathfinder from LayoutManager if not already created.

        Returns:
            Pathfinder instance configured for this layout
        """
        if not hasattr(self.layout_manager, 'pathfinder'):
            from .pathfinder import Pathfinder
            self.layout_manager.pathfinder = Pathfinder(
                self.layout_manager.collision_matrix,
                self.layout_manager.grid_width,
                self.layout_manager.grid_height
            )
        return self.layout_manager.pathfinder

    def get_all_skus(self) -> List[str]:
        """
        Return list of all known SKU codes.
        
        NEW in V12: Used for order validation.
        
        Returns:
            List of SKU code strings
        """
        return list(self.sku_catalog.keys())

    def get_sku_info(self, sku_code: str) -> Optional[Dict[str, Any]]:
        """
        Get information about a specific SKU.
        
        Args:
            sku_code: The SKU code to look up
            
        Returns:
            Dict with SKU info or None if not found
        """
        return self.sku_catalog.get(sku_code)

    def sku_exists(self, sku_code: str) -> bool:
        """
        Check if a SKU exists in the catalog.
        
        Args:
            sku_code: The SKU code to check
            
        Returns:
            True if SKU exists, False otherwise
        """
        return sku_code in self.sku_catalog

    # ========================================================================
    # PRIVATE METHODS - SQLite Data Loading (V12 NEW)
    # ========================================================================

    def _load_from_database(self):
        """
        Load warehouse data from SQLite database.
        
        Uses the v_picking_sequence view for efficient data retrieval.
        Falls back to direct table queries if view doesn't exist.
        """
        print(f"[DATA-MANAGER] Loading from SQLite: {self.db_path}")
        
        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            
            # Load SKU catalog
            self._load_sku_catalog(conn)
            
            # Load picking locations with inventory
            self._load_picking_locations_from_db(conn)
            
            # Load staging areas
            self._load_staging_areas_from_db(conn)
            
            conn.close()
            
            print(f"[DATA-MANAGER] SQLite loading complete:")
            print(f"  - SKUs loaded: {len(self.sku_catalog)}")
            print(f"  - Picking locations: {len(self.puntos_de_picking_ordenados)}")
            print(f"  - Staging areas: {len(self.outbound_staging_locations)}")
            
        except sqlite3.Error as e:
            conn.close() if 'conn' in locals() else None
            raise DataManagerError(f"SQLite error: {e}")

    def _load_sku_catalog(self, conn: sqlite3.Connection):
        """Load SKU catalog from database."""
        cursor = conn.execute("""
            SELECT sku_code, description, volume_m3, weight_kg, 
                   category, equipment_required
            FROM sku_catalog
        """)
        
        for row in cursor:
            self.sku_catalog[row['sku_code']] = {
                'sku_code': row['sku_code'],
                'description': row['description'],
                'volume_m3': row['volume_m3'] or 0.01,
                'weight_kg': row['weight_kg'],
                'category': row['category'] or 'GENERAL',
                'equipment_required': row['equipment_required'] or 'GroundOperator'
            }

    def _load_picking_locations_from_db(self, conn: sqlite3.Connection):
        """Load picking locations with inventory from database."""
        # Try using the view first (more efficient)
        try:
            cursor = conn.execute("""
                SELECT 
                    location_id,
                    pick_sequence,
                    work_area,
                    work_group,
                    equipment_required,
                    legacy_x,
                    legacy_y,
                    sku_code,
                    qty_available
                FROM v_picking_sequence
                ORDER BY pick_sequence
            """)
        except sqlite3.OperationalError:
            # View doesn't exist, use direct join
            cursor = conn.execute("""
                SELECT 
                    l.location_id,
                    l.pick_sequence,
                    l.work_area,
                    l.work_group,
                    l.equipment_required,
                    l.legacy_x,
                    l.legacy_y,
                    i.sku_code,
                    i.qty_available
                FROM locations l
                LEFT JOIN inventory i ON l.location_id = i.location_id
                WHERE l.location_type = 'PICKING'
                ORDER BY l.pick_sequence
            """)
        
        for row in cursor:
            x = row['legacy_x'] or 0
            y = row['legacy_y'] or 0
            
            punto = {
                'x': x,
                'y': y,
                'ubicacion_grilla': (x, y),
                'pick_sequence': row['pick_sequence'] or 0,
                'equipment_required': row['equipment_required'] or 'GroundOperator',
                'sku_initial': row['sku_code'] or 'SKU000',
                'qty_initial': row['qty_available'] or 1,
                'WorkGroup': row['work_group'] or 'WG_Default',
                'WorkArea': row['work_area'] or 'Area_Ground',
                'location_id': row['location_id']  # NEW: Include location_id
            }
            
            self.puntos_de_picking_ordenados.append(punto)

    def _load_staging_areas_from_db(self, conn: sqlite3.Connection):
        """Load staging areas from database."""
        cursor = conn.execute("""
            SELECT staging_id, legacy_x, legacy_y
            FROM staging_areas
            WHERE staging_type = 'OUTBOUND'
        """)
        
        for row in cursor:
            staging_id = row['staging_id']
            x = row['legacy_x'] or 0
            y = row['legacy_y'] or 0
            
            if staging_id > 0:
                self.outbound_staging_locations[staging_id] = (x, y)
        
        # Generate defaults if none found
        if not self.outbound_staging_locations:
            print("[DATA-MANAGER] No staging areas in DB, generating defaults")
            self._generate_default_staging_locations()

    # ========================================================================
    # PRIVATE METHODS - Excel Data Loading (Legacy Fallback)
    # ========================================================================

    def _load_excel_data(self):
        """
        Load and parse Warehouse_Logic.xlsx (LEGACY FALLBACK)

        Expects two sheets:
        - PickingLocations: x, y, pick_sequence, WorkArea, etc.
        - OutboundStaging: staging_id, x, y

        Raises:
            DataManagerError: If Excel loading fails
        """
        import openpyxl
        
        print(f"[DATA-MANAGER] Leyendo archivo Excel: {self.excel_file_path}")

        try:
            workbook = openpyxl.load_workbook(self.excel_file_path, data_only=True)
        except FileNotFoundError:
            raise DataManagerError(
                f"Excel file not found: {self.excel_file_path}"
            )
        except Exception as e:
            raise DataManagerError(f"Error loading Excel file: {e}")

        # Process PickingLocations sheet (REQUIRED)
        if 'PickingLocations' in workbook.sheetnames:
            sheet = workbook['PickingLocations']
            self._process_picking_locations(sheet)
        else:
            workbook.close()
            raise DataManagerError(
                "Sheet 'PickingLocations' not found in Excel file. "
                "This sheet is required for warehouse operation."
            )

        # Process OutboundStaging sheet (OPTIONAL with fallback)
        if 'OutboundStaging' in workbook.sheetnames:
            sheet = workbook['OutboundStaging']
            self._process_outbound_staging(sheet)
        else:
            print("[DATA-MANAGER WARNING] Sheet 'OutboundStaging' not found - "
                  "generating default staging locations")
            self._generate_default_staging_locations()

        workbook.close()
        
        # Build SKU catalog from Excel data
        self._build_sku_catalog_from_picking_points()

    def _build_sku_catalog_from_picking_points(self):
        """Build SKU catalog from loaded picking points (Excel fallback)."""
        for punto in self.puntos_de_picking_ordenados:
            sku_code = punto.get('sku_initial', '')
            if sku_code and sku_code not in self.sku_catalog:
                self.sku_catalog[sku_code] = {
                    'sku_code': sku_code,
                    'description': f'SKU {sku_code}',
                    'volume_m3': 0.01,
                    'equipment_required': punto.get('equipment_required', 'GroundOperator')
                }

    def _process_picking_locations(self, sheet):
        """
        Process PickingLocations sheet from Excel

        Expected columns:
        - x: Grid X coordinate
        - y: Grid Y coordinate
        - pick_sequence: Order in master plan
        - equipment_required: Agent type
        - sku_initial: Initial SKU code
        - qty_initial: Initial quantity
        - WorkGroup: Work group identifier
        - WorkArea: Work area identifier

        Args:
            sheet: openpyxl worksheet object

        Raises:
            DataManagerError: If required columns are missing
        """
        # Find header row (should contain 'x' column)
        header_row = None
        for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            if row and any(str(cell).lower() == 'x' for cell in row if cell):
                header_row = idx
                break

        if not header_row:
            raise DataManagerError(
                "Header row with 'x' column not found in PickingLocations sheet"
            )

        # Read header
        headers = [cell.value for cell in sheet[header_row]]

        # Process data rows
        data_rows = []
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not row[0]:  # Skip empty rows
                continue

            row_dict = {}
            for idx, header in enumerate(headers):
                if header and idx < len(row):
                    row_dict[header] = row[idx]

            data_rows.append(row_dict)

        print(f"[DATA-MANAGER] Hoja 'PickingLocations' cargada con "
              f"{len(data_rows)} registros")

        # Convert to standardized format
        for row_data in data_rows:
            try:
                pick_seq = int(row_data.get('pick_sequence', 0))
                
                punto = {
                    'x': int(row_data.get('x', 0)),
                    'y': int(row_data.get('y', 0)),
                    'pick_sequence': pick_seq,
                    'equipment_required': str(row_data.get('equipment_required', 'GroundOperator')),
                    'sku_initial': str(row_data.get('sku_initial', 'SKU000')),
                    'qty_initial': int(row_data.get('qty_initial', 1)),
                    'WorkGroup': str(row_data.get('WorkGroup', 'WG_Default')),
                    'WorkArea': str(row_data.get('WorkArea', 'Area_Ground')),
                    'location_id': f"LOC-{pick_seq:03d}"  # Generate location_id
                }

                # Add grid location tuple for convenience
                punto['ubicacion_grilla'] = (punto['x'], punto['y'])

                self.puntos_de_picking_ordenados.append(punto)

            except (ValueError, TypeError) as e:
                print(f"[DATA-MANAGER WARNING] Skipping invalid row: {row_data} - {e}")
                continue

        # Sort by pick_sequence
        self.puntos_de_picking_ordenados.sort(key=lambda p: p['pick_sequence'])

        print(f"[DATA-MANAGER] Columna 'WorkGroup' procesada correctamente")
        print(f"[DATA-MANAGER] Columna 'WorkArea' procesada correctamente")

    def _process_outbound_staging(self, sheet):
        """
        Process OutboundStaging sheet from Excel

        Expected columns:
        - staging_id: Unique identifier (integer)
        - x: Grid X coordinate
        - y: Grid Y coordinate

        Args:
            sheet: openpyxl worksheet object
        """
        # Find header row
        header_row = None
        for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            if row and any('staging' in str(cell).lower() for cell in row if cell):
                header_row = idx
                break

        if not header_row:
            # Fallback: assume first row is header
            header_row = 1

        headers = [cell.value for cell in sheet[header_row]]

        # Process staging locations
        staging_dict = {}
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or not row[0]:
                continue

            row_dict = {}
            for idx, header in enumerate(headers):
                if header and idx < len(row):
                    row_dict[header] = row[idx]

            try:
                # Handle different header variations
                staging_id_key = None
                for key in row_dict.keys():
                    if 'staging' in str(key).lower() and 'id' in str(key).lower():
                        staging_id_key = key
                        break

                if not staging_id_key:
                    staging_id_key = 'staging_id'

                staging_id = int(row_dict.get(staging_id_key, 0))
                x = int(row_dict.get('x', 0))
                y = int(row_dict.get('y', 0))

                if staging_id > 0:  # Only add valid IDs
                    staging_dict[staging_id] = (x, y)

            except (ValueError, TypeError, KeyError) as e:
                print(f"[DATA-MANAGER WARNING] Skipping invalid staging row: {row_dict} - {e}")
                continue

        self.outbound_staging_locations = staging_dict

        print(f"[DATA-MANAGER] Hoja 'OutboundStaging' cargada con "
              f"{len(staging_dict)} registros")
        print(f"[DATA-MANAGER] OutboundStaging procesado: {staging_dict}")

    def _generate_default_staging_locations(self):
        """
        Generate default staging locations if OutboundStaging sheet is missing

        Creates 7 staging areas along the bottom edge of the map
        """
        grid_h = self.layout_manager.grid_height

        # Create 7 staging areas along bottom edge (y = grid_h - 1)
        # Distributed evenly across the width
        num_staging = 7
        spacing = max(1, self.layout_manager.grid_width // (num_staging + 1))

        staging_dict = {}
        for i in range(1, num_staging + 1):
            x = i * spacing
            y = grid_h - 1
            staging_dict[i] = (x, y)

        self.outbound_staging_locations = staging_dict

        print(f"[DATA-MANAGER] Default staging locations generated: {staging_dict}")

    def _validate_data_consistency(self):
        """
        Validate that coordinates are within TMX grid bounds

        Checks:
        - All picking points are within grid
        - All staging locations are within grid
        - Coordinates are non-negative

        Raises:
            DataManagerError: If critical validation fails
        """
        grid_w = self.layout_manager.grid_width
        grid_h = self.layout_manager.grid_height

        # Validate picking points
        invalid_points = []
        for punto in self.puntos_de_picking_ordenados:
            x, y = punto['x'], punto['y']
            if not (0 <= x < grid_w and 0 <= y < grid_h):
                invalid_points.append((x, y, punto['pick_sequence']))

        if invalid_points:
            # Warning but not fatal - some points may be intentionally outside
            print(f"[DATA-MANAGER WARNING] {len(invalid_points)} puntos fuera "
                  f"del grid {grid_w}x{grid_h}")
            if len(invalid_points) <= 5:
                print(f"  Puntos invalidos: {invalid_points}")
            else:
                print(f"  Primeros 5 puntos invalidos: {invalid_points[:5]}")

        # Validate staging locations (CRITICAL - must be valid)
        invalid_staging = []
        for staging_id, (x, y) in self.outbound_staging_locations.items():
            if not (0 <= x < grid_w and 0 <= y < grid_h):
                invalid_staging.append((staging_id, x, y))

        if invalid_staging:
            raise DataManagerError(
                f"Staging locations out of bounds: {invalid_staging}. "
                f"Valid grid range: 0-{grid_w-1} x 0-{grid_h-1}"
            )

        print(f"[DATA-MANAGER] Validacion completada:")
        print(f"  - {len(self.puntos_de_picking_ordenados)} picking points validados")
        print(f"  - {len(self.outbound_staging_locations)} staging areas validadas")
        print(f"  - {len(self.sku_catalog)} SKUs en catalogo")
        print(f"  - Grid bounds: {grid_w}x{grid_h}")

    # ========================================================================
    # ALLOCATION LAYER - Stock Availability Methods (V12.1)
    # ========================================================================

    def get_available_stock(self, sku_code: str) -> int:
        """
        Get total available stock for a SKU across all locations.
        
        Used by allocation layer to check stock before creating WorkOrders.
        
        Args:
            sku_code: The SKU code to look up
            
        Returns:
            Total qty_available (summed across all locations), 0 if not found
        """
        if not os.path.exists(self.db_path):
            return 0
        
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.execute("""
                SELECT COALESCE(SUM(qty_available), 0) 
                FROM inventory 
                WHERE sku_code = ?
            """, (sku_code,))
            result = cursor.fetchone()[0]
            conn.close()
            return result
        except sqlite3.Error as e:
            print(f"[DATA-MANAGER] Error querying stock for {sku_code}: {e}")
            return 0

    def get_all_available_stock(self) -> Dict[str, int]:
        """
        Get available stock for ALL SKUs in a single efficient query.
        
        Used by DeterministicOrderStrategy for FCFS allocation.
        Returns snapshot of stock at query time.
        
        Returns:
            Dict mapping sku_code -> total_available quantity
        """
        if not os.path.exists(self.db_path):
            print("[DATA-MANAGER] No database found for stock query")
            return {}
        
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.execute("""
                SELECT sku_code, COALESCE(SUM(qty_available), 0) as total
                FROM inventory 
                GROUP BY sku_code
            """)
            result = {row[0]: row[1] for row in cursor.fetchall()}
            conn.close()
            
            total_units = sum(result.values())
            print(f"[DATA-MANAGER] Stock snapshot loaded: {len(result)} SKUs, {total_units} total units")
            return result
        except sqlite3.Error as e:
            print(f"[DATA-MANAGER] Error querying all stock: {e}")
            return {}

    def get_inventory_by_location(self) -> Dict[str, List[Dict[str, Any]]]:
        """
        Get per-LOCATION inventory snapshot, grouped by SKU (V12.1 - Initiative #1).

        Unlike get_all_available_stock (which aggregates by SKU), this returns the
        actual locations that physically hold each SKU, so the allocation layer can
        assign WorkOrders to REAL picking locations instead of random ones.

        Each location entry includes qty_free (= qty_available - qty_reserved) and
        the real grid coordinates / pick_sequence / work_area from the locations
        table. Locations are sorted by pick_sequence ASC (FCFS, reproducible).

        Returns:
            Dict mapping sku_code -> list of location dicts:
            {
                sku_code: [
                    {location_id, x, y, pick_sequence, work_area,
                     equipment_required, qty_available, qty_reserved, qty_free},
                    ...  # ordered by pick_sequence ASC
                ]
            }
        """
        result: Dict[str, List[Dict[str, Any]]] = {}

        if not os.path.exists(self.db_path):
            print("[DATA-MANAGER] No database found for per-location inventory query")
            return result

        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.execute("""
                SELECT
                    i.location_id,
                    i.sku_code,
                    i.qty_available,
                    i.qty_reserved,
                    l.pick_sequence,
                    l.work_area,
                    l.equipment_required,
                    l.legacy_x,
                    l.legacy_y
                FROM inventory i
                JOIN locations l ON i.location_id = l.location_id
                WHERE l.location_type = 'PICKING'
                ORDER BY i.sku_code, l.pick_sequence
            """)

            total_locs = 0
            for row in cursor:
                sku_code = row['sku_code']
                qty_available = row['qty_available'] or 0
                qty_reserved = row['qty_reserved'] or 0
                entry = {
                    'location_id': row['location_id'],
                    'x': row['legacy_x'] or 0,
                    'y': row['legacy_y'] or 0,
                    'pick_sequence': row['pick_sequence'] or 0,
                    'work_area': row['work_area'] or 'Area_Ground',
                    'equipment_required': row['equipment_required'] or 'GroundOperator',
                    'qty_available': qty_available,
                    'qty_reserved': qty_reserved,
                    'qty_free': max(0, qty_available - qty_reserved),
                }
                result.setdefault(sku_code, []).append(entry)
                total_locs += 1

            conn.close()
            print(f"[DATA-MANAGER] Per-location inventory loaded: "
                  f"{len(result)} SKUs across {total_locs} locations")
            return result
        except sqlite3.Error as e:
            conn.close() if 'conn' in locals() else None
            print(f"[DATA-MANAGER] Error querying inventory by location: {e}")
            return {}

    def commit_reservations(self, reservations: Dict[str, int]) -> bool:
        """
        Persist per-location reservations to inventory.qty_reserved (V12.1 - Init #1).

        Idempotent by design: first resets ALL qty_reserved to 0, then sets the
        absolute reserved quantity for each location in `reservations`. This way
        re-running the simulation does NOT accumulate stale reservations; each run
        leaves qty_reserved reflecting exactly the WorkOrders of that run.

        Both operations run in a single transaction.

        Args:
            reservations: Dict mapping location_id -> reserved quantity (absolute)

        Returns:
            True on success, False on error.
        """
        if not os.path.exists(self.db_path):
            print("[DATA-MANAGER] No database found - cannot commit reservations")
            return False

        try:
            conn = sqlite3.connect(self.db_path)
            try:
                conn.execute("BEGIN")
                conn.execute("UPDATE inventory SET qty_reserved = 0")
                for location_id, qty in reservations.items():
                    if qty and qty > 0:
                        conn.execute(
                            "UPDATE inventory SET qty_reserved = ? WHERE location_id = ?",
                            (int(qty), location_id)
                        )
                conn.commit()
            except sqlite3.Error:
                conn.rollback()
                raise
            finally:
                conn.close()

            total_reserved = sum(q for q in reservations.values() if q and q > 0)
            print(f"[DATA-MANAGER] Reservations committed: "
                  f"{len(reservations)} locations, {total_reserved} units reserved")
            return True
        except sqlite3.Error as e:
            print(f"[DATA-MANAGER] Error committing reservations: {e}")
            return False

    def reset_reservations(self) -> bool:
        """
        Reset ALL qty_reserved to 0 (V12.1 - Init #1).

        Standalone helper for callers that want to clear reservations without
        committing new ones (commit_reservations already resets internally).

        Returns:
            True on success, False on error.
        """
        if not os.path.exists(self.db_path):
            return False
        try:
            conn = sqlite3.connect(self.db_path)
            conn.execute("UPDATE inventory SET qty_reserved = 0")
            conn.commit()
            conn.close()
            print("[DATA-MANAGER] All reservations reset to 0")
            return True
        except sqlite3.Error as e:
            print(f"[DATA-MANAGER] Error resetting reservations: {e}")
            return False

    def consume_stock(self, location_id: str, qty: int,
                      sim_now: Optional[float] = None) -> Optional[Tuple[int, int]]:
        """
        Consume stock at a location when an item is physically picked (Fase 2).

        Decrements qty_available and releases the matching qty_reserved for the
        given location, guarding against negatives. Called once per WorkOrder at
        the pick moment during the simulation.

        Args:
            location_id: physical inventory location being picked.
            qty: units picked (= WorkOrder.cantidad_inicial).
            sim_now: optional SimPy timestamp to store in last_updated.

        Returns:
            (new_qty_available, new_qty_reserved) on success, None on error/no-op.
        """
        if not location_id or qty is None or qty <= 0:
            return None
        if not os.path.exists(self.db_path):
            return None

        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            row = conn.execute(
                "SELECT qty_available, qty_reserved FROM inventory WHERE location_id = ?",
                (location_id,)
            ).fetchone()
            if row is None:
                conn.close()
                print(f"[STOCK][WARN] location '{location_id}' not in inventory - skip consume")
                return None

            avail = row['qty_available'] or 0
            reserved = row['qty_reserved'] or 0

            if qty > avail:
                print(f"[STOCK][WARN] pick {qty} > available {avail} at {location_id} "
                      f"-> capping qty_available to 0")
            new_avail = max(0, avail - qty)
            new_reserved = max(0, reserved - qty)

            conn.execute(
                "UPDATE inventory SET qty_available = ?, qty_reserved = ?, last_updated = ? "
                "WHERE location_id = ?",
                (new_avail, new_reserved, sim_now, location_id)
            )
            conn.commit()
            conn.close()
            return (new_avail, new_reserved)
        except sqlite3.Error as e:
            conn.close() if 'conn' in locals() else None
            print(f"[STOCK][ERROR] consume_stock failed for {location_id}: {e}")
            return None

    def restore_inventory_baseline(self) -> bool:
        """
        Restore qty_available from a baseline snapshot and reset reservations (Fase 2).

        Makes each simulation run reproducible: the first time it runs it captures
        the CURRENT qty_available into the auxiliary table inventory_baseline
        (created with CREATE TABLE IF NOT EXISTS - the canonical schema is NOT
        touched). On every run it restores qty_available from that baseline and
        sets qty_reserved = 0, so the per-location allocation/ledger always starts
        from full stock and the in-sim decrement does not accumulate across runs.

        Returns:
            True on success, False on error.
        """
        if not os.path.exists(self.db_path):
            return False
        try:
            conn = sqlite3.connect(self.db_path)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS inventory_baseline (
                    location_id TEXT PRIMARY KEY,
                    qty_baseline INTEGER NOT NULL
                )
            """)
            # Lazy snapshot: populate baseline once from current qty_available
            count = conn.execute("SELECT COUNT(*) FROM inventory_baseline").fetchone()[0]
            if count == 0:
                conn.execute("""
                    INSERT INTO inventory_baseline (location_id, qty_baseline)
                    SELECT location_id, qty_available FROM inventory
                """)
                print("[STOCK] inventory_baseline created from current qty_available")

            # Restore qty_available from baseline; clear reservations
            conn.execute("""
                UPDATE inventory
                SET qty_available = (
                    SELECT b.qty_baseline FROM inventory_baseline b
                    WHERE b.location_id = inventory.location_id
                ),
                qty_reserved = 0
                WHERE location_id IN (SELECT location_id FROM inventory_baseline)
            """)
            conn.commit()
            total = conn.execute("SELECT SUM(qty_available) FROM inventory").fetchone()[0]
            conn.close()
            print(f"[STOCK] Inventory restored to baseline ({total} units), reservations reset")
            return True
        except sqlite3.Error as e:
            conn.close() if 'conn' in locals() else None
            print(f"[STOCK][ERROR] restore_inventory_baseline failed: {e}")
            return False

    def __repr__(self):
        return (f"DataManager(picking_points={len(self.puntos_de_picking_ordenados)}, "
                f"staging_areas={len(self.outbound_staging_locations)}, "
                f"skus={len(self.sku_catalog)}, "
                f"grid={self.layout_manager.grid_width}x{self.layout_manager.grid_height})")
